import base64
import io
import logging

_logger = logging.getLogger(__name__)

SUPPORTED_FORMATS = {
    "pdf": "application/pdf",
    "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "doc": "application/msword",
    "txt": "text/plain",
    "csv": "text/csv",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def get_file_extension(filename):
    """Extract and normalize file extension from filename."""
    if not filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def extract_text_from_binary(binary_data, filename):
    """Extract text from a base64-encoded binary file.

    Args:
        binary_data: Base64-encoded file content (as stored in Odoo Binary fields).
        filename: Original filename used to determine format.

    Returns:
        str: Extracted text content.

    Raises:
        ValueError: If the file format is unsupported or extraction fails.
    """
    if not binary_data:
        raise ValueError("Empty file content")

    file_bytes = base64.b64decode(binary_data)
    ext = get_file_extension(filename)

    if ext == "pdf":
        return _extract_pdf(file_bytes)
    elif ext in ("docx",):
        return _extract_docx(file_bytes)
    elif ext in ("txt", "csv", "md"):
        return _extract_text(file_bytes)
    elif ext == "xlsx":
        return _extract_xlsx(file_bytes)
    else:
        raise ValueError(
            f"Unsupported file format: .{ext}. "
            f"Supported formats: {', '.join(SUPPORTED_FORMATS.keys())}"
        )


def extract_text_from_attachments(attachments):
    """Extract text from multiple ir.attachment records.

    Args:
        attachments: Odoo recordset of ir.attachment records.

    Returns:
        str: Combined text from all documents, separated by headers.
    """
    parts = []
    for attachment in attachments:
        try:
            text = extract_text_from_binary(attachment.datas, attachment.name)
            parts.append(f"--- Document: {attachment.name} ---\n{text}")
            _logger.info(
                "Extracted text from %s (%d chars)", attachment.name, len(text)
            )
        except ValueError as e:
            _logger.warning("Skipping %s: %s", attachment.name, e)
            parts.append(
                f"--- Document: {attachment.name} ---\n[Could not extract text: {e}]"
            )
        except Exception as e:
            _logger.error("Failed to extract text from %s: %s", attachment.name, e)
            parts.append(
                f"--- Document: {attachment.name} ---\n[Extraction error: {e}]"
            )
    return "\n\n".join(parts)


def _extract_pdf(file_bytes):
    """Extract text from PDF using PyMuPDF (fitz)."""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ValueError("PyMuPDF is not installed. Run: pip install PyMuPDF")

    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages = []
    for page_num, page in enumerate(doc, 1):
        text = page.get_text("text")
        if text.strip():
            pages.append(f"[Page {page_num}]\n{text.strip()}")
    doc.close()

    if not pages:
        raise ValueError(
            "PDF contains no extractable text (might be scanned/image-based)"
        )
    return "\n\n".join(pages)


def _extract_docx(file_bytes):
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
    except ImportError:
        raise ValueError("python-docx is not installed. Run: pip install python-docx")

    doc = Document(io.BytesIO(file_bytes))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

    if not paragraphs:
        raise ValueError("DOCX contains no text content")
    return "\n".join(paragraphs)


def _extract_text(file_bytes):
    """Extract text from plain text files."""
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            return file_bytes.decode(encoding)
        except (UnicodeDecodeError, AttributeError):
            continue
    raise ValueError("Could not decode text file with any supported encoding")


def _extract_xlsx(file_bytes):
    """Extract text from XLSX using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    sheets_text = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                rows.append(row_text)
        if rows:
            sheets_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    wb.close()
    if not sheets_text:
        raise ValueError("XLSX contains no data")
    return "\n\n".join(sheets_text)
