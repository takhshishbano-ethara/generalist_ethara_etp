import base64
import csv
import io
import json
import logging

from odoo import http
from odoo.http import request

from odoo.addons.api_auth_gateway.controllers.utility import (
    return_Response,
    validate_token,
    validate_request,
)

_logger = logging.getLogger(__name__)

LEVIATHAN_ADMIN_ROLE_XMLIDS = (
    "api_auth_gateway.role_cto_technical",
    "api_auth_gateway.role_tpm_technical",
    "api_auth_gateway.role_pl_technical",
    "api_auth_gateway.role_pl_stem",
    "api_auth_gateway.role_pl_non_stem",
)

LEVIATHAN_TASKER_ROLE_XMLIDS = (
    "api_auth_gateway.role_tasker_technical",
    "api_auth_gateway.role_tasker_stem",
    "api_auth_gateway.role_tasker_non_stem",
)

LEVIATHAN_USER_ROLE_XMLIDS = LEVIATHAN_ADMIN_ROLE_XMLIDS + (
    "api_auth_gateway.role_qc_technical",
    "api_auth_gateway.role_qc_stem",
    "api_auth_gateway.role_qc_non_stem",
) + LEVIATHAN_TASKER_ROLE_XMLIDS

SKIP_DUPLICATE_STATES = ("submitted", "cancelled")
BULK_REQUIRED_COLUMNS = ("url",)
BULK_OPTIONAL_COLUMNS = ("category_id", "category", "tasker", "tasker_id")
SUPPORTED_BULK_EXTENSIONS = (".csv", ".xlsx")

LIST_DEFAULT_LIMIT = 50
LIST_MAX_LIMIT = 500
JOB_LIST_ALLOWED_ORDERS = (
    "create_date desc", "create_date asc",
    "write_date desc", "write_date asc",
    "name asc", "name desc",
    "score desc", "score asc",
    "state asc", "state desc",
    "completed_at desc", "completed_at asc",
)


def _get_role_ids(env, xmlids):
    ids = []
    for xmlid in xmlids:
        rec = env.ref(xmlid, raise_if_not_found=False)
        if rec:
            ids.append(rec.id)
    return ids


def _user_has_role(env, xmlids):
    role = env.user.user_role
    if not role:
        return False
    return role.id in _get_role_ids(env, xmlids)


def _require_admin():
    if not _user_has_role(request.env, LEVIATHAN_ADMIN_ROLE_XMLIDS):
        return return_Response(
            message="You are not allowed to create Leviathan jobs.",
            status=403,
        )
    return None


def _require_leviathan_user():
    if not _user_has_role(request.env, LEVIATHAN_USER_ROLE_XMLIDS):
        return return_Response(
            message="You are not allowed to access Leviathan data.",
            status=403,
        )
    return None


def _require_tasker():
    if not _user_has_role(request.env, LEVIATHAN_TASKER_ROLE_XMLIDS):
        return return_Response(
            message="Only taskers are allowed to claim Leviathan jobs.",
            status=403,
        )
    return None


def _normalize_url(raw):
    if not raw:
        return ""
    url = str(raw).strip()
    if not url:
        return ""
    if not (url.startswith("http://") or url.startswith("https://")):
        url = "https://" + url
    return url


def _resolve_category(env, category_id=None, category_name=None):
    Category = env["leviathan.category"].sudo()
    if category_id:
        try:
            cid = int(category_id)
        except (TypeError, ValueError):
            return False, f"Invalid category_id '{category_id}'."
        cat = Category.browse(cid).exists()
        if not cat:
            return False, f"Category id {cid} not found."
        return cat, None
    if category_name:
        name = str(category_name).strip()
        if not name:
            return False, None
        cat = Category.search([("name", "=ilike", name)], limit=1)
        if not cat:
            return False, f"Category '{name}' not found."
        return cat, None
    return False, None


def _resolve_tasker(env, tasker_value):
    if tasker_value in (None, "", False):
        return False, None
    Users = env["res.users"].sudo()
    if isinstance(tasker_value, int):
        user = Users.browse(tasker_value).exists()
        if not user:
            return False, f"Tasker id {tasker_value} not found."
        return user, None
    raw = str(tasker_value).strip()
    if not raw:
        return False, None
    if raw.isdigit():
        user = Users.browse(int(raw)).exists()
        if user:
            return user, None
    user = Users.search([("login", "=", raw)], limit=1)
    if not user and "@" in raw:
        user = Users.search([("email", "=", raw)], limit=1)
    if not user:
        return False, f"Tasker '{raw}' not found."
    return user, None


def _job_already_exists(env, url):
    return bool(
        env["leviathan.job"].sudo().search(
            [
                ("url", "=", url),
                ("state", "not in", list(SKIP_DUPLICATE_STATES)),
            ],
            limit=1,
        )
    )


def _serialize_job(job):
    return {
        "id": job.id,
        "name": job.name,
        "url": job.url,
        "state": job.state,
        "category_id": job.category_id.id or False,
        "category_name": job.category_id.name or "",
        "user_id": job.user_id.id or False,
        "user_name": job.user_id.name or "",
    }


def _serialize_category(cat):
    return {
        "id": cat.id,
        "name": cat.name,
        "technical_key": cat.technical_key,
        "active": cat.active,
    }


def _coerce_int(value, default):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


CATEGORY_LIST_DEFAULT_LIMIT = 100
CATEGORY_LIST_MAX_LIMIT = 500


def _parse_csv(file_bytes):
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({(k or "").strip().lower(): (v if v is not None else "") for k, v in row.items()})
    return rows


def _parse_xlsx(file_bytes):
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl is required to parse XLSX files but is not installed.")
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [(str(h).strip().lower() if h is not None else "") for h in header]
    rows = []
    for raw_row in rows_iter:
        if raw_row is None:
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw_row):
            continue
        row = {}
        for idx, value in enumerate(raw_row):
            if idx >= len(headers):
                break
            key = headers[idx]
            if not key:
                continue
            row[key] = "" if value is None else value
        rows.append(row)
    return rows


def _build_job_vals(env, url, category_id=None, category_name=None, tasker_value=None):
    warnings = []
    vals = {"url": url}

    category, cat_warn = _resolve_category(env, category_id=category_id, category_name=category_name)
    if cat_warn:
        warnings.append(cat_warn)
    if category:
        vals["category_id"] = category.id

    user, user_warn = _resolve_tasker(env, tasker_value)
    if user_warn:
        warnings.append(user_warn)
    if user:
        vals["user_id"] = user.id

    return vals, warnings


def _resolve_role_specs(env, specs):
    Role = env["api.role"].sudo()
    ids = []
    warnings = []
    for spec in specs or []:
        if spec is None:
            continue
        raw = str(spec).strip()
        if not raw:
            continue
        role = False
        if raw.isdigit():
            role = Role.browse(int(raw)).exists()
        if not role:
            role = Role.search([("name", "=ilike", raw)], limit=1)
        if not role and "-" in raw:
            role = Role.search([("user_type", "=ilike", raw)], limit=1)
        if role:
            if role.id not in ids:
                ids.append(role.id)
        else:
            warnings.append(f"Role '{raw}' not found.")
    return ids, warnings


def _split_csv_param(value):
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return [str(v).strip() for v in value if str(v).strip()]
    return [p.strip() for p in str(value).split(",") if p.strip()]


def _try_get_xmlid(record):
    try:
        data = record.get_external_id()
        xmlid = data.get(record.id) if isinstance(data, dict) else ""
        return xmlid or ""
    except Exception:
        return ""


def _serialize_role(role):
    return {
        "id": role.id,
        "name": role.name or "",
        "user_type": role.user_type or "",
        "project_type": role.project_type or "",
        "xmlid": _try_get_xmlid(role),
        "permission_count": len(role.line_ids),
    }


def _serialize_user(user):
    role = user.user_role
    role_info = _serialize_role(role) if role else None
    return {
        "id": user.id,
        "name": user.name or "",
        "login": user.login or "",
        "email": user.email or (user.partner_id.email or ""),
        "active": user.active,
        "partner_id": user.partner_id.id or False,
        "company_id": user.company_id.id or False,
        "company_name": user.company_id.name or "",
        "user_role_id": role.id if role else False,
        "user_role_name": role.name if role else "",
        "user_type": role.user_type if role else "",
        "project_type": role.project_type if role else "",
        "user_role": role_info,
    }


def _iso(dt):
    return dt.isoformat() if dt else ""


def _serialize_job_summary(job):
    return {
        "id": job.id,
        "name": job.name,
        "url": job.url or "",
        "site_name": job.site_name or "",
        "state": job.state,
        "category_id": job.category_id.id or False,
        "category_name": job.category_id.name or "",
        "category_key": job.category_key or "",
        "user_id": job.user_id.id or False,
        "user_name": job.user_id.name or "",
        "user_login": job.user_id.login or "",
        "score": job.score or 0.0,
        "grade": job.grade or "",
        "qc_verdict": job.qc_verdict or "",
        "create_date": _iso(job.create_date),
        "write_date": _iso(job.write_date),
        "started_at": _iso(job.started_at),
        "completed_at": _iso(job.completed_at),
        "duration_seconds": job.duration_seconds or 0.0,
    }


def _serialize_job_detail(job):
    data = _serialize_job_summary(job)
    data.update({
        "page_count": job.page_count or 0,
        "prd_text": job.prd_text or "",
        "prd_prompt": job.prd_prompt or "",
        "qc_report": job.qc_report or "",
        "tech_stack": job.tech_stack or "",
        "error_message": job.error_message or "",
        "extraction_warnings": job.extraction_warnings or "",
        "prd_url": job.prd_url or "",
        "artifacts_url": job.artifacts_url or "",
        "deliverables_url": job.deliverables_url or "",
        "screenshot_keys": job.screenshot_keys or [],
        "asset_keys": job.asset_keys or [],
        "score_report": job.score_report_json or None,
        "site_discovery": job.site_discovery_json or None,
        "lambda_callback": job.lambda_callback_json or None,
        "llm_trace": job.llm_trace_json or None,
        "llm_attempts": job.llm_attempts or 0,
        "last_heartbeat": _iso(job.last_heartbeat),
        "started_processing_at": _iso(job.started_processing_at),
        "watchdog_retry_count": job.watchdog_retry_count or 0,
        "cancel_requested": bool(job.cancel_requested),
        "via_batch": bool(job.via_batch),
    })
    return data


def _build_jobs_domain(env, params):
    domain = []
    warnings = []

    states = _split_csv_param(params.get("state"))
    if states:
        domain.append(("state", "in", states))

    name = (params.get("name") or "").strip()
    if name:
        domain.append(("name", "ilike", name))

    url_q = (params.get("url") or "").strip()
    if url_q:
        domain.append(("url", "ilike", url_q))

    search = (params.get("search") or "").strip()
    if search:
        domain += [
            "|", "|",
            ("name", "ilike", search),
            ("url", "ilike", search),
            ("site_name", "ilike", search),
        ]

    user_id_raw = (params.get("user_id") or "").strip()
    if user_id_raw:
        if user_id_raw.lower() in ("none", "false", "null", "unassigned"):
            domain.append(("user_id", "=", False))
        elif user_id_raw.isdigit():
            domain.append(("user_id", "=", int(user_id_raw)))
        else:
            warnings.append(f"Invalid user_id '{user_id_raw}'.")

    has_user = params.get("has_user")
    if has_user not in (None, ""):
        val = str(has_user).strip().lower()
        if val in ("true", "1", "yes", "assigned"):
            domain.append(("user_id", "!=", False))
        elif val in ("false", "0", "no", "unassigned"):
            domain.append(("user_id", "=", False))
        else:
            warnings.append(f"Invalid has_user '{has_user}'.")

    cat_id_raw = (params.get("category_id") or "").strip()
    if cat_id_raw:
        if cat_id_raw.isdigit():
            domain.append(("category_id", "=", int(cat_id_raw)))
        else:
            warnings.append(f"Invalid category_id '{cat_id_raw}'.")

    category_name = (params.get("category") or "").strip()
    if category_name:
        domain.append(("category_id.name", "=ilike", category_name))

    qc = (params.get("qc_verdict") or "").strip()
    if qc:
        domain.append(("qc_verdict", "=", qc))

    grade = (params.get("grade") or "").strip()
    if grade:
        domain.append(("grade", "=", grade))

    smin = params.get("score_min")
    if smin not in (None, ""):
        try:
            domain.append(("score", ">=", float(smin)))
        except (TypeError, ValueError):
            warnings.append(f"Invalid score_min '{smin}'.")

    smax = params.get("score_max")
    if smax not in (None, ""):
        try:
            domain.append(("score", "<=", float(smax)))
        except (TypeError, ValueError):
            warnings.append(f"Invalid score_max '{smax}'.")

    date_from = (params.get("date_from") or "").strip()
    if date_from:
        domain.append(("create_date", ">=", date_from))

    date_to = (params.get("date_to") or "").strip()
    if date_to:
        domain.append(("create_date", "<=", f"{date_to} 23:59:59"))

    return domain, warnings


class LeviathanExtensionController(http.Controller):

    @http.route(
        "/api/v1/leviathan_ext/jobs/create",
        type="http",
        auth="none",
        methods=["POST"],
        csrf=False,
        cors="*",
    )
    @validate_token
    @validate_request(
        {
            "url": {"type": "string", "required": True},
            "category_id": {"type": "int", "required": True},
            "tasker_id": {"type": "int", "required": False},
        }
    )
    def leviathan_ext_create_job(self, **kwargs):
        guard = _require_admin()
        if guard is not None:
            return guard

        jdata = kwargs.get("jdata") or {}
        url = _normalize_url(jdata.get("url"))
        if not url:
            return return_Response(message="'url' is required and cannot be empty.", status=400)

        category, cat_warn = _resolve_category(request.env, category_id=jdata.get("category_id"))
        if not category:
            return return_Response(
                message=cat_warn or "category_id is required.",
                status=400,
            )

        tasker_id = jdata.get("tasker_id")
        user = False
        if tasker_id:
            user, user_warn = _resolve_tasker(request.env, tasker_id)
            if not user:
                return return_Response(message=user_warn or "Tasker not found.", status=400)

        if _job_already_exists(request.env, url):
            return return_Response(
                message=f"An active job already exists for url '{url}'.",
                status=400,
            )

        vals = {"url": url, "category_id": category.id}
        if user:
            vals["user_id"] = user.id

        try:
            job = request.env["leviathan.job"].sudo().create(vals)
        except Exception as exc:
            _logger.exception("leviathan_ext create job failed: %s", exc)
            return return_Response(message=f"Failed to create job: {exc}", status=400)

        return return_Response(
            message="Leviathan job created successfully.",
            status=200,
            data={"job": _serialize_job(job)},
        )

    @http.route(
        "/api/v1/leviathan_ext/jobs/bulk_create",
        type="http",
        auth="none",
        methods=["POST"],
        csrf=False,
        cors="*",
    )
    @validate_token
    @validate_request(
        {
            "file": {"type": "string", "required": True},
            "filename": {"type": "string", "required": True},
        }
    )
    def leviathan_ext_bulk_create_jobs(self, **kwargs):
        guard = _require_admin()
        if guard is not None:
            return guard

        jdata = kwargs.get("jdata") or {}
        b64_payload = jdata.get("file") or ""
        filename = (jdata.get("filename") or "").strip().lower()

        if not filename.endswith(SUPPORTED_BULK_EXTENSIONS):
            return return_Response(
                message="Unsupported file type. Only .csv and .xlsx are supported.",
                status=400,
            )

        try:
            if "," in b64_payload and b64_payload.lstrip().startswith("data:"):
                b64_payload = b64_payload.split(",", 1)[1]
            file_bytes = base64.b64decode(b64_payload, validate=False)
        except Exception as exc:
            return return_Response(message=f"Invalid base64 file payload: {exc}", status=400)

        try:
            if filename.endswith(".csv"):
                rows = _parse_csv(file_bytes)
            else:
                rows = _parse_xlsx(file_bytes)
        except RuntimeError as exc:
            return return_Response(message=str(exc), status=400)
        except Exception as exc:
            _logger.exception("leviathan_ext bulk parse failed: %s", exc)
            return return_Response(message=f"Failed to parse file: {exc}", status=400)

        if not rows:
            return return_Response(message="No data rows found in file.", status=400)

        header_keys = set(rows[0].keys())
        missing_required = [c for c in BULK_REQUIRED_COLUMNS if c not in header_keys]
        if missing_required:
            return return_Response(
                message=(
                    f"Missing required column(s): {', '.join(missing_required)}. "
                    f"Required: {', '.join(BULK_REQUIRED_COLUMNS)}. "
                    f"Optional: {', '.join(BULK_OPTIONAL_COLUMNS)}."
                ),
                status=400,
            )

        created_jobs = []
        skipped = []
        errors = []
        Job = request.env["leviathan.job"].sudo()

        HEADER_ROW_OFFSET = 2
        for idx, row in enumerate(rows, start=HEADER_ROW_OFFSET):
            url = _normalize_url(row.get("url"))
            if not url:
                skipped.append({"row": idx, "reason": "empty url"})
                continue

            if _job_already_exists(request.env, url):
                skipped.append({"row": idx, "url": url, "reason": "duplicate active job"})
                continue

            vals, warnings = _build_job_vals(
                request.env,
                url=url,
                category_id=row.get("category_id"),
                category_name=row.get("category"),
                tasker_value=row.get("tasker") or row.get("tasker_id"),
            )

            try:
                job = Job.create(vals)
            except Exception as exc:
                errors.append({"row": idx, "url": url, "error": str(exc)})
                continue

            entry = _serialize_job(job)
            if warnings:
                entry["warnings"] = warnings
            created_jobs.append(entry)

        return return_Response(
            message=(
                f"Bulk create complete. created={len(created_jobs)} "
                f"skipped={len(skipped)} errors={len(errors)}"
            ),
            status=200,
            data={
                "created": len(created_jobs),
                "skipped_count": len(skipped),
                "error_count": len(errors),
                "jobs": created_jobs,
                "skipped": skipped,
                "row_errors": errors,
            },
        )

    @http.route(
        "/api/v1/leviathan_ext/categories",
        type="http",
        auth="none",
        methods=["GET"],
        csrf=False,
        cors="*",
    )
    @validate_token
    def leviathan_ext_list_categories(self, **kwargs):
        params = request.params or {}
        active_raw = str(params.get("active", "true")).strip().lower()
        active_filter = active_raw not in ("false", "0", "no")
        search = (params.get("search") or "").strip()
        limit = _coerce_int(params.get("limit"), CATEGORY_LIST_DEFAULT_LIMIT)
        limit = max(1, min(limit, CATEGORY_LIST_MAX_LIMIT))
        offset = max(0, _coerce_int(params.get("offset"), 0))

        domain = []
        if search:
            domain.append(("name", "ilike", search))

        Category = request.env["leviathan.category"].sudo()
        if not active_filter:
            Category = Category.with_context(active_test=False)

        total = Category.search_count(domain)
        records = Category.search(domain, limit=limit, offset=offset)

        return return_Response(
            message="OK",
            status=200,
            data={
                "categories": [_serialize_category(c) for c in records],
                "count": len(records),
                "total": total,
                "offset": offset,
                "limit": limit,
            },
        )

    @http.route(
        "/api/v1/leviathan_ext/users",
        type="http",
        auth="none",
        methods=["GET"],
        csrf=False,
        cors="*",
    )
    @validate_token
    def leviathan_ext_list_users(self, **kwargs):
        guard = _require_leviathan_user()
        if guard is not None:
            return guard

        params = request.params or {}
        role_specs = _split_csv_param(params.get("role")) + _split_csv_param(params.get("roles"))
        active_raw = str(params.get("active", "true")).strip().lower()
        active_filter = active_raw not in ("false", "0", "no")
        search = (params.get("search") or "").strip()
        has_role_raw = params.get("has_role")
        limit = _coerce_int(params.get("limit"), LIST_DEFAULT_LIMIT)
        limit = max(1, min(limit, LIST_MAX_LIMIT))
        offset = max(0, _coerce_int(params.get("offset"), 0))

        warnings = []
        domain = []
        if not active_filter:
            domain.append(("active", "in", (True, False)))

        if search:
            domain += [
                "|", "|",
                ("name", "ilike", search),
                ("login", "ilike", search),
                ("email", "ilike", search),
            ]

        role_ids = []
        if role_specs:
            role_ids, role_warnings = _resolve_role_specs(request.env, role_specs)
            warnings.extend(role_warnings)
            if not role_ids:
                return return_Response(
                    message="No matching roles found.",
                    status=400,
                    errors=warnings,
                )
            domain.append(("user_role", "in", role_ids))

        if has_role_raw not in (None, ""):
            val = str(has_role_raw).strip().lower()
            if val in ("true", "1", "yes"):
                domain.append(("user_role", "!=", False))
            elif val in ("false", "0", "no"):
                domain.append(("user_role", "=", False))
            else:
                warnings.append(f"Invalid has_role '{has_role_raw}'.")

        Users = request.env["res.users"].sudo()
        if not active_filter:
            Users = Users.with_context(active_test=False)

        total = Users.search_count(domain)
        records = Users.search(domain, limit=limit, offset=offset, order="name asc, id asc")

        Role = request.env["api.role"].sudo()
        roles_resolved = []
        if role_ids:
            roles_resolved = [
                _serialize_role(r)
                for r in Role.browse(role_ids).exists()
            ]
        available_roles = [_serialize_role(r) for r in Role.search([], order="name asc")]

        return return_Response(
            message="OK",
            status=200,
            errors=warnings,
            data={
                "users": [_serialize_user(u) for u in records],
                "count": len(records),
                "total": total,
                "offset": offset,
                "limit": limit,
                "roles_resolved": roles_resolved,
                "available_roles": available_roles,
            },
        )

    @http.route(
        "/api/v1/leviathan_ext/jobs",
        type="http",
        auth="none",
        methods=["GET"],
        csrf=False,
        cors="*",
    )
    @validate_token
    def leviathan_ext_list_jobs(self, **kwargs):
        guard = _require_leviathan_user()
        if guard is not None:
            return guard

        params = request.params or {}
        limit = _coerce_int(params.get("limit"), LIST_DEFAULT_LIMIT)
        limit = max(1, min(limit, LIST_MAX_LIMIT))
        offset = max(0, _coerce_int(params.get("offset"), 0))
        order = (params.get("order") or "create_date desc").strip().lower()
        if order not in JOB_LIST_ALLOWED_ORDERS:
            return return_Response(
                message=(
                    f"Invalid order '{order}'. Allowed: "
                    f"{', '.join(JOB_LIST_ALLOWED_ORDERS)}."
                ),
                status=400,
            )

        domain, warnings = _build_jobs_domain(request.env, params)

        Job = request.env["leviathan.job"].sudo()
        total = Job.search_count(domain)
        records = Job.search(domain, limit=limit, offset=offset, order=order)

        return return_Response(
            message="OK",
            status=200,
            errors=warnings,
            data={
                "jobs": [_serialize_job_summary(j) for j in records],
                "count": len(records),
                "total": total,
                "offset": offset,
                "limit": limit,
                "order": order,
            },
        )

    @http.route(
        "/api/v1/leviathan_ext/jobs/<int:job_id>",
        type="http",
        auth="none",
        methods=["GET"],
        csrf=False,
        cors="*",
    )
    @validate_token
    def leviathan_ext_get_job(self, job_id, **kwargs):
        guard = _require_leviathan_user()
        if guard is not None:
            return guard

        job = request.env["leviathan.job"].sudo().browse(job_id).exists()
        if not job:
            return return_Response(
                message=f"Leviathan job id {job_id} not found.",
                status=404,
            )

        return return_Response(
            message="OK",
            status=200,
            data={"job": _serialize_job_detail(job)},
        )

    @http.route(
        "/api/v1/leviathan_ext/jobs/<int:job_id>/claim",
        type="http",
        auth="none",
        methods=["POST"],
        csrf=False,
        cors="*",
    )
    @validate_token
    def leviathan_ext_claim_job(self, job_id, **kwargs):
        guard = _require_tasker()
        if guard is not None:
            return guard

        job = request.env["leviathan.job"].sudo().browse(job_id).exists()
        if not job:
            return return_Response(
                message=f"Leviathan job id {job_id} not found.",
                status=404,
            )

        if job.state != "not_assigned":
            return return_Response(
                message=(
                    f"Job '{job.name}' cannot be claimed because its state is "
                    f"'{job.state}'. Only 'not_assigned' jobs can be claimed."
                ),
                status=409,
            )

        tasker = request.env.user
        try:
            job.write({"user_id": tasker.id})
        except Exception as exc:
            _logger.exception("[leviathan_ext] claim job %s failed", job_id)
            return return_Response(
                message=f"Failed to assign job: {exc}",
                status=400,
            )

        warnings = []
        try:
            job.sudo().with_context(force_extract=True).action_run()
        except Exception as exc:
            _logger.exception(
                "[leviathan_ext] action_run failed for job %s after claim", job_id
            )
            warnings.append(f"Job claimed but action_run failed: {exc}")

        job = request.env["leviathan.job"].sudo().browse(job_id).exists()
        return return_Response(
            message="OK",
            status=200,
            errors=warnings,
            data={"job": _serialize_job_detail(job)},
        )

    @http.route(
        "/api/v1/leviathan_ext/jobs/<int:job_id>/update",
        type="http",
        auth="none",
        methods=["POST"],
        csrf=False,
        cors="*",
    )
    @validate_token
    def leviathan_ext_update_job(self, job_id, **kwargs):
        guard = _require_admin()
        if guard is not None:
            return guard

        raw_body = request.httprequest.data or b""
        if isinstance(raw_body, bytes):
            try:
                raw_body = raw_body.decode("utf-8")
            except UnicodeDecodeError:
                return return_Response(
                    message="Request body must be valid UTF-8 JSON.",
                    status=400,
                )
        body = (raw_body or "").strip()
        if not body:
            return return_Response(
                message="Request body is empty; provide a JSON object.",
                status=400,
            )
        try:
            payload = json.loads(body)
        except (ValueError, TypeError):
            return return_Response(
                message="Request body must be valid JSON.",
                status=400,
            )
        if not isinstance(payload, dict):
            return return_Response(
                message="Request body must be a JSON object.",
                status=400,
            )

        job = request.env["leviathan.job"].sudo().browse(job_id).exists()
        if not job:
            return return_Response(
                message=f"Leviathan job id {job_id} not found.",
                status=404,
            )

        env = request.env
        warnings = []
        vals = {}

        if "url" in payload:
            url_val = payload.get("url")
            if not url_val or not str(url_val).strip():
                return return_Response(
                    message="Field 'url' cannot be empty.",
                    status=400,
                )
            vals["url"] = _normalize_url(str(url_val).strip())

        if "site_name" in payload:
            site_name_val = payload.get("site_name")
            vals["site_name"] = (
                str(site_name_val).strip() if site_name_val else False
            )

        if "category_id" in payload:
            category_val = payload.get("category_id")
            if category_val in (None, False, "", 0, "0"):
                vals["category_id"] = False
            else:
                try:
                    category_id_int = int(category_val)
                except (TypeError, ValueError):
                    return return_Response(
                        message="Field 'category_id' must be an integer.",
                        status=400,
                    )
                category = _resolve_category(env, category_id=category_id_int)
                if not category:
                    return return_Response(
                        message=f"Category id {category_id_int} not found.",
                        status=400,
                    )
                vals["category_id"] = category.id

        tasker_assigned = False
        if "tasker_id" in payload:
            tasker_val = payload.get("tasker_id")
            if tasker_val in (None, False, "", 0, "0"):
                vals["user_id"] = False
            else:
                tasker = _resolve_tasker(env, tasker_val)
                if not tasker:
                    return return_Response(
                        message=f"Tasker '{tasker_val}' not found.",
                        status=400,
                    )
                vals["user_id"] = tasker.id
                if tasker.id != (job.user_id.id or False):
                    tasker_assigned = True

        if not vals:
            return return_Response(
                message=(
                    "No updatable fields provided. Supported fields: "
                    "url, site_name, category_id, tasker_id."
                ),
                status=400,
            )

        try:
            job.write(vals)
        except Exception as exc:
            _logger.exception("[leviathan_ext] update job %s failed", job_id)
            return return_Response(
                message=f"Failed to update job: {exc}",
                status=400,
            )

        job = request.env["leviathan.job"].sudo().browse(job_id).exists()

        if tasker_assigned and job.state in ("draft", "not_assigned"):
            try:
                job.sudo().with_context(force_extract=True).action_run()
            except Exception as exc:
                _logger.exception(
                    "[leviathan_ext] action_run failed for job %s after update",
                    job_id,
                )
                warnings.append(f"Job updated but action_run failed: {exc}")
            job = request.env["leviathan.job"].sudo().browse(job_id).exists()

        return return_Response(
            message="OK",
            status=200,
            errors=warnings,
            data={"job": _serialize_job_detail(job)},
        )
