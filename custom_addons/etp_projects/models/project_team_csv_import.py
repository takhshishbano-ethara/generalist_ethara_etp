import base64
import csv
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


ROLE_TO_FIELD = {
    'tasker': 'project_tasker',
    'qr': 'project_qc_reviewer',
    'qc': 'project_qc_reviewer',
    'pl': 'project_lead',
    'tpm': 'project_tpm',
}

FIELD_LABELS = {
    'project_tasker': 'Tasker',
    'project_qc_reviewer': 'QR',
    'project_lead': 'PL',
    'project_tpm': 'TPM',
}


class ProjectTeamCsvImport(models.TransientModel):
    _name = 'etp.project.team.csv.import'
    _description = 'Project Team CSV Import Wizard'

    project_id = fields.Many2one(
        'project.project',
        string='Project',
        required=True,
        default=lambda self: self.env.context.get('active_id'),
        ondelete='cascade',
    )
    csv_file = fields.Binary(string='CSV File', required=True)
    csv_filename = fields.Char(string='Filename')
    replace_existing = fields.Boolean(
        string='Replace Existing Team Members',
        default=False,
        help='If checked, the project Tasker/QR/PL lists are replaced by the '
             'CSV. Otherwise CSV rows are added on top of the existing team.',
    )
    import_summary = fields.Text(string='Import Summary', readonly=True)

    def action_import(self):
        self.ensure_one()
        if not self.csv_file:
            raise UserError(_("Please upload a CSV file."))
        if not self.project_id:
            raise UserError(_("Project is required."))

        rows = self._read_csv()
        per_field_ids, issues = self._resolve_rows(rows)

        write_vals = self._build_write_vals(per_field_ids)
        if write_vals:
            self.project_id.write(write_vals)

        self.import_summary = self._format_summary(per_field_ids, issues)

        return {
            'type': 'ir.actions.act_window',
            'name': _('Team Import Summary'),
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
            'context': self.env.context,
        }

    def _read_csv(self):
        try:
            raw = base64.b64decode(self.csv_file)
        except Exception as exc:
            raise UserError(_("Unable to decode file: %s") % exc) from exc

        name = (self.csv_filename or '').lower()
        if name.endswith(('.xlsx', '.xlsm')):
            return self._read_excel(raw)
        return self._read_csv_bytes(raw)

    def _read_csv_bytes(self, raw):
        try:
            text = raw.decode('utf-8-sig')
        except UnicodeDecodeError as exc:
            raise UserError(_(
                "CSV must be UTF-8 encoded. Decoding failed: %s"
            ) % exc) from exc

        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise UserError(_("CSV file is empty or has no header row."))

        normalized = {(f or '').strip().lower() for f in reader.fieldnames}
        required = {'name', 'email', 'role'}
        if not required.issubset(normalized):
            raise UserError(_(
                "CSV must contain columns: Name, Email, Role. Found: %s"
            ) % ', '.join(reader.fieldnames))

        rows = []
        for line_no, row in enumerate(reader, start=2):
            rows.append((line_no, {
                (k or '').strip().lower(): (v or '').strip()
                for k, v in row.items()
            }))
        return rows

    def _read_excel(self, raw):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError(_(
                "openpyxl is required to import Excel files: %s"
            ) % exc) from exc
        try:
            wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise UserError(_("Failed to read Excel file: %s") % exc) from exc

        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                raise UserError(_("Excel file is empty."))
            if not header:
                raise UserError(_("Excel header row is empty."))

            header_norm = [
                (str(c).strip().lower() if c is not None else '')
                for c in header
            ]
            required = {'name', 'email', 'role'}
            if not required.issubset(set(header_norm)):
                raise UserError(_(
                    "Excel must contain columns: Name, Email, Role. Found: %s"
                ) % ', '.join([h or '' for h in header_norm]))

            rows = []
            for line_no, raw_row in enumerate(rows_iter, start=2):
                if raw_row is None:
                    continue
                mapped = {}
                for col_name, val in zip(header_norm, raw_row):
                    if not col_name:
                        continue
                    if val is None:
                        mapped[col_name] = ''
                    else:
                        mapped[col_name] = str(val).strip()
                if not any(mapped.values()):
                    continue
                rows.append((line_no, mapped))
            return rows
        finally:
            wb.close()

    def _resolve_rows(self, rows):
        Employee = self.env['hr.employee'].sudo()
        per_field_ids = {field: set() for field in set(ROLE_TO_FIELD.values())}
        issues = []

        for line_no, row in rows:
            role = (row.get('role') or '').lower()
            email = (row.get('email') or '').lower()
            field = ROLE_TO_FIELD.get(role)
            if not field:
                issues.append(_(
                    "Line %(line)s: unknown role '%(role)s' (expected one of: "
                    "Tasker, QR/QC, PL, TPM)"
                ) % {'line': line_no, 'role': row.get('role') or ''})
                continue
            if not email:
                issues.append(_("Line %s: missing email") % line_no)
                continue
            employee = Employee.search([('work_email', '=ilike', email)], limit=1)
            if not employee:
                issues.append(_(
                    "Line %(line)s: no employee found with work email '%(email)s'"
                ) % {'line': line_no, 'email': email})
                continue
            per_field_ids[field].add(employee.id)
        return per_field_ids, issues

    def _build_write_vals(self, per_field_ids):
        write_vals = {}
        for field, emp_ids in per_field_ids.items():
            if self.replace_existing:
                write_vals[field] = [(6, 0, sorted(emp_ids))]
            elif emp_ids:
                write_vals[field] = [(4, eid) for eid in sorted(emp_ids)]
        return write_vals

    def _format_summary(self, per_field_ids, issues):
        lines = [_("Imported team for project '%s':") % self.project_id.name]
        for field, label in FIELD_LABELS.items():
            lines.append(
                "  %s: %d employee(s)" % (label, len(per_field_ids.get(field, set())))
            )
        if issues:
            lines.append('')
            lines.append(_("Issues (%d):") % len(issues))
            lines.extend("  - " + issue for issue in issues)
        return '\n'.join(lines)
