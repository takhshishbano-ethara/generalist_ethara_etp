from __future__ import annotations

import base64
import io
import logging
from datetime import datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

_PASSED_COLUMNS = [
    "item_id", "category", "sub_category", "style", "priority", "complexity",
    "language", "topic", "prompt", "video_file", "duration_seconds",
    "resolution", "fps", "audio_description", "contains_dialogue",
    "speaker_count", "dialogue_transcript", "visual_description",
]
_PASSED_HEADERS = [
    "item_id", "category", "sub-category", "Style", "Priority", "Complexity",
    "Language", "topic", "prompt", "video_file", "duration_seconds",
    "resolution", "fps", "audio_description", "contains_dialogue",
    "speaker_count", "dialogue_transcript", "visual_description",
]
_FAILED_COLUMNS = [
    "item_id", "category", "sub_category", "style", "priority", "complexity",
    "language", "topic", "prompt", "video_file", "issues",
]
_FAILED_HEADERS = [
    "item_id", "category", "sub-category", "Style", "Priority", "Complexity",
    "Language", "topic", "prompt", "video_file", "Issues",
]

_MAX_EXPORT_ROWS = 10000


class T2AVSequenceSheetExportWizard(models.TransientModel):
    _name = "t2av.sequence.sheet.export.wizard"
    _description = "T2AV Sequence Sheet Export Wizard"

    date_from = fields.Date()
    date_to = fields.Date()
    category = fields.Char(
        help="Optional snake_case category filter (e.g. 'human_activities'). "
             "Leave blank for all categories.",
    )
    include_passed = fields.Boolean(default=True)
    include_failed = fields.Boolean(default=True)

    row_count = fields.Integer(readonly=True)
    exported_file = fields.Binary(readonly=True, attachment=False)
    exported_filename = fields.Char(readonly=True)

    def _build_domain(self, sheet_type):
        domain = [("sheet_type", "=", sheet_type)]
        if self.date_from:
            domain.append(("create_date", ">=", str(self.date_from)))
        if self.date_to:
            domain.append(("create_date", "<=", str(self.date_to) + " 23:59:59"))
        if self.category:
            domain.append(("category", "=", self.category.strip()))
        return domain

    def action_build_xlsx(self):
        self.ensure_one()
        try:
            from openpyxl import Workbook
        except ImportError as e:
            raise UserError(_(
                "openpyxl is required for sequence-sheet export. "
                "Install it: pip install openpyxl"
            )) from e

        if not (self.include_passed or self.include_failed):
            raise UserError(_("Select at least one of Passed / Failed."))

        Row = self.env["t2av.sequence.sheet.row"]
        wb = Workbook(write_only=True)
        total = 0

        if self.include_passed:
            ws = wb.create_sheet(title="Master")
            ws.append(_PASSED_HEADERS)
            rows = Row.search(self._build_domain("passed"), limit=_MAX_EXPORT_ROWS)
            for r in rows:
                ws.append([self._cell(r, col) for col in _PASSED_COLUMNS])
            total += len(rows)

        if self.include_failed:
            ws = wb.create_sheet(title="Failures")
            ws.append(_FAILED_HEADERS)
            rows = Row.search(self._build_domain("failed"), limit=_MAX_EXPORT_ROWS)
            for r in rows:
                ws.append([self._cell(r, col) for col in _FAILED_COLUMNS])
            total += len(rows)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        today = datetime.utcnow().strftime("%Y%m%d")
        filename = f"T2AV-{today}-VideoGen-FINAL.xlsx"

        self.write({
            "exported_file": base64.b64encode(buf.getvalue()),
            "exported_filename": filename,
            "row_count": total,
        })

        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    @api.model
    def _cell(self, record, col):
        value = record[col]
        if isinstance(value, bool):
            return value
        if value is False or value is None:
            return ""
        return value
