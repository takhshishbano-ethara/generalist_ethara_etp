#!/usr/bin/env python3
"""Ingest per-patient Excel workbooks into the unified patients.json contract.

Supports two layouts (auto-detected):

  nested:
      <data-dir>/Patient_<pid>/structured/*.xlsx
      <data-dir>/Patient_<pid>/records/<category>/*.pdf
      <data-dir>/Patient_<pid>/wsi/*.svs

  flat:
      <data-dir>/Data/Clinical Data_<pid>.xlsx
      <data-dir>/Data/WSI_<pid>.<n>.svs
      <data-dir>/Data/Clinical Data_<pid>.pdf       (overview)
      <data-dir>/De-identified Records/*.pdf        (attributed to --records-pid)

PDFs are copied into the module's static/docs/<pid>/<category>/ so Odoo
auto-serves them at /loki_dashboard_2/static/docs/<pid>/<cat>/<file>.

Usage:
    python tools/ingest_excel.py [--data-dir PATH] [--output PATH] \
                                 [--static-docs-dir PATH] [--records-pid 4]
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shutil
import sys
from typing import Any, Iterable

try:
    import openpyxl  # noqa: F401
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl not installed.\n"
        "  Install it with:  pip install openpyxl\n"
    )
    sys.exit(1)

from openpyxl import load_workbook


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MODULE_DIR = os.path.dirname(SCRIPT_DIR)

DEFAULT_DATA_DIR = os.path.join(MODULE_DIR, "data", "Clinical_Data")
DEFAULT_OUTPUT = os.path.join(DEFAULT_DATA_DIR, "patients.json")
DEFAULT_STATIC_DOCS = os.path.join(MODULE_DIR, "static", "docs")

SHEET_ALIASES = {
    # Excel truncates sheet names to 31 chars. Map truncated → canonical.
    "diagnostic - molecular patholog": "diagnostic - molecular pathology",
    "diagnostic - clinical biochemis": "diagnostic - clinical biochemistry",
    "diagnostic - staging informatio": "diagnostic - staging information",
}

CATEGORY_MAP = [
    # (regex on uppercased filename, category, modality-or-subtype)
    (re.compile(r"^DS\s+CHEMO"), "discharge_summary", "chemo"),
    (re.compile(r"^DS\s+SURGERY"), "discharge_summary", "surgery"),
    (re.compile(r"^DS\b"), "discharge_summary", None),
    (re.compile(r"^ENDOSCOPY"), "endoscopy", None),
    (re.compile(r"^HAEMATOLOGY|^HEMATOLOGY"), "hematology", None),
    (re.compile(r"^HISTOPATHOLOGY"), "pathology", None),
    (re.compile(r"^MOLECULAR"), "molecular", None),
    (re.compile(r"^RADIOLOGY\s*\(CT"), "imaging", "ct"),
    (re.compile(r"^RADIOLOGY\s*\(MRI"), "imaging", "mri"),
    (re.compile(r"^RADIOLOGY\s*\(PET"), "imaging", "pet_ct"),
    (re.compile(r"^RADIOLOGY\s*\(USG"), "imaging", "usg"),
    (re.compile(r"^RADIOLOGY\s*\(X-RAY"), "imaging", "xray"),
    (re.compile(r"^CLINICAL\s+DATA"), "overview", None),
]


# ────────────────────────────────────────────────────────────────────────
# Generic helpers
# ────────────────────────────────────────────────────────────────────────

def coerce_date(value: Any) -> str | None:
    """Normalise various date encodings to ISO yyyy-mm-dd. Range strings
    keep their delimiter so caller can split start/end."""
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value).strip()
    if not s or s.upper() == "NA":
        return None
    # range like '11-07-2020 to 15-02-2021'
    if " to " in s.lower():
        parts = re.split(r"\s+to\s+", s, flags=re.IGNORECASE)
        starts = [coerce_date(p) for p in parts]
        return " to ".join(p for p in starts if p)
    # remove stray spaces inside `dd /mm /yyyy` or `dd- m-yyyy`
    s = re.sub(r"\s+", "", s)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y", "%d-%m-%y"):
        try:
            return dt.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # last-ditch: pull dd, mm, yyyy
    m = re.match(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = ("20" + y) if int(y) < 70 else ("19" + y)
        try:
            return dt.date(int(y), int(mo), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            return None
    return None


def split_date_range(value: Any) -> tuple[str | None, str | None]:
    """Split a coerced date range into (start, end). End may equal start."""
    iso = coerce_date(value)
    if iso is None:
        return (None, None)
    if " to " in iso:
        a, _, b = iso.partition(" to ")
        return (a.strip() or None, b.strip() or None)
    return (iso, iso)


def parse_bracket_list(value: Any) -> list[str]:
    """`'[CEA, PSA]'` → `['CEA', 'PSA']`. Plain string → `[value]`."""
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    if s.startswith("[") and s.endswith("]"):
        s = s[1:-1]
    parts = [p.strip() for p in s.split(",")]
    return [p for p in parts if p]


def parse_csv_list(value: Any) -> list[str]:
    """Comma-separated, no brackets. Used for molecular gene/variation lists."""
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.upper() == "NA":
        return []
    return [p.strip() for p in s.split(",") if p.strip()]


def s(value: Any) -> str | None:
    """String-ify a cell, treating 'NA' and empties as None."""
    if value is None:
        return None
    out = str(value).strip()
    if not out or out.upper() == "NA":
        return None
    return out


def tnm_v(value: Any) -> str | None:
    """TNM value: normalize whole-number floats so M=0 doesn't render as 'M0.0'."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return s(value)


def num(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        v = str(value).strip().rstrip("%")
        if not v or v.upper() == "NA":
            return None
        if "." in v:
            return float(v)
        return int(v)
    except (TypeError, ValueError):
        return None


def safe_filename(name: str) -> str:
    """Strip characters Odoo's static handler won't pass through cleanly."""
    out = re.sub(r"[^A-Za-z0-9._\-]+", "_", name)
    out = re.sub(r"_+", "_", out).strip("._-")
    return out or "file"


# ────────────────────────────────────────────────────────────────────────
# Vertical key-value sheet reader
# ────────────────────────────────────────────────────────────────────────

_NO_KEY = re.compile(r"^no\s*[.:]?\s*(\d+)?\s*$", re.IGNORECASE)


def iter_clean_rows(ws) -> Iterable[tuple]:
    """Yield non-empty rows with trailing Nones stripped."""
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        if all(c is None for c in row):
            continue
        r = list(row)
        while r and r[-1] is None:
            r.pop()
        yield tuple(r)


def split_records(rows: Iterable[tuple]) -> list[list[tuple]]:
    """Group rows into records delimited by `No.` / `No:N` header rows.

    A record contains all rows after its `No*` row, up to the next `No*` or end.
    Rows before the first `No*` (e.g. a title row) are discarded.
    """
    groups: list[list[tuple]] = []
    cur: list[tuple] | None = None
    for r in rows:
        first = r[0]
        if isinstance(first, str) and _NO_KEY.match(first.strip()):
            if cur is not None:
                groups.append(cur)
            cur = []
            continue
        if cur is None:
            continue
        cur.append(r)
    if cur is not None:
        groups.append(cur)
    return [g for g in groups if g]


def record_to_dict(rows: list[tuple]) -> dict[str, Any]:
    """Flatten a record's rows into a dict. Duplicate keys become lists.

    Multi-column rows are stored as the value list (col[1:])."""
    out: dict[str, Any] = {}
    for r in rows:
        if not r:
            continue
        k = r[0]
        if not isinstance(k, str):
            continue
        key = k.strip()
        if not key:
            continue
        if len(r) <= 1:
            value: Any = None
        elif len(r) == 2:
            value = r[1]
        else:
            value = list(r[1:])
        if key in out:
            existing = out[key]
            if isinstance(existing, list) and not (len(r) > 2):
                existing.append(value)
            else:
                out[key] = [existing, value]
        else:
            out[key] = value
    return out


def is_empty_sheet(rows: list[tuple]) -> bool:
    """`Not given`, `Not prescribed`, etc., as the only payload."""
    if not rows:
        return True
    for r in rows:
        if not r:
            continue
        v = r[0]
        if isinstance(v, str) and re.match(r"^not\s+(given|prescribed|provided|applicable)", v.strip(), re.IGNORECASE):
            return True
    return False


def get_sheet(wb, canonical_name: str):
    """Look up a sheet by canonical (full-length) name; tolerate truncation."""
    target = canonical_name.lower().strip()
    for sn in wb.sheetnames:
        norm = sn.lower().strip()
        if norm == target:
            return wb[sn]
        if SHEET_ALIASES.get(norm) == target:
            return wb[sn]
    return None


# ────────────────────────────────────────────────────────────────────────
# Sheet parsers
# ────────────────────────────────────────────────────────────────────────

def parse_personal(ws) -> dict[str, Any]:
    info: dict[str, Any] = {}
    for r in iter_clean_rows(ws):
        if len(r) < 2 or not isinstance(r[0], str):
            continue
        key = r[0].strip().rstrip(":").lower()
        val = r[1]
        if key == "age":
            info["age_years"] = num(val)
        elif key == "gender":
            g = s(val)
            info["gender"] = (g[0].upper() if g else None)
        elif key == "height (cm)":
            info["height_cm"] = num(val)
        elif key == "weight (kg)":
            info["weight_kg"] = num(val)
        elif key == "bmi":
            info["bmi"] = num(val)
        elif key == "diet":
            info["diet"] = s(val)
        elif key == "marital status":
            info["marital_status"] = s(val)
        elif key == "no. of children":
            info["children_count"] = num(val)
        elif key == "religion":
            info["religion"] = s(val)
        elif key == "mother tongue":
            info["mother_tongue"] = s(val)
        elif key == "occupation":
            info["occupation"] = s(val)
    return info


def parse_medical(ws) -> dict[str, Any]:
    info: dict[str, Any] = {
        "site": None, "icd_o_3_topo": None, "morphology": None,
        "icd_o_3_morph": None, "location": None, "icd_o_3_location": None,
        "laterality": None, "age_at_diagnosis": None,
        "symptoms": [], "risk_factors": [], "comorbidities": [],
        "family_history": [],
    }
    initial_tnm = {"t": None, "n": None, "m": None}
    section = None
    for r in iter_clean_rows(ws):
        if not r:
            continue
        first = r[0]
        # section header rows like ('TNM Staging',)
        if len(r) == 1 and isinstance(first, str):
            section = first.strip().lower()
            continue
        if not isinstance(first, str):
            continue
        key = first.strip().lower().rstrip(":")
        val = r[1] if len(r) > 1 else None
        extras = list(r[2:]) if len(r) > 2 else []
        if key == "cancer site":
            info["site"] = s(val)
        elif key == "icd-o-3 code (site)":
            info["icd_o_3_topo"] = s(val)
        elif key == "cancer location":
            info["location"] = s(val)
        elif key == "icd-o-3 code (location)":
            info["icd_o_3_location"] = s(val)
        elif key == "morphology":
            info["morphology"] = s(val)
        elif key == "icd-o-3 code (morphology)":
            info["icd_o_3_morph"] = s(val)
        elif key == "laterality":
            info["laterality"] = s(val)
        elif key == "age at diagnosis":
            info["age_at_diagnosis"] = num(val)
        elif key == "symptoms before diagnosis":
            for piece in parse_csv_list(val):
                info["symptoms"].append(piece)
        elif key == "t" and section and "staging" in section:
            initial_tnm["t"] = tnm_v(val)
        elif key == "n" and section and "staging" in section:
            initial_tnm["n"] = tnm_v(val)
        elif key == "m" and section and "staging" in section:
            initial_tnm["m"] = tnm_v(val)
        elif key == "smoking" and s(val):
            info["risk_factors"].append(f"Smoking: {s(val)}")
        elif key == "alcohol consumption" and s(val):
            info["risk_factors"].append(f"Alcohol: {s(val)}")
        elif key == "other forms of tobacco use":
            spec = s(extras[0]) if extras else None
            if spec:
                info["risk_factors"].append(f"Other tobacco: {spec}")
        elif key == "history of immunosuppression":
            spec = s(extras[0]) if extras else None
            if spec:
                info["risk_factors"].append(f"Immunosuppression: {spec}")
        elif key == "allergies":
            spec = s(extras[0]) if extras else None
            if spec:
                info["risk_factors"].append(f"Allergy: {spec}")
        elif key == "co-morbidity":
            for piece in ([s(val)] + [s(x) for x in extras]):
                if piece:
                    info["comorbidities"].append(piece)
        elif key == "relationship with patient":
            rel = s(val)
            if rel:
                info["family_history"].append({"relationship": rel})
        elif key == "cancer type" and info["family_history"]:
            info["family_history"][-1]["cancer_type"] = s(val)
        elif key == "age of diagnosis" and info["family_history"]:
            info["family_history"][-1]["age_at_diagnosis"] = num(val) or s(val)
    info["initial_tnm"] = {k: v for k, v in initial_tnm.items() if v is not None}
    # dedupe lists, drop empties
    for k in ("symptoms", "risk_factors", "comorbidities"):
        seen = set()
        out = []
        for v in info[k]:
            if v and v not in seen:
                seen.add(v)
                out.append(v)
        info[k] = out
    info["family_history"] = [fh for fh in info["family_history"] if fh.get("cancer_type") and fh["cancer_type"].upper() != "NA"]
    return info


def parse_events_index(ws) -> list[dict[str, Any]]:
    out = []
    saw_header = False
    for r in iter_clean_rows(ws):
        if len(r) < 2:
            continue
        a, b = r[0], r[1]
        if not saw_header:
            if isinstance(a, str) and a.strip().lower() == "date":
                saw_header = True
            continue
        date = coerce_date(a)
        if not date:
            continue
        category = s(b) or "other"
        out.append({"date": date, "category": category.lower(), "title": (s(b) or "").title(), "detail": "", "ref": "events_index"})
    return out


def parse_treatment_chemo(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        start, end = split_date_range(d.get("Date"))
        drugs = parse_bracket_list(d.get("Drug")) or parse_csv_list(d.get("Drug"))
        out.append({
            "drug": ", ".join(drugs) if drugs else s(d.get("Drug")),
            "start": start,
            "end": end,
            "cycles": num(d.get("cycles") or d.get("Cycles") or d.get("cycles ")),
            "dosage": s(d.get("dosage") or d.get("Dosage") or d.get("dosage ")),
            "comments": s(d.get("Comments")),
        })
    return out


def parse_treatment_targeted(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        start, end = split_date_range(d.get("Date"))
        drugs = parse_bracket_list(d.get("Drug")) or parse_csv_list(d.get("Drug"))
        out.append({
            "drug": ", ".join(drugs) if drugs else s(d.get("Drug")),
            "start": start,
            "end": end,
            "cycles": num(d.get("cycles") or d.get("Cycles") or d.get("cycles ")),
            "dosage": s(d.get("dosage") or d.get("Dosage") or d.get("dosage ")),
            "comments": s(d.get("Comments")),
        })
    return out


def parse_treatment_immuno(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        start, end = split_date_range(d.get("Date"))
        drugs = parse_bracket_list(d.get("Drug")) or parse_csv_list(d.get("Drug"))
        dose_mg = num(d.get("Dose (mg)") or d.get("dose_mg") or d.get("dose"))
        out.append({
            "drug": ", ".join(drugs) if drugs else s(d.get("Drug")),
            "start": start,
            "end": end,
            "cycles": num(d.get("cycles") or d.get("Cycles") or d.get("cycles ")),
            "dose_mg": dose_mg,
            "comments": s(d.get("Comments")),
        })
    return out


def parse_treatment_radiation(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        date = coerce_date(d.get("Date"))
        out.append({
            "type": s(d.get("Type")) or s(d.get("Modality")),
            "date": date,
            "fractions": num(d.get("Fractions")),
            "dose": s(d.get("Dose")),
            "comments": s(d.get("Comments")),
        })
    return out


def parse_treatment_surgery(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        out.append({
            "procedure": s(d.get("Surgery")) or s(d.get("Procedure")),
            "date": coerce_date(d.get("Date")),
            "findings": s(d.get("Comments")),
        })
    return out


def parse_performance(ws) -> tuple[list[dict[str, Any]], dict[str, Any] | None]:
    """Returns (history, latest_extras) where extras carries survival_status / date.

    Date can appear under any of: "Date", "Date of Criteria", "Date Of Criteria",
    "Date of  Criteria" (double space — seen in real source files). We match
    case-insensitively after collapsing whitespace.
    """
    history: list[dict[str, Any]] = []
    survival: dict[str, Any] = {}
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        # Tolerant key lookup — collapse whitespace, lowercase
        d_norm = {re.sub(r"\s+", " ", str(k).strip().lower()): v for k, v in d.items()}
        date = (coerce_date(d_norm.get("date"))
                or coerce_date(d_norm.get("date of criteria")))
        label = s(d_norm.get("value"))
        criteria = s(d_norm.get("type of criteria")) or "Recist Criteria"
        if "survival status" in d_norm:
            survival["status"] = s(d_norm["survival status"])
            survival["as_of"] = date
        if date and label:
            history.append({"date": date, "label": label, "recist": _recist_from_label(label), "criteria": criteria})
    return history, (survival or None)


_RECIST_PATTERNS = [
    (re.compile(r"\bicr\b|immune\s+complete", re.IGNORECASE), "iCR"),
    (re.compile(r"\bipr\b|immune\s+partial", re.IGNORECASE), "iPR"),
    (re.compile(r"\bisd\b|immune\s+stable", re.IGNORECASE), "iSD"),
    (re.compile(r"\bipd\b|immune\s+progressive", re.IGNORECASE), "iPD"),
    (re.compile(r"complete\s+response|\bcr\b", re.IGNORECASE), "CR"),
    (re.compile(r"partial\s+response|\bpr\b", re.IGNORECASE), "PR"),
    (re.compile(r"stable\s+disease|\bsd\b", re.IGNORECASE), "SD"),
    (re.compile(r"progressive\s+disease|progression|\bpd\b", re.IGNORECASE), "PD"),
]


def _recist_from_label(label: str | None) -> str | None:
    if not label:
        return None
    for rx, code in _RECIST_PATTERNS:
        if rx.search(label):
            return code
    return None


def parse_pathology(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        out.append({
            "date": coerce_date(d.get("Date")),
            "type": s(d.get("Type Of Sample")) or s(d.get("Type")),
            "tissue": _join(parse_bracket_list(d.get("Tissue"))),
            "location": _join(parse_bracket_list(d.get("CancerSite Location")) or parse_bracket_list(d.get("CancerSite"))),
            "pt": s(d.get("PT")),
            "pn": s(d.get("PN")),
            "pm": s(d.get("PM")) if not isinstance(d.get("PM"), (int, float)) else str(int(d["PM"]) if float(d["PM"]).is_integer() else d["PM"]),
            "grade": _join(parse_bracket_list(d.get("CancerGrade")) or parse_bracket_list(d.get("Grade"))),
            "morphology": _join(parse_bracket_list(d.get("Morphology"))),
        })
    return out


def _join(items: list[str]) -> str | None:
    items = [i for i in items if i and i.upper() != "NA"]
    return ", ".join(items) if items else None


def parse_molecular(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        date = coerce_date(d.get("Date"))
        method = s(d.get("Method")) or ""
        method_u = method.upper()
        if method_u in ("NGS",):
            genes = parse_csv_list(d.get("Genes")) or parse_bracket_list(d.get("Genes"))
            variations = parse_csv_list(d.get("Type Of Variations")) or parse_bracket_list(d.get("Type Of Variations"))
            assay = s(d.get("NGS Assay"))
            for i, gene in enumerate(genes):
                out.append({
                    "date": date,
                    "category": "NGS",
                    "gene": gene,
                    "variation": variations[i] if i < len(variations) else None,
                    "result": None,
                    "clone": None,
                    "assay": assay,
                    "comments": s(d.get("Comments")),
                })
        elif method_u in ("IMMUNOHISTOCHEMISTRY", "IHC"):
            markers = parse_csv_list(d.get("Marker")) or parse_bracket_list(d.get("Marker"))
            results = parse_csv_list(d.get("Result")) or parse_bracket_list(d.get("Result"))
            clone = s(d.get("Clone"))
            if not markers and s(d.get("Marker")):
                markers = [s(d.get("Marker"))]
                results = [s(d.get("Result"))]
            for i, marker in enumerate(markers):
                out.append({
                    "date": date,
                    "category": "IHC",
                    "marker": marker,
                    "result": results[i] if i < len(results) else None,
                    "clone": clone,
                    "comments": s(d.get("Comments")),
                })
        elif method_u in ("PCR",):
            marker = s(d.get("Marker"))
            out.append({
                "date": date,
                "category": "PCR",
                "marker": marker,
                "result": s(d.get("Result")),
                "comments": s(d.get("Comments")),
            })
        else:
            out.append({
                "date": date,
                "category": method or "Other",
                "comments": s(d.get("Comments")),
            })
    return out


def parse_biomarkers(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        date = coerce_date(d.get("Date"))
        tests = parse_bracket_list(d.get("Test"))
        units = parse_bracket_list(d.get("Unit"))
        values = parse_bracket_list(d.get("Value"))
        for i, t in enumerate(tests):
            v = values[i] if i < len(values) else None
            u = units[i] if i < len(units) else None
            out.append({
                "date": date,
                "test": t,
                "value": num(v) if num(v) is not None else s(v),
                "unit": u,
            })
    return out


def parse_lab_panel(ws, panel_label: str) -> list[dict[str, Any]]:
    """Hematology / Clinical BioChem: triples of (Parameter, Value, Units)."""
    out = []
    cur_date = None
    cur_test = None
    cur_param = None
    cur_value = None
    for r in iter_clean_rows(ws):
        if not r:
            continue
        first = r[0]
        if isinstance(first, str) and _NO_KEY.match(first.strip()):
            cur_param = cur_value = None
            continue
        if not isinstance(first, str):
            continue
        key = first.strip()
        val = r[1] if len(r) > 1 else None
        if key.lower() == "date":
            cur_date = coerce_date(val)
        elif key.lower() == "test":
            cur_test = s(val)
            cur_param = cur_value = None
        elif key.lower().endswith("- parameter"):
            cur_param = s(val)
        elif key.lower() == "value":
            cur_value = val
        elif key.lower().endswith("- value"):
            # e.g. 'ESR - Value' (single-parameter test)
            cur_param = cur_param or (cur_test or key.split(" - ")[0])
            cur_value = val
        elif key.lower() in ("units", "unit") and (cur_value is not None or cur_param):
            if cur_date and (cur_param or cur_test):
                out.append({
                    "date": cur_date,
                    "panel": panel_label,
                    "test": cur_test or cur_param,
                    "parameter": cur_param,
                    "value": num(cur_value) if num(cur_value) is not None else s(cur_value),
                    "unit": s(val),
                })
            cur_value = None
    return out


def parse_imaging(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        out.append({
            "date": coerce_date(d.get("Date")),
            "modality": s(d.get("Type")) or s(d.get("Modality")),
            "impression": s(d.get("Impression")),
        })
    return out


def parse_staging(ws) -> list[dict[str, Any]]:
    out = []
    for rec in split_records(iter_clean_rows(ws)):
        if is_empty_sheet(rec):
            continue
        d = record_to_dict(rec)
        # Two dates can appear: 'Date' and 'Date ' (trailing space)
        date = coerce_date(d.get("Date")) or coerce_date(d.get("Date "))
        out.append({
            "date": date,
            "system": s(d.get("Staging System")),
            "edition": s(d.get("Edition")) or s(d.get("Edition ")),
            "type": s(d.get("Classification Type")) or s(d.get("Classification Type ")),
            "t": tnm_v(d.get("Value T")) or tnm_v(d.get("Value T ")),
            "n": tnm_v(d.get("Value N")) or tnm_v(d.get("Value N ")),
            "m": tnm_v(d.get("Value M")) or tnm_v(d.get("Value M ")),
            "metastatic_site": s(d.get("MetastaticSite")) or s(d.get("MetastaticSite ")),
            "grade": s(d.get("CancerGrade")) or s(d.get("CancerGrade ")),
            "stage": s(d.get("CancerStage")) or s(d.get("CancerStage ")),
        })
    return out


def parse_consultation(ws) -> list[dict[str, Any]]:
    out = []
    saw_header = False
    for r in iter_clean_rows(ws):
        if not saw_header:
            if r and isinstance(r[0], str) and r[0].strip().lower() == "date":
                saw_header = True
            continue
        if not r or len(r) < 2:
            continue
        date = coerce_date(r[0])
        if not date:
            continue
        out.append({
            "date": date,
            "comments": s(r[1]) if len(r) > 1 else None,
            "advice": s(r[2]) if len(r) > 2 else None,
        })
    return out


# ────────────────────────────────────────────────────────────────────────
# Per-patient assembly
# ────────────────────────────────────────────────────────────────────────

def empty_patient(pid: str, code: str) -> dict[str, Any]:
    return {
        "id": pid,
        "code": code,
        "demographics": {},
        "cancer_profile": {"symptoms": [], "risk_factors": [], "comorbidities": [], "family_history": []},
        "current_stage": {},
        "latest_performance": {},
        "status": "unknown",
        "days_since_diagnosis": None,
        "events": [],
        "treatments": {"chemotherapy": [], "immunotherapy": [], "targeted": [], "radiation": [], "surgery": []},
        "stagings": [],
        "pathology": [],
        "molecular": [],
        "biomarkers": [],
        "labs": [],
        "imaging": [],
        "performance_history": [],
        "consultations": [],
        "wsi_slides": [],
        "documents": [],
    }


def parse_workbook(path: str, pid: str, code: str) -> dict[str, Any]:
    print(f"  reading workbook: {os.path.basename(path)}", file=sys.stderr)
    wb = load_workbook(path, read_only=True, data_only=True)
    p = empty_patient(pid, code)

    if (ws := get_sheet(wb, "Personal Information")):
        p["demographics"] = parse_personal(ws)
    if (ws := get_sheet(wb, "Medical Information")):
        med = parse_medical(ws)
        initial_tnm = med.pop("initial_tnm", {})
        p["cancer_profile"].update({
            "site": med.get("site"),
            "icd_o_3_topo": med.get("icd_o_3_topo"),
            "location": med.get("location"),
            "icd_o_3_location": med.get("icd_o_3_location"),
            "morphology": med.get("morphology"),
            "icd_o_3_morph": med.get("icd_o_3_morph"),
            "laterality": med.get("laterality"),
            "age_at_diagnosis": med.get("age_at_diagnosis"),
            "symptoms": med.get("symptoms", []),
            "risk_factors": med.get("risk_factors", []),
            "comorbidities": med.get("comorbidities", []),
            "family_history": med.get("family_history", []),
        })
        if initial_tnm:
            p["current_stage"].update({"t": initial_tnm.get("t"), "n": initial_tnm.get("n"), "m": initial_tnm.get("m")})

    if (ws := get_sheet(wb, "Events Index")):
        p["events"] = parse_events_index(ws)
    if (ws := get_sheet(wb, "Treatment - Chemotherapy")):
        p["treatments"]["chemotherapy"] = parse_treatment_chemo(ws)
    if (ws := get_sheet(wb, "Treatment - Immunotherapy")):
        p["treatments"]["immunotherapy"] = parse_treatment_immuno(ws)
    if (ws := get_sheet(wb, "Treatment - Targeted Therapy")):
        p["treatments"]["targeted"] = parse_treatment_targeted(ws)
    if (ws := get_sheet(wb, "Treatment (Radiation)")):
        p["treatments"]["radiation"] = parse_treatment_radiation(ws)
    if (ws := get_sheet(wb, "Treatment - HT (Radiation)")):
        p["treatments"]["radiation"].extend(parse_treatment_radiation(ws))
    if (ws := get_sheet(wb, "Treatment (Surgery)")):
        p["treatments"]["surgery"] = parse_treatment_surgery(ws)

    if (ws := get_sheet(wb, "Performance Status")):
        history, survival = parse_performance(ws)
        p["performance_history"] = sorted(history, key=lambda x: x["date"])
        if p["performance_history"]:
            latest = p["performance_history"][-1]
            p["latest_performance"] = {"as_of": latest["date"], "label": latest["label"], "recist": latest["recist"]}
        if survival:
            label = (survival.get("status") or "").lower()
            if "alive" in label:
                p["status"] = "alive"
            elif "decease" in label or "expired" in label or "death" in label:
                p["status"] = "deceased"
            if survival.get("as_of"):
                p["latest_performance"]["as_of"] = survival["as_of"]

    if (ws := get_sheet(wb, "Diagnostic - Pathology")):
        p["pathology"] = parse_pathology(ws)
    if (ws := get_sheet(wb, "Diagnostic - Molecular Pathology")):
        p["molecular"] = parse_molecular(ws)
    if (ws := get_sheet(wb, "Diagnostic - Serum Biomarker")):
        p["biomarkers"] = parse_biomarkers(ws)
    if (ws := get_sheet(wb, "Diagnostic - Hematology")):
        p["labs"].extend(parse_lab_panel(ws, "Hematology"))
    if (ws := get_sheet(wb, "Diagnostic - Clinical BioChemistry")):
        p["labs"].extend(parse_lab_panel(ws, "Clinical BioChemistry"))
    if (ws := get_sheet(wb, "Diagnostic - ImageSite")):
        p["imaging"] = sorted(parse_imaging(ws), key=lambda x: x.get("date") or "")
    if (ws := get_sheet(wb, "Diagnostic - Staging Information")):
        p["stagings"] = parse_staging(ws)
        if p["stagings"]:
            latest_stage = p["stagings"][-1]
            p["current_stage"] = {
                "edition": latest_stage.get("edition"),
                "t": latest_stage.get("t") or p["current_stage"].get("t"),
                "n": latest_stage.get("n") or p["current_stage"].get("n"),
                "m": latest_stage.get("m") or p["current_stage"].get("m"),
                "stage": latest_stage.get("stage"),
                "grade": latest_stage.get("grade"),
                "as_of": latest_stage.get("date"),
            }
    if (ws := get_sheet(wb, "Consultation")):
        p["consultations"] = parse_consultation(ws)

    wb.close()

    # derived: days_since_diagnosis
    earliest = None
    for ev in p["events"]:
        d = ev.get("date")
        if d and (earliest is None or d < earliest):
            earliest = d
    if earliest:
        try:
            dx_date = dt.datetime.strptime(earliest, "%Y-%m-%d").date()
            today = dt.date.today()
            p["days_since_diagnosis"] = (today - dx_date).days
        except ValueError:
            pass

    # enrich events with treatment/imaging anchors so the timeline isn't bare
    p["events"] = build_unified_events(p)
    return p


def build_unified_events(p: dict[str, Any]) -> list[dict[str, Any]]:
    """Merge events-index entries with first-of-each treatment/imaging milestone."""
    by_date: dict[str, list[dict[str, Any]]] = {}
    def add(date, category, title, detail="", ref=""):
        if not date:
            return
        by_date.setdefault(date, []).append({
            "date": date, "category": category, "title": title, "detail": detail, "ref": ref,
        })
    for ev in p["events"]:
        add(ev["date"], ev["category"], ev["title"], ev.get("detail", ""), ev.get("ref", ""))
    for s_rec in p["treatments"]["surgery"]:
        add(s_rec["date"], "surgery", s_rec.get("procedure") or "Surgery", (s_rec.get("findings") or "")[:300], "surgery")
    for grp, cat in (("chemotherapy", "chemo"), ("immunotherapy", "immuno"), ("targeted", "targeted")):
        for t in p["treatments"][grp]:
            start, drug = t.get("start"), t.get("drug")
            if start and drug:
                add(start, cat, drug, f"Cycles: {t.get('cycles') or '?'}", grp)
    for r in p["treatments"]["radiation"]:
        if r.get("date"):
            add(r["date"], "radiation", r.get("type") or "Radiation", r.get("comments") or "", "radiation")
    for img in p["imaging"]:
        if img.get("date"):
            add(img["date"], "imaging", img.get("modality") or "Imaging", (img.get("impression") or "")[:300], "imaging")
    for path in p["pathology"]:
        if path.get("date"):
            add(path["date"], "pathology", f"Pathology: {path.get('type') or 'Biopsy'}", path.get("location") or "", "pathology")
    for mol in p["molecular"]:
        if mol.get("date"):
            title = f"{mol.get('category', 'Molecular')}: " + (mol.get("gene") or mol.get("marker") or "")
            add(mol["date"], "molecular", title.strip(": "), mol.get("variation") or mol.get("result") or "", "molecular")
    out = []
    for date in sorted(by_date.keys()):
        out.extend(by_date[date])
    return out


# ────────────────────────────────────────────────────────────────────────
# Document discovery & copying
# ────────────────────────────────────────────────────────────────────────

def categorise(filename: str) -> tuple[str, str | None]:
    upper = filename.upper()
    for rx, cat, sub in CATEGORY_MAP:
        if rx.search(upper):
            return cat, sub
    return "other", None


def attach_documents(patient: dict[str, Any], pdfs: list[str], static_docs_dir: str, module_static_url: str) -> None:
    for src in pdfs:
        if not os.path.isfile(src):
            continue
        base = os.path.basename(src)
        cat, sub = categorise(base)
        safe = safe_filename(base)
        rel_dir = os.path.join(patient["id"], cat, sub) if sub else os.path.join(patient["id"], cat)
        out_dir = os.path.join(static_docs_dir, rel_dir)
        os.makedirs(out_dir, exist_ok=True)
        dest = os.path.join(out_dir, safe)
        if not os.path.exists(dest) or os.path.getsize(dest) != os.path.getsize(src):
            shutil.copy2(src, dest)
        url_parts = [module_static_url, patient["id"], cat]
        if sub:
            url_parts.append(sub)
        url_parts.append(safe)
        patient["documents"].append({
            "category": cat,
            "subcategory": sub,
            "filename": base,
            "title": base.rsplit(".", 1)[0],
            "url": "/" + "/".join(url_parts),
        })


def attach_wsi_slides(patient: dict[str, Any], svs_paths: list[str], module_dzi_url: str) -> None:
    for src in sorted(svs_paths):
        if not os.path.isfile(src):
            continue
        base = os.path.basename(src)
        stem = base.rsplit(".", 1)[0]
        patient["wsi_slides"].append({
            "slide": base,
            "dzi_path": f"/{module_dzi_url}/{patient['id']}/{stem}.dzi",
            "source": src,
        })


# ────────────────────────────────────────────────────────────────────────
# Layout discovery
# ────────────────────────────────────────────────────────────────────────

def discover_flat(data_dir: str) -> list[tuple[str, str, str, list[str]]]:
    """Find (pid, code, xlsx_path, wsi_paths[]) for the flat layout."""
    data_subdir = os.path.join(data_dir, "Data")
    base = data_subdir if os.path.isdir(data_subdir) else data_dir
    out = []
    rx = re.compile(r"^Clinical\s*Data\s*[_]\s*(.+?)\.xlsx$", re.IGNORECASE)
    for fn in sorted(os.listdir(base)):
        m = rx.match(fn)
        if not m:
            continue
        raw_pid = m.group(1).strip()
        pid = re.sub(r"[^A-Za-z0-9_]+", "_", raw_pid).strip("_")
        code = f"P{raw_pid.replace('_', '-')}"
        xlsx = os.path.join(base, fn)
        stem_m = re.match(r"\d+", raw_pid)
        pid_stem = stem_m.group(0) if stem_m else raw_pid
        wsi_rx_full = re.compile(rf"^WSI[_]{re.escape(raw_pid)}\.\d+\.svs$", re.IGNORECASE)
        wsi_rx_stem = re.compile(rf"^WSI[_]{re.escape(pid_stem)}\.\d+\.svs$", re.IGNORECASE)
        wsi = [os.path.join(base, f) for f in os.listdir(base)
               if wsi_rx_full.match(f) or wsi_rx_stem.match(f)]
        out.append((pid, code, xlsx, wsi))
    return out


def discover_nested(data_dir: str) -> list[tuple[str, str, str, list[str]]]:
    out = []
    rx = re.compile(r"^Patient_(.+)$")
    for sub in sorted(os.listdir(data_dir)):
        if not os.path.isdir(os.path.join(data_dir, sub)):
            continue
        m = rx.match(sub)
        if not m:
            continue
        pid = m.group(1)
        code = f"P{pid.replace('_', '-')}"
        structured = os.path.join(data_dir, sub, "structured")
        if not os.path.isdir(structured):
            continue
        xlsxs = [os.path.join(structured, f) for f in os.listdir(structured) if f.lower().endswith(".xlsx")]
        if not xlsxs:
            continue
        wsi_dir = os.path.join(data_dir, sub, "wsi")
        wsi = []
        if os.path.isdir(wsi_dir):
            wsi = [os.path.join(wsi_dir, f) for f in os.listdir(wsi_dir) if f.lower().endswith(".svs")]
        out.append((pid, code, xlsxs[0], wsi))
    return out


# ────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────

def main(argv=None):
    ap = argparse.ArgumentParser(description="Ingest patient Excel workbooks → patients.json")
    ap.add_argument("--data-dir", default=DEFAULT_DATA_DIR)
    ap.add_argument("--output", default=DEFAULT_OUTPUT)
    ap.add_argument("--static-docs-dir", default=DEFAULT_STATIC_DOCS,
                    help="PDFs are copied under here; URLs reference /loki_dashboard_2/static/docs/...")
    ap.add_argument("--records-pid", default="4",
                    help="Patient id to attribute the De-identified Records folder to (flat layout).")
    ap.add_argument("--module-url-prefix", default="loki_dashboard_2/static/docs")
    ap.add_argument("--dzi-url-prefix", default="loki_dashboard_2/static/src/wsi/dzi")
    args = ap.parse_args(argv)

    data_dir = os.path.abspath(args.data_dir)
    if not os.path.isdir(data_dir):
        sys.stderr.write(f"ERROR: data-dir not found: {data_dir}\n")
        return 1

    # auto-detect layout
    flat = discover_flat(data_dir)
    nested = discover_nested(data_dir)
    if flat and not nested:
        layout = "flat"
        patients_meta = flat
    elif nested and not flat:
        layout = "nested"
        patients_meta = nested
    elif flat and nested:
        layout = "flat"  # prefer flat when both present
        patients_meta = flat
    else:
        sys.stderr.write(f"ERROR: no workbooks found under {data_dir}\n")
        return 1
    print(f"layout: {layout}, patients: {[p[1] for p in patients_meta]}", file=sys.stderr)

    static_docs_dir = os.path.abspath(args.static_docs_dir)
    os.makedirs(static_docs_dir, exist_ok=True)

    patients: list[dict[str, Any]] = []
    for pid, code, xlsx, wsi in patients_meta:
        try:
            p = parse_workbook(xlsx, pid, code)
        except Exception as exc:
            sys.stderr.write(f"WARNING: failed to parse {xlsx}: {exc}\n")
            continue

        attach_wsi_slides(p, wsi, args.dzi_url_prefix)

        # Documents: nested layout has per-patient records/<cat>/*.pdf
        if layout == "nested":
            recs_root = os.path.join(data_dir, f"Patient_{pid}", "records")
            pdfs = []
            if os.path.isdir(recs_root):
                for root, _dirs, files in os.walk(recs_root):
                    for f in files:
                        if f.lower().endswith(".pdf"):
                            pdfs.append(os.path.join(root, f))
            attach_documents(p, pdfs, static_docs_dir, args.module_url_prefix)

        if layout == "flat":
            base = os.path.dirname(xlsx)
            # source filenames vary: "Clinical Data_3.pdf", "Clinical Data _4.pdf", "Clinical Data_7- GB.pdf"
            def _norm(x: str) -> str:
                return re.sub(r"[\s_\-]+", "", x).lower()
            want = _norm(f"ClinicalData{pid}.pdf")
            pdfs = [os.path.join(base, f) for f in os.listdir(base)
                    if f.lower().endswith(".pdf") and _norm(f) == want]
            attach_documents(p, pdfs, static_docs_dir, args.module_url_prefix)

        patients.append(p)

    # Flat layout: attribute the De-identified Records folder to --records-pid
    if layout == "flat":
        di_dir = os.path.join(data_dir, "De-identified Records")
        if os.path.isdir(di_dir):
            target = next((p for p in patients if p["id"] == args.records_pid), None)
            if target is None:
                sys.stderr.write(f"WARNING: --records-pid {args.records_pid} not found; skipping De-identified Records\n")
            else:
                pdfs = [os.path.join(di_dir, f) for f in sorted(os.listdir(di_dir)) if f.lower().endswith(".pdf")]
                attach_documents(target, pdfs, static_docs_dir, args.module_url_prefix)
                print(f"attributed {len(pdfs)} De-identified Records → {target['code']}", file=sys.stderr)

    payload = {
        "schema_version": 1,
        "generated_at": dt.datetime.utcnow().isoformat(timespec="seconds"),
        "source_layout": layout,
        "patients": patients,
    }

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False, default=str)
    print(f"wrote {args.output}  ({len(patients)} patients)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
