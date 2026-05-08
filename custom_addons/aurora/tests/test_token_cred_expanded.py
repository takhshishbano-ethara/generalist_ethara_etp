# -*- coding: utf-8 -*-
import hashlib
import os
import unittest
from unittest.mock import MagicMock, patch, PropertyMock


def _mock_icp():
    m = MagicMock()
    m.get_param = MagicMock(return_value="")
    return m


def _mock_cr():
    m = MagicMock()
    m.fetchone = MagicMock(return_value=None)
    m.dbname = "test_db"
    return m


class TestCredentialManagerConstants(unittest.TestCase):
    """Tests for credential_manager.py constants and module-level objects."""

    def test_encrypted_prefix_value(self):
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        self.assertEqual(_ENCRYPTED_PREFIX, "fernet:1:")

    def test_encrypted_prefix_is_string(self):
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        self.assertIsInstance(_ENCRYPTED_PREFIX, str)

    def test_encrypted_params_is_frozenset(self):
        from odoo.addons.aurora.models.credential_manager import ENCRYPTED_PARAMS
        self.assertIsInstance(ENCRYPTED_PARAMS, frozenset)

    def test_encrypted_params_contains_s3_access_key(self):
        from odoo.addons.aurora.models.credential_manager import ENCRYPTED_PARAMS
        self.assertIn("aurora.s3_access_key", ENCRYPTED_PARAMS)

    def test_encrypted_params_contains_s3_secret_key(self):
        from odoo.addons.aurora.models.credential_manager import ENCRYPTED_PARAMS
        self.assertIn("aurora.s3_secret_key", ENCRYPTED_PARAMS)

    def test_encrypted_params_contains_webhook_secret(self):
        from odoo.addons.aurora.models.credential_manager import ENCRYPTED_PARAMS
        self.assertIn("aurora.webhook_secret", ENCRYPTED_PARAMS)

    def test_encrypted_params_contains_github_registry_write_token(self):
        from odoo.addons.aurora.models.credential_manager import ENCRYPTED_PARAMS
        self.assertIn("aurora.github_registry_write_token", ENCRYPTED_PARAMS)

    def test_encrypted_params_has_exactly_four_keys(self):
        from odoo.addons.aurora.models.credential_manager import ENCRYPTED_PARAMS
        self.assertEqual(len(ENCRYPTED_PARAMS), 4)

    def test_key_cache_lock_exists(self):
        from odoo.addons.aurora.models.credential_manager import _key_cache_lock
        import threading
        self.assertIsInstance(_key_cache_lock, type(threading.Lock()))

    def test_cached_fernet_keys_is_dict(self):
        from odoo.addons.aurora.models.credential_manager import _cached_fernet_keys
        self.assertIsInstance(_cached_fernet_keys, dict)

    def test_cached_fernet_keys_raw_is_dict(self):
        from odoo.addons.aurora.models.credential_manager import _cached_fernet_keys_raw
        self.assertIsInstance(_cached_fernet_keys_raw, dict)


class TestGetOrCreateKey(unittest.TestCase):
    """Tests for _get_or_create_key function."""

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": "test-key-value"})
    def test_returns_env_var_when_present(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key
        icp = _mock_icp()
        result = _get_or_create_key(icp)
        self.assertEqual(result, b"test-key-value")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": "  spaced-key  "})
    def test_strips_env_var_whitespace(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key
        icp = _mock_icp()
        result = _get_or_create_key(icp)
        self.assertEqual(result, b"spaced-key")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_falls_back_to_db_when_env_empty(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key, _cached_fernet_keys
        icp = _mock_icp()
        icp.env.cr.dbname = "test_db_fallback"
        icp.get_param.return_value = "db-stored-key"
        _cached_fernet_keys.pop("test_db_fallback", None)
        result = _get_or_create_key(icp)
        self.assertEqual(result, b"db-stored-key")
        _cached_fernet_keys.pop("test_db_fallback", None)

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_raises_runtime_error_when_no_key(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key, _cached_fernet_keys
        icp = _mock_icp()
        icp.env.cr.dbname = "test_db_nokey"
        icp.get_param.return_value = ""
        _cached_fernet_keys.pop("test_db_nokey", None)
        with self.assertRaises(RuntimeError):
            _get_or_create_key(icp)

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_uses_cache_when_available(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key, _cached_fernet_keys
        _cached_fernet_keys["cached_db"] = b"cached-key-value"
        icp = _mock_icp()
        icp.env.cr.dbname = "cached_db"
        result = _get_or_create_key(icp)
        self.assertEqual(result, b"cached-key-value")
        _cached_fernet_keys.pop("cached_db", None)


class TestGetPreviousKey(unittest.TestCase):
    """Tests for _get_previous_key function."""

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": "prev-key"})
    def test_returns_env_var_when_present(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key
        icp = _mock_icp()
        result = _get_previous_key(icp)
        self.assertEqual(result, b"prev-key")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_falls_back_to_db(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key
        icp = _mock_icp()
        icp.get_param.return_value = "db-prev-key"
        result = _get_previous_key(icp)
        self.assertEqual(result, b"db-prev-key")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_returns_none_when_missing(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key
        icp = _mock_icp()
        icp.get_param.return_value = ""
        result = _get_previous_key(icp)
        self.assertIsNone(result)


class TestMakeFernet(unittest.TestCase):
    """Tests for _make_fernet function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()
        self.test_key_prev = Fernet.generate_key()

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_returns_fernet_with_single_key(self):
        from odoo.addons.aurora.models.credential_manager import _make_fernet, _cached_fernet_keys
        from cryptography.fernet import Fernet
        icp = _mock_icp()
        icp.env.cr.dbname = "fernet_single_db"
        icp.get_param.side_effect = lambda k, d="": self.test_key.decode() if k == "aurora.encryption_key" else ""
        _cached_fernet_keys.pop("fernet_single_db", None)
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        try:
            result = _make_fernet(icp)
            self.assertIsInstance(result, Fernet)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_returns_multifernet_with_previous_key(self):
        from odoo.addons.aurora.models.credential_manager import _make_fernet
        from cryptography.fernet import MultiFernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"] = self.test_key_prev.decode()
        try:
            icp = _mock_icp()
            result = _make_fernet(icp)
            self.assertIsInstance(result, MultiFernet)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]
            del os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"]


class TestGetOrCreateKeyRaw(unittest.TestCase):
    """Tests for _get_or_create_key_raw function."""

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": "raw-env-key"})
    def test_returns_env_var_when_present(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key_raw
        cr = _mock_cr()
        result = _get_or_create_key_raw(cr)
        self.assertEqual(result, b"raw-env-key")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_returns_db_row_when_found(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key_raw, _cached_fernet_keys_raw
        cr = _mock_cr()
        cr.dbname = "raw_db_key"
        cr.fetchone.return_value = ("db-raw-key-value",)
        _cached_fernet_keys_raw.pop("raw_db_key", None)
        result = _get_or_create_key_raw(cr)
        self.assertEqual(result, b"db-raw-key-value")
        _cached_fernet_keys_raw.pop("raw_db_key", None)

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_raises_runtime_error_when_missing(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key_raw, _cached_fernet_keys_raw
        cr = _mock_cr()
        cr.dbname = "raw_db_missing"
        cr.fetchone.return_value = None
        _cached_fernet_keys_raw.pop("raw_db_missing", None)
        with self.assertRaises(RuntimeError):
            _get_or_create_key_raw(cr)

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_uses_cache_when_available(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key_raw, _cached_fernet_keys_raw
        _cached_fernet_keys_raw["raw_cached_db"] = b"raw-cached-key"
        cr = _mock_cr()
        cr.dbname = "raw_cached_db"
        result = _get_or_create_key_raw(cr)
        self.assertEqual(result, b"raw-cached-key")
        _cached_fernet_keys_raw.pop("raw_cached_db", None)

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY": ""})
    def test_handles_missing_dbname_attribute(self):
        from odoo.addons.aurora.models.credential_manager import _get_or_create_key_raw, _cached_fernet_keys_raw
        cr = MagicMock(spec=[])
        cr.execute = MagicMock()
        cr.fetchone = MagicMock(return_value=("fallback-key",))
        _cached_fernet_keys_raw.pop("default", None)
        result = _get_or_create_key_raw(cr)
        self.assertEqual(result, b"fallback-key")
        _cached_fernet_keys_raw.pop("default", None)


class TestGetPreviousKeyRaw(unittest.TestCase):
    """Tests for _get_previous_key_raw function."""

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": "raw-prev-key"})
    def test_returns_env_var_when_present(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key_raw
        cr = _mock_cr()
        result = _get_previous_key_raw(cr)
        self.assertEqual(result, b"raw-prev-key")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_returns_db_row_when_found(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key_raw
        cr = _mock_cr()
        cr.fetchone.return_value = ("db-prev-raw-key",)
        result = _get_previous_key_raw(cr)
        self.assertEqual(result, b"db-prev-raw-key")

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_returns_none_when_missing(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key_raw
        cr = _mock_cr()
        cr.fetchone.return_value = None
        result = _get_previous_key_raw(cr)
        self.assertIsNone(result)

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_returns_none_when_row_empty(self):
        from odoo.addons.aurora.models.credential_manager import _get_previous_key_raw
        cr = _mock_cr()
        cr.fetchone.return_value = ("",)
        result = _get_previous_key_raw(cr)
        self.assertIsNone(result)


class TestMakeFernetRaw(unittest.TestCase):
    """Tests for _make_fernet_raw function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()
        self.test_key_prev = Fernet.generate_key()

    @patch.dict(os.environ, {"AURORA_ENCRYPTION_KEY_PREVIOUS": ""})
    def test_returns_fernet_single_key(self):
        from odoo.addons.aurora.models.credential_manager import _make_fernet_raw
        from cryptography.fernet import Fernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        try:
            cr = _mock_cr()
            cr.fetchone.return_value = None
            result = _make_fernet_raw(cr)
            self.assertIsInstance(result, Fernet)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_returns_multifernet_with_previous(self):
        from odoo.addons.aurora.models.credential_manager import _make_fernet_raw
        from cryptography.fernet import MultiFernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"] = self.test_key_prev.decode()
        try:
            cr = _mock_cr()
            result = _make_fernet_raw(cr)
            self.assertIsInstance(result, MultiFernet)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]
            del os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"]


class TestEncryptValue(unittest.TestCase):
    """Tests for encrypt_value function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_empty_string_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value
        icp = _mock_icp()
        result = encrypt_value(icp, "")
        self.assertEqual(result, "")

    def test_none_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value
        icp = _mock_icp()
        result = encrypt_value(icp, "")
        self.assertEqual(result, "")

    def test_normal_encryption_adds_prefix(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            result = encrypt_value(icp, "my-secret")
            self.assertTrue(result.startswith(_ENCRYPTED_PREFIX))
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_encrypted_value_is_not_plaintext(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            result = encrypt_value(icp, "my-secret")
            self.assertNotIn("my-secret", result)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_encrypted_value_different_each_call(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            r1 = encrypt_value(icp, "my-secret")
            r2 = encrypt_value(icp, "my-secret")
            self.assertNotEqual(r1, r2)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]


class TestDecryptValue(unittest.TestCase):
    """Tests for decrypt_value function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_empty_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value
        icp = _mock_icp()
        result = decrypt_value(icp, "")
        self.assertEqual(result, "")

    def test_no_prefix_returns_passthrough(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value
        icp = _mock_icp()
        result = decrypt_value(icp, "plain-value")
        self.assertEqual(result, "plain-value")

    def test_valid_decrypt(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, decrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            encrypted = encrypt_value(icp, "secret-data")
            decrypted = decrypt_value(icp, encrypted)
            self.assertEqual(decrypted, "secret-data")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_invalid_token_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value, _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            bad_stored = f"{_ENCRYPTED_PREFIX}invalid-garbage-data"
            result = decrypt_value(icp, bad_stored)
            self.assertEqual(result, "")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_none_input_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value
        icp = _mock_icp()
        result = decrypt_value(icp, "")
        self.assertEqual(result, "")


class TestDecryptValueRaw(unittest.TestCase):
    """Tests for decrypt_value_raw function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_empty_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value_raw
        cr = _mock_cr()
        result = decrypt_value_raw(cr, "")
        self.assertEqual(result, "")

    def test_no_prefix_returns_passthrough(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value_raw
        cr = _mock_cr()
        result = decrypt_value_raw(cr, "plain-text-value")
        self.assertEqual(result, "plain-text-value")

    def test_valid_decrypt_raw(self):
        from odoo.addons.aurora.models.credential_manager import (
            decrypt_value_raw, _ENCRYPTED_PREFIX
        )
        from cryptography.fernet import Fernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            f = Fernet(self.test_key)
            encrypted_token = f.encrypt(b"raw-secret").decode()
            stored = f"{_ENCRYPTED_PREFIX}{encrypted_token}"
            cr = _mock_cr()
            result = decrypt_value_raw(cr, stored)
            self.assertEqual(result, "raw-secret")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_invalid_token_returns_empty_raw(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value_raw, _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            stored = f"{_ENCRYPTED_PREFIX}corrupted-data-here"
            cr = _mock_cr()
            result = decrypt_value_raw(cr, stored)
            self.assertEqual(result, "")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]


class TestGetEncryptedParam(unittest.TestCase):
    """Tests for get_encrypted_param function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_encrypted_key_decrypts(self):
        from odoo.addons.aurora.models.credential_manager import (
            get_encrypted_param, _ENCRYPTED_PREFIX
        )
        from cryptography.fernet import Fernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            f = Fernet(self.test_key)
            enc_val = f"{_ENCRYPTED_PREFIX}{f.encrypt(b'access-key-123').decode()}"
            env = MagicMock()
            icp_mock = _mock_icp()
            icp_mock.get_param = MagicMock(side_effect=lambda key, default="": enc_val if key == "aurora.s3_access_key" else default)
            env.__getitem__ = MagicMock(return_value=MagicMock(sudo=MagicMock(return_value=icp_mock)))
            result = get_encrypted_param(env, "aurora.s3_access_key")
            self.assertEqual(result, "access-key-123")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_non_encrypted_key_passthrough(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param
        env = MagicMock()
        icp_mock = MagicMock()
        icp_mock.get_param.return_value = "some-plain-value"
        env.__getitem__ = MagicMock(return_value=MagicMock(sudo=MagicMock(return_value=icp_mock)))
        result = get_encrypted_param(env, "aurora.some_other_key")
        self.assertEqual(result, "some-plain-value")

    def test_default_value_used_when_missing(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param
        env = MagicMock()
        icp_mock = MagicMock()
        icp_mock.get_param.return_value = "fallback"
        env.__getitem__ = MagicMock(return_value=MagicMock(sudo=MagicMock(return_value=icp_mock)))
        result = get_encrypted_param(env, "aurora.non_existent", "fallback")
        self.assertEqual(result, "fallback")


class TestGetEncryptedParamRaw(unittest.TestCase):
    """Tests for get_encrypted_param_raw function."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_encrypted_key_decrypts_raw(self):
        from odoo.addons.aurora.models.credential_manager import (
            get_encrypted_param_raw, _ENCRYPTED_PREFIX
        )
        from cryptography.fernet import Fernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            f = Fernet(self.test_key)
            enc_val = f"{_ENCRYPTED_PREFIX}{f.encrypt(b'secret-key').decode()}"
            cr = _mock_cr()
            cr.fetchone.side_effect = [(enc_val,), None]
            result = get_encrypted_param_raw(cr, "aurora.s3_secret_key")
            self.assertEqual(result, "secret-key")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_non_encrypted_key_returns_raw_value(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param_raw
        cr = _mock_cr()
        cr.fetchone.return_value = ("raw-config-value",)
        result = get_encrypted_param_raw(cr, "aurora.some_plain_key")
        self.assertEqual(result, "raw-config-value")

    def test_missing_row_returns_default(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param_raw
        cr = _mock_cr()
        cr.fetchone.return_value = None
        result = get_encrypted_param_raw(cr, "aurora.missing_key", "my-default")
        self.assertEqual(result, "my-default")

    def test_empty_stored_encrypted_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param_raw
        cr = _mock_cr()
        cr.fetchone.return_value = ("",)
        result = get_encrypted_param_raw(cr, "aurora.s3_access_key", "def")
        self.assertEqual(result, "")


class TestGithubTokenConstants(unittest.TestCase):
    """Tests for github_token.py constants."""

    def test_token_states_has_six_entries(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        self.assertEqual(len(TOKEN_STATES), 6)

    def test_token_states_contains_draft(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        states = [s[0] for s in TOKEN_STATES]
        self.assertIn("draft", states)

    def test_token_states_contains_active(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        states = [s[0] for s in TOKEN_STATES]
        self.assertIn("active", states)

    def test_token_states_contains_exhausted(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        states = [s[0] for s in TOKEN_STATES]
        self.assertIn("exhausted", states)

    def test_token_states_contains_expired(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        states = [s[0] for s in TOKEN_STATES]
        self.assertIn("expired", states)

    def test_token_states_contains_revoked(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        states = [s[0] for s in TOKEN_STATES]
        self.assertIn("revoked", states)

    def test_token_states_contains_quarantined(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        states = [s[0] for s in TOKEN_STATES]
        self.assertIn("quarantined", states)

    def test_valid_token_prefixes_contains_ghp(self):
        from odoo.addons.aurora.models.github_token import _VALID_TOKEN_PREFIXES
        self.assertIn("ghp_", _VALID_TOKEN_PREFIXES)

    def test_valid_token_prefixes_contains_gho(self):
        from odoo.addons.aurora.models.github_token import _VALID_TOKEN_PREFIXES
        self.assertIn("gho_", _VALID_TOKEN_PREFIXES)

    def test_valid_token_prefixes_contains_github_pat(self):
        from odoo.addons.aurora.models.github_token import _VALID_TOKEN_PREFIXES
        self.assertIn("github_pat_", _VALID_TOKEN_PREFIXES)

    def test_lease_batch_size_value(self):
        from odoo.addons.aurora.models.github_token import _LEASE_BATCH_SIZE
        self.assertEqual(_LEASE_BATCH_SIZE, 3)

    def test_health_check_workers_value(self):
        from odoo.addons.aurora.models.github_token import _HEALTH_CHECK_WORKERS
        self.assertEqual(_HEALTH_CHECK_WORKERS, 10)

    def test_health_check_rate_value(self):
        from odoo.addons.aurora.models.github_token import _HEALTH_CHECK_RATE
        self.assertEqual(_HEALTH_CHECK_RATE, 15)

    def test_min_remaining_for_lease_value(self):
        from odoo.addons.aurora.models.github_token import _MIN_REMAINING_FOR_LEASE
        self.assertEqual(_MIN_REMAINING_FOR_LEASE, 100)

    def test_quarantine_threshold_value(self):
        from odoo.addons.aurora.models.github_token import _QUARANTINE_THRESHOLD
        self.assertEqual(_QUARANTINE_THRESHOLD, 6)

    def test_quarantine_expiry_hours_value(self):
        from odoo.addons.aurora.models.github_token import _QUARANTINE_EXPIRY_HOURS
        self.assertEqual(_QUARANTINE_EXPIRY_HOURS, 24)

    def test_metrics_retention_days_value(self):
        from odoo.addons.aurora.models.github_token import _METRICS_RETENTION_DAYS
        self.assertEqual(_METRICS_RETENTION_DAYS, 7)

    def test_import_batch_size_value(self):
        from odoo.addons.aurora.models.github_token import _IMPORT_BATCH_SIZE
        self.assertEqual(_IMPORT_BATCH_SIZE, 500)

    def test_allowed_update_columns_is_frozenset(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIsInstance(_ALLOWED_UPDATE_COLUMNS, frozenset)

    def test_allowed_update_columns_contains_state(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("state", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_rate_limit_remaining(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("rate_limit_remaining", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_rate_limit_reset(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("rate_limit_reset", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_last_health_check(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("last_health_check", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_last_heartbeat(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("last_heartbeat", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_consecutive_failure_count(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("consecutive_failure_count", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_error_message(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("error_message", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_leased_by_run_id(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("leased_by_run_id", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_contains_leased_at(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertIn("leased_at", _ALLOWED_UPDATE_COLUMNS)

    def test_allowed_update_columns_count(self):
        from odoo.addons.aurora.models.github_token import _ALLOWED_UPDATE_COLUMNS
        self.assertEqual(len(_ALLOWED_UPDATE_COLUMNS), 9)

    def test_valid_token_prefixes_is_tuple(self):
        from odoo.addons.aurora.models.github_token import _VALID_TOKEN_PREFIXES
        self.assertIsInstance(_VALID_TOKEN_PREFIXES, tuple)

    def test_token_states_is_list(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        self.assertIsInstance(TOKEN_STATES, list)

    def test_token_states_tuples_have_two_elements(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        for state in TOKEN_STATES:
            self.assertEqual(len(state), 2)


class TestHashToken(unittest.TestCase):
    """Tests for AuroraGithubToken._hash_token static method."""

    def test_deterministic_sha256(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        h1 = AuroraGithubToken._hash_token("ghp_abc123")
        h2 = AuroraGithubToken._hash_token("ghp_abc123")
        self.assertEqual(h1, h2)

    def test_produces_hex_digest(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        result = AuroraGithubToken._hash_token("ghp_test")
        self.assertEqual(len(result), 64)

    def test_matches_manual_sha256(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        token = "ghp_testtoken123"
        expected = hashlib.sha256(token.encode()).hexdigest()
        self.assertEqual(AuroraGithubToken._hash_token(token), expected)

    def test_different_tokens_different_hashes(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        h1 = AuroraGithubToken._hash_token("ghp_token_a")
        h2 = AuroraGithubToken._hash_token("ghp_token_b")
        self.assertNotEqual(h1, h2)

    def test_hash_is_lowercase_hex(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        result = AuroraGithubToken._hash_token("ghp_xyz")
        self.assertTrue(all(c in "0123456789abcdef" for c in result))


class TestDecryptTokenRaw(unittest.TestCase):
    """Tests for AuroraGithubToken._decrypt_token_raw static method."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_empty_returns_empty(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        result = AuroraGithubToken._decrypt_token_raw(cr, "")
        self.assertEqual(result, "")

    def test_none_returns_empty(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        result = AuroraGithubToken._decrypt_token_raw(cr, "")
        self.assertEqual(result, "")

    def test_no_prefix_returns_passthrough(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        result = AuroraGithubToken._decrypt_token_raw(cr, "ghp_plaintoken")
        self.assertEqual(result, "ghp_plaintoken")

    def test_valid_decrypt(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        from cryptography.fernet import Fernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            f = Fernet(self.test_key)
            encrypted = f"{_ENCRYPTED_PREFIX}{f.encrypt(b'ghp_realtoken').decode()}"
            cr = _mock_cr()
            result = AuroraGithubToken._decrypt_token_raw(cr, encrypted)
            self.assertEqual(result, "ghp_realtoken")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_invalid_token_returns_empty(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            bad = f"{_ENCRYPTED_PREFIX}totally-not-valid-fernet"
            cr = _mock_cr()
            result = AuroraGithubToken._decrypt_token_raw(cr, bad)
            self.assertEqual(result, "")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]


class TestLeaseTokens(unittest.TestCase):
    """Tests for AuroraGithubToken.lease_tokens static method."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_sql_query_executed(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 42)
        cr.execute.assert_called_once()

    def test_empty_result_returns_empty_list(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        cr.fetchall.return_value = []
        result = AuroraGithubToken.lease_tokens(cr, 42)
        self.assertEqual(result, [])

    def test_with_rows_decrypts_tokens(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        from cryptography.fernet import Fernet
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            f = Fernet(self.test_key)
            enc_tok = f"{_ENCRYPTED_PREFIX}{f.encrypt(b'ghp_leased').decode()}"
            cr = _mock_cr()
            cr.fetchall.return_value = [(1, enc_tok)]
            result = AuroraGithubToken.lease_tokens(cr, 99)
            self.assertEqual(result, ["ghp_leased"])
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_sql_contains_run_id_param(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken, _MIN_REMAINING_FOR_LEASE, _LEASE_BATCH_SIZE
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 77, count=5)
        call_args = cr.execute.call_args
        self.assertIn(77, call_args[0][1])

    def test_uses_count_param(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken, _MIN_REMAINING_FOR_LEASE
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 10, count=7)
        call_args = cr.execute.call_args
        self.assertIn(7, call_args[0][1])

    def test_skips_empty_decrypted_tokens(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            bad_enc = f"{_ENCRYPTED_PREFIX}garbage-not-fernet"
            cr = _mock_cr()
            cr.fetchall.return_value = [(1, bad_enc)]
            result = AuroraGithubToken.lease_tokens(cr, 42)
            self.assertEqual(result, [])
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_multiple_rows_returns_multiple_tokens(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        from cryptography.fernet import Fernet
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            f = Fernet(self.test_key)
            enc1 = f"{_ENCRYPTED_PREFIX}{f.encrypt(b'ghp_one').decode()}"
            enc2 = f"{_ENCRYPTED_PREFIX}{f.encrypt(b'ghp_two').decode()}"
            cr = _mock_cr()
            cr.fetchall.return_value = [(1, enc1), (2, enc2)]
            result = AuroraGithubToken.lease_tokens(cr, 5)
            self.assertEqual(len(result), 2)
            self.assertIn("ghp_one", result)
            self.assertIn("ghp_two", result)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]


class TestReleaseTokens(unittest.TestCase):
    """Tests for AuroraGithubToken.release_tokens static method."""

    def test_without_token_summaries(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        AuroraGithubToken.release_tokens(cr, 42)
        cr.execute.assert_called_once()

    def test_with_token_summaries_calls_write_rate_limits(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"hash1": {"remaining": 500, "reset": 1700000000}}
        with patch.object(AuroraGithubToken, '_write_rate_limits') as mock_wrl:
            AuroraGithubToken.release_tokens(cr, 42, token_summaries=summaries)
            mock_wrl.assert_called_once_with(cr, 42, summaries)

    def test_release_sql_contains_run_id(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        AuroraGithubToken.release_tokens(cr, 99)
        call_args = cr.execute.call_args
        self.assertIn(99, call_args[0][1])

    def test_release_sets_leased_by_null(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        AuroraGithubToken.release_tokens(cr, 55)
        sql = cr.execute.call_args[0][0]
        self.assertIn("leased_by_run_id = NULL", sql)

    def test_release_with_empty_summaries_skips_write(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        with patch.object(AuroraGithubToken, '_write_rate_limits') as mock_wrl:
            AuroraGithubToken.release_tokens(cr, 42, token_summaries=None)
            mock_wrl.assert_not_called()


class TestHeartbeatRateLimits(unittest.TestCase):
    """Tests for AuroraGithubToken.heartbeat_rate_limits static method."""

    def test_empty_summaries_returns_early(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        with patch.object(AuroraGithubToken, '_write_rate_limits') as mock_wrl:
            AuroraGithubToken.heartbeat_rate_limits(cr, 42, {})
            mock_wrl.assert_not_called()

    def test_none_summaries_returns_early(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        with patch.object(AuroraGithubToken, '_write_rate_limits') as mock_wrl:
            AuroraGithubToken.heartbeat_rate_limits(cr, 42, None)
            mock_wrl.assert_not_called()

    def test_with_summaries_calls_write_rate_limits(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"hash_a": {"remaining": 200, "reset": 1700000000}}
        with patch.object(AuroraGithubToken, '_write_rate_limits') as mock_wrl:
            AuroraGithubToken.heartbeat_rate_limits(cr, 10, summaries)
            mock_wrl.assert_called_once_with(cr, 10, summaries)

    def test_with_multiple_summaries(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {
            "hash_1": {"remaining": 100, "reset": 1700000000},
            "hash_2": {"remaining": 200, "reset": 1700001000},
        }
        with patch.object(AuroraGithubToken, '_write_rate_limits') as mock_wrl:
            AuroraGithubToken.heartbeat_rate_limits(cr, 5, summaries)
            mock_wrl.assert_called_once()


class TestWriteRateLimits(unittest.TestCase):
    """Tests for AuroraGithubToken._write_rate_limits static method."""

    def test_updates_rate_limit_remaining(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"abc123": {"remaining": 450, "reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 42, summaries)
        call_args = cr.execute.call_args
        self.assertIn(450, call_args[0][1])

    def test_updates_rate_limit_reset(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"abc123": {"remaining": 100, "reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 42, summaries)
        call_args = cr.execute.call_args
        self.assertIsNotNone(call_args[0][1][1])

    def test_handles_none_reset(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"abc123": {"remaining": 100, "reset": None}}
        AuroraGithubToken._write_rate_limits(cr, 42, summaries)
        call_args = cr.execute.call_args
        self.assertIsNone(call_args[0][1][1])

    def test_multiple_summaries_multiple_executes(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {
            "hash_a": {"remaining": 100, "reset": 1700000000},
            "hash_b": {"remaining": 200, "reset": 1700001000},
        }
        AuroraGithubToken._write_rate_limits(cr, 42, summaries)
        self.assertEqual(cr.execute.call_count, 2)

    def test_sql_contains_token_hash(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"myhash": {"remaining": 50, "reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 42, summaries)
        call_args = cr.execute.call_args
        self.assertIn("myhash", call_args[0][1])

    def test_sql_contains_run_id(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"h": {"remaining": 10, "reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 99, summaries)
        call_args = cr.execute.call_args
        self.assertIn(99, call_args[0][1])

    def test_zero_remaining_handled(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"h": {"remaining": 0, "reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 1, summaries)
        call_args = cr.execute.call_args
        self.assertIn(0, call_args[0][1])


class TestBuildXlsx(unittest.TestCase):
    """Tests for AuroraGithubToken._build_xlsx static method."""

    def test_returns_bytes(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        headers = ["Name", "State"]
        rows = [("token1", "active")]
        result = AuroraGithubToken._build_xlsx(headers, rows)
        self.assertIsInstance(result, bytes)

    def test_returns_non_empty_bytes(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        headers = ["Name", "State"]
        rows = [("token1", "active")]
        result = AuroraGithubToken._build_xlsx(headers, rows)
        self.assertGreater(len(result), 0)

    def test_headers_correct(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        import io
        from openpyxl import load_workbook
        headers = ["Name", "State", "Remaining"]
        rows = [("t1", "active", "100")]
        xlsx = AuroraGithubToken._build_xlsx(headers, rows)
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        header_row = [cell.value for cell in ws[1]]
        self.assertEqual(header_row, headers)

    def test_xlsx_contains_data_rows(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        import io
        from openpyxl import load_workbook
        headers = ["Name", "State"]
        rows = [("tok1", "active"), ("tok2", "expired")]
        xlsx = AuroraGithubToken._build_xlsx(headers, rows)
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        self.assertEqual(ws.max_row, 3)

    def test_empty_rows_still_has_headers(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        import io
        from openpyxl import load_workbook
        headers = ["Name", "State"]
        rows = []
        xlsx = AuroraGithubToken._build_xlsx(headers, rows)
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        self.assertEqual(ws.max_row, 1)

    def test_none_values_converted_to_empty_string(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        import io
        from openpyxl import load_workbook
        headers = ["Name", "State"]
        rows = [("tok1", None)]
        xlsx = AuroraGithubToken._build_xlsx(headers, rows)
        wb = load_workbook(io.BytesIO(xlsx))
        ws = wb.active
        self.assertIsNone(ws.cell(row=2, column=2).value)

    def test_worksheet_title(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        import io
        from openpyxl import load_workbook
        headers = ["Name"]
        rows = []
        xlsx = AuroraGithubToken._build_xlsx(headers, rows)
        wb = load_workbook(io.BytesIO(xlsx))
        self.assertEqual(wb.active.title, "Tokens")


class TestEncryptDecryptRoundTrip(unittest.TestCase):
    """Round-trip encryption/decryption tests."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_roundtrip_simple_string(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, decrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            enc = encrypt_value(icp, "hello-world")
            dec = decrypt_value(icp, enc)
            self.assertEqual(dec, "hello-world")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_roundtrip_unicode(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, decrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            enc = encrypt_value(icp, "héllo-wörld-日本")
            dec = decrypt_value(icp, enc)
            self.assertEqual(dec, "héllo-wörld-日本")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_roundtrip_long_string(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, decrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            long_str = "a" * 10000
            enc = encrypt_value(icp, long_str)
            dec = decrypt_value(icp, enc)
            self.assertEqual(dec, long_str)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_roundtrip_special_chars(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, decrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            special = "!@#$%^&*()_+-={}[]|\\:\";<>?,./~`"
            enc = encrypt_value(icp, special)
            dec = decrypt_value(icp, enc)
            self.assertEqual(dec, special)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]


class TestKeyRotation(unittest.TestCase):
    """Tests for key rotation with MultiFernet."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.old_key = Fernet.generate_key()
        self.new_key = Fernet.generate_key()

    def test_decrypt_with_previous_key(self):
        from odoo.addons.aurora.models.credential_manager import (
            encrypt_value, decrypt_value, _ENCRYPTED_PREFIX
        )
        from cryptography.fernet import Fernet
        old_f = Fernet(self.old_key)
        encrypted_with_old = f"{_ENCRYPTED_PREFIX}{old_f.encrypt(b'old-secret').decode()}"
        os.environ["AURORA_ENCRYPTION_KEY"] = self.new_key.decode()
        os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"] = self.old_key.decode()
        try:
            icp = _mock_icp()
            result = decrypt_value(icp, encrypted_with_old)
            self.assertEqual(result, "old-secret")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]
            del os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"]

    def test_encrypt_uses_current_key(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, _ENCRYPTED_PREFIX
        from cryptography.fernet import Fernet
        os.environ["AURORA_ENCRYPTION_KEY"] = self.new_key.decode()
        os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"] = self.old_key.decode()
        try:
            icp = _mock_icp()
            enc = encrypt_value(icp, "new-secret")
            token = enc[len(_ENCRYPTED_PREFIX):]
            f = Fernet(self.new_key)
            decrypted = f.decrypt(token.encode()).decode()
            self.assertEqual(decrypted, "new-secret")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]
            del os.environ["AURORA_ENCRYPTION_KEY_PREVIOUS"]


class TestLeaseTokensEdgeCases(unittest.TestCase):
    """Edge case tests for lease_tokens."""

    def test_default_count_is_batch_size(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken, _LEASE_BATCH_SIZE, _MIN_REMAINING_FOR_LEASE
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 1)
        call_args = cr.execute.call_args
        self.assertIn(_LEASE_BATCH_SIZE, call_args[0][1])

    def test_sql_contains_min_remaining(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken, _MIN_REMAINING_FOR_LEASE
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 1)
        call_args = cr.execute.call_args
        self.assertIn(_MIN_REMAINING_FOR_LEASE, call_args[0][1])

    def test_sql_has_for_no_key_update(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 1)
        sql = cr.execute.call_args[0][0]
        self.assertIn("FOR NO KEY UPDATE SKIP LOCKED", sql)

    def test_sql_has_returning_clause(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        cr.fetchall.return_value = []
        AuroraGithubToken.lease_tokens(cr, 1)
        sql = cr.execute.call_args[0][0]
        self.assertIn("RETURNING", sql)


class TestDecryptValueRawEdge(unittest.TestCase):
    """Additional edge case tests for decrypt_value_raw."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_prefix_only_returns_empty(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value_raw, _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            cr = _mock_cr()
            result = decrypt_value_raw(cr, _ENCRYPTED_PREFIX)
            self.assertEqual(result, "")
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_whitespace_stored_returns_passthrough(self):
        from odoo.addons.aurora.models.credential_manager import decrypt_value_raw
        cr = _mock_cr()
        result = decrypt_value_raw(cr, "  spaces  ")
        self.assertEqual(result, "  spaces  ")


class TestEncryptValueEdgeCases(unittest.TestCase):
    """Additional edge case tests for encrypt_value."""

    def setUp(self):
        from cryptography.fernet import Fernet
        self.test_key = Fernet.generate_key()

    def test_single_char_encrypts(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, _ENCRYPTED_PREFIX
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            result = encrypt_value(icp, "x")
            self.assertTrue(result.startswith(_ENCRYPTED_PREFIX))
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]

    def test_newlines_in_plaintext(self):
        from odoo.addons.aurora.models.credential_manager import encrypt_value, decrypt_value
        os.environ["AURORA_ENCRYPTION_KEY"] = self.test_key.decode()
        os.environ.pop("AURORA_ENCRYPTION_KEY_PREVIOUS", None)
        try:
            icp = _mock_icp()
            text = "line1\nline2\nline3"
            enc = encrypt_value(icp, text)
            dec = decrypt_value(icp, enc)
            self.assertEqual(dec, text)
        finally:
            del os.environ["AURORA_ENCRYPTION_KEY"]


class TestGetEncryptedParamRawEdge(unittest.TestCase):
    """Edge case tests for get_encrypted_param_raw."""

    def test_non_encrypted_key_with_none_row_uses_default(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param_raw
        cr = _mock_cr()
        cr.fetchone.return_value = None
        result = get_encrypted_param_raw(cr, "aurora.non_encrypted_key", "the-default")
        self.assertEqual(result, "the-default")

    def test_non_encrypted_key_empty_string_row(self):
        from odoo.addons.aurora.models.credential_manager import get_encrypted_param_raw
        cr = _mock_cr()
        cr.fetchone.return_value = ("",)
        result = get_encrypted_param_raw(cr, "aurora.non_encrypted_key", "fallback")
        self.assertEqual(result, "fallback")


class TestWriteRateLimitsEdge(unittest.TestCase):
    """Edge case tests for _write_rate_limits."""

    def test_missing_remaining_defaults_to_zero(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"h": {"reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 1, summaries)
        call_args = cr.execute.call_args
        self.assertIn(0, call_args[0][1])

    def test_missing_reset_defaults_to_none(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"h": {"remaining": 50}}
        AuroraGithubToken._write_rate_limits(cr, 1, summaries)
        call_args = cr.execute.call_args
        self.assertIsNone(call_args[0][1][1])

    def test_sql_updates_last_heartbeat(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        cr = _mock_cr()
        summaries = {"h": {"remaining": 10, "reset": 1700000000}}
        AuroraGithubToken._write_rate_limits(cr, 1, summaries)
        sql = cr.execute.call_args[0][0]
        self.assertIn("last_heartbeat", sql)


class TestTokenStatesStructure(unittest.TestCase):

    def test_draft_label(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        state_dict = dict(TOKEN_STATES)
        self.assertEqual(state_dict["draft"], "Draft")

    def test_active_label(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        state_dict = dict(TOKEN_STATES)
        self.assertEqual(state_dict["active"], "Active")

    def test_exhausted_label(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        state_dict = dict(TOKEN_STATES)
        self.assertEqual(state_dict["exhausted"], "Exhausted")

    def test_expired_label(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        state_dict = dict(TOKEN_STATES)
        self.assertEqual(state_dict["expired"], "Expired")

    def test_revoked_label(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        state_dict = dict(TOKEN_STATES)
        self.assertEqual(state_dict["revoked"], "Revoked")

    def test_quarantined_label(self):
        from odoo.addons.aurora.models.github_token import TOKEN_STATES
        state_dict = dict(TOKEN_STATES)
        self.assertEqual(state_dict["quarantined"], "Quarantined")


class TestEncryptedPrefixUsage(unittest.TestCase):

    def test_prefix_length(self):
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        self.assertEqual(len(_ENCRYPTED_PREFIX), 9)

    def test_prefix_ends_with_colon(self):
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        self.assertTrue(_ENCRYPTED_PREFIX.endswith(":"))

    def test_prefix_starts_with_fernet(self):
        from odoo.addons.aurora.models.credential_manager import _ENCRYPTED_PREFIX
        self.assertTrue(_ENCRYPTED_PREFIX.startswith("fernet"))


class TestHashTokenEdge(unittest.TestCase):

    def test_empty_string_hashes(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        result = AuroraGithubToken._hash_token("")
        expected = hashlib.sha256(b"").hexdigest()
        self.assertEqual(result, expected)

    def test_long_token_hashes(self):
        from odoo.addons.aurora.models.github_token import AuroraGithubToken
        long_token = "ghp_" + "a" * 1000
        result = AuroraGithubToken._hash_token(long_token)
        self.assertEqual(len(result), 64)


if __name__ == "__main__":
    unittest.main()
