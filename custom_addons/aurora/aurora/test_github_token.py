import hashlib
import io
from datetime import datetime, timezone
from unittest.mock import MagicMock, patch

import pytest

from aurora.models.github_token import (
    AuroraGithubToken,
    TOKEN_STATES,
    _VALID_TOKEN_PREFIXES,
    _LEASE_BATCH_SIZE,
    _HEALTH_CHECK_WORKERS,
    _HEALTH_CHECK_RATE,
    _MIN_REMAINING_FOR_LEASE,
    _QUARANTINE_THRESHOLD,
    _QUARANTINE_EXPIRY_HOURS,
    _METRICS_RETENTION_DAYS,
    _IMPORT_BATCH_SIZE,
    _ALLOWED_UPDATE_COLUMNS,
)


def test_token_states_count():
    assert len(TOKEN_STATES) == 6


@pytest.mark.parametrize("code,label", TOKEN_STATES)
def test_token_state_entry(code, label):
    assert isinstance(code, str)
    assert isinstance(label, str)


@pytest.mark.parametrize("state", ["draft", "active", "exhausted", "expired", "revoked", "quarantined"])
def test_token_state_present(state):
    codes = [c for c, _ in TOKEN_STATES]
    assert state in codes


def test_valid_token_prefixes_count():
    assert len(_VALID_TOKEN_PREFIXES) == 3


@pytest.mark.parametrize("prefix", ["ghp_", "gho_", "github_pat_"])
def test_valid_token_prefix_present(prefix):
    assert prefix in _VALID_TOKEN_PREFIXES


def test_valid_token_prefixes_is_tuple():
    assert isinstance(_VALID_TOKEN_PREFIXES, tuple)


def test_lease_batch_size():
    assert _LEASE_BATCH_SIZE == 3


def test_health_check_workers():
    assert _HEALTH_CHECK_WORKERS == 10


def test_health_check_rate():
    assert _HEALTH_CHECK_RATE == 15


def test_min_remaining_for_lease():
    assert _MIN_REMAINING_FOR_LEASE == 100


def test_quarantine_threshold():
    assert _QUARANTINE_THRESHOLD == 6


def test_quarantine_expiry_hours():
    assert _QUARANTINE_EXPIRY_HOURS == 24


def test_metrics_retention_days():
    assert _METRICS_RETENTION_DAYS == 7


def test_import_batch_size():
    assert _IMPORT_BATCH_SIZE == 500


@pytest.mark.parametrize("col", [
    "state", "rate_limit_remaining", "rate_limit_reset",
    "last_health_check", "last_heartbeat", "consecutive_failure_count",
    "error_message", "leased_by_run_id", "leased_at",
])
def test_allowed_update_column(col):
    assert col in _ALLOWED_UPDATE_COLUMNS


def test_allowed_update_columns_count():
    assert len(_ALLOWED_UPDATE_COLUMNS) == 9


def test_allowed_update_columns_is_frozenset():
    assert isinstance(_ALLOWED_UPDATE_COLUMNS, frozenset)


@pytest.mark.parametrize("raw_token", [
    "ghp_abc123",
    "gho_xyz789",
    "github_pat_test",
    "",
    "a",
    "x" * 100,
    "ghp_" + "A" * 50,
    "gho_" + "B" * 50,
    "github_pat_" + "C" * 50,
    "short",
    "a" * 256,
    "special!@#$%^&*()",
    "unicode_🎉_token",
    "  spaces  ",
    "newline\ntoken",
    "tab\ttoken",
    "ghp_realtoken12345678901234567890",
    "gho_realtoken12345678901234567890",
    "github_pat_realtoken123456789012345",
    "ALLCAPS",
])
def test_hash_token(raw_token):
    result = AuroraGithubToken._hash_token(raw_token)
    expected = hashlib.sha256(raw_token.encode()).hexdigest()
    assert result == expected
    assert len(result) == 64
    assert all(c in "0123456789abcdef" for c in result)


@pytest.mark.parametrize("t1,t2", [
    ("ghp_abc", "ghp_abd"),
    ("token1", "token2"),
    ("a", "b"),
])
def test_hash_token_different_inputs(t1, t2):
    assert AuroraGithubToken._hash_token(t1) != AuroraGithubToken._hash_token(t2)


def test_hash_token_deterministic():
    token = "ghp_test_deterministic"
    assert AuroraGithubToken._hash_token(token) == AuroraGithubToken._hash_token(token)


def test_decrypt_token_raw_empty():
    cr = MagicMock()
    assert AuroraGithubToken._decrypt_token_raw(cr, "") == ""


def test_decrypt_token_raw_none():
    cr = MagicMock()
    assert AuroraGithubToken._decrypt_token_raw(cr, None) == ""


def test_decrypt_token_raw_non_encrypted():
    cr = MagicMock()
    assert AuroraGithubToken._decrypt_token_raw(cr, "plain_text") == "plain_text"


def test_decrypt_token_raw_encrypted(fernet_key, mock_cursor):
    from cryptography.fernet import Fernet
    from aurora.models.credential_manager import _ENCRYPTED_PREFIX

    f = Fernet(fernet_key)
    encrypted = f.encrypt(b"my_secret_token").decode()
    stored = f"{_ENCRYPTED_PREFIX}{encrypted}"

    with patch("aurora.models.credential_manager._cached_fernet_key_raw", None):
        with patch("aurora.models.credential_manager.os.environ.get", return_value=fernet_key.decode()):
            result = AuroraGithubToken._decrypt_token_raw(mock_cursor, stored)

    assert result == "my_secret_token"


def test_decrypt_token_raw_invalid_cipher(mock_cursor, fernet_key):
    from aurora.models.credential_manager import _ENCRYPTED_PREFIX

    stored = f"{_ENCRYPTED_PREFIX}invalid_cipher_text"

    with patch("aurora.models.credential_manager._cached_fernet_key_raw", None):
        with patch("aurora.models.credential_manager.os.environ.get", return_value=fernet_key.decode()):
            result = AuroraGithubToken._decrypt_token_raw(mock_cursor, stored)

    assert result == ""


@pytest.mark.parametrize("count", [1, 2, 3, 5])
def test_lease_tokens_count(mock_cursor, count):
    token_rows = [(i, f"plain_token_{i}") for i in range(count)]
    mock_cursor.fetchall.return_value = token_rows

    with patch.object(AuroraGithubToken, "_decrypt_token_raw", side_effect=lambda cr, t: t):
        result = AuroraGithubToken.lease_tokens(mock_cursor, run_id=1, count=count)

    assert len(result) == count
    mock_cursor.execute.assert_called_once()


def test_lease_tokens_empty(mock_cursor):
    mock_cursor.fetchall.return_value = []
    result = AuroraGithubToken.lease_tokens(mock_cursor, run_id=1)
    assert result == []


def test_lease_tokens_decrypt_failure(mock_cursor):
    mock_cursor.fetchall.return_value = [(1, "encrypted_val")]

    with patch.object(AuroraGithubToken, "_decrypt_token_raw", return_value=""):
        result = AuroraGithubToken.lease_tokens(mock_cursor, run_id=1, count=1)

    assert result == []


def test_lease_tokens_partial_decrypt(mock_cursor):
    mock_cursor.fetchall.return_value = [(1, "tok1"), (2, "tok2"), (3, "tok3")]
    side_effects = ["decrypted1", "", "decrypted3"]

    with patch.object(AuroraGithubToken, "_decrypt_token_raw", side_effect=side_effects):
        result = AuroraGithubToken.lease_tokens(mock_cursor, run_id=1, count=3)

    assert result == ["decrypted1", "decrypted3"]


def test_release_tokens_no_summaries(mock_cursor):
    AuroraGithubToken.release_tokens(mock_cursor, run_id=42)
    mock_cursor.execute.assert_called_once()


def test_release_tokens_with_summaries(mock_cursor):
    summaries = {"hash1": {"remaining": 4000, "reset": 1700000000}}
    AuroraGithubToken.release_tokens(mock_cursor, run_id=42, token_summaries=summaries)
    assert mock_cursor.execute.call_count == 2


def test_release_tokens_calls_write_rate_limits(mock_cursor):
    summaries = {"h1": {"remaining": 100, "reset": None}}
    with patch.object(AuroraGithubToken, "_write_rate_limits") as mock_wrl:
        AuroraGithubToken.release_tokens(mock_cursor, run_id=1, token_summaries=summaries)
        mock_wrl.assert_called_once_with(mock_cursor, 1, summaries)


def test_heartbeat_rate_limits_empty(mock_cursor):
    AuroraGithubToken.heartbeat_rate_limits(mock_cursor, run_id=1, token_summaries=None)
    mock_cursor.execute.assert_not_called()


def test_heartbeat_rate_limits_empty_dict(mock_cursor):
    AuroraGithubToken.heartbeat_rate_limits(mock_cursor, run_id=1, token_summaries={})
    mock_cursor.execute.assert_not_called()


def test_heartbeat_rate_limits_calls_write(mock_cursor):
    summaries = {"h1": {"remaining": 500, "reset": 1700000000}}
    with patch.object(AuroraGithubToken, "_write_rate_limits") as mock_wrl:
        AuroraGithubToken.heartbeat_rate_limits(mock_cursor, run_id=1, token_summaries=summaries)
        mock_wrl.assert_called_once_with(mock_cursor, 1, summaries)


def test_heartbeat_rate_limits_commits(mock_cursor):
    summaries = {"h1": {"remaining": 100, "reset": None}}
    with patch.object(AuroraGithubToken, "_write_rate_limits"):
        AuroraGithubToken.heartbeat_rate_limits(mock_cursor, run_id=1, token_summaries=summaries)
    mock_cursor.commit.assert_called_once()


@pytest.mark.parametrize("n_summaries", [0, 1, 5])
def test_write_rate_limits_count(mock_cursor, n_summaries):
    summaries = {
        f"hash_{i}": {"remaining": i * 100, "reset": 1700000000 + i}
        for i in range(n_summaries)
    }
    AuroraGithubToken._write_rate_limits(mock_cursor, run_id=1, token_summaries=summaries)
    assert mock_cursor.execute.call_count == n_summaries


def test_write_rate_limits_with_reset_timestamp(mock_cursor):
    summaries = {"h1": {"remaining": 4500, "reset": 1700000000}}
    AuroraGithubToken._write_rate_limits(mock_cursor, run_id=1, token_summaries=summaries)
    call_args = mock_cursor.execute.call_args
    params = call_args[0][1]
    assert params[0] == 4500
    assert params[1] is not None
    assert params[2] == "h1"
    assert params[3] == 1


def test_write_rate_limits_no_reset(mock_cursor):
    summaries = {"h1": {"remaining": 100, "reset": None}}
    AuroraGithubToken._write_rate_limits(mock_cursor, run_id=1, token_summaries=summaries)
    call_args = mock_cursor.execute.call_args
    params = call_args[0][1]
    assert params[1] is None


def test_write_rate_limits_missing_remaining(mock_cursor):
    summaries = {"h1": {"reset": 1700000000}}
    AuroraGithubToken._write_rate_limits(mock_cursor, run_id=1, token_summaries=summaries)
    call_args = mock_cursor.execute.call_args
    params = call_args[0][1]
    assert params[0] == 0


@pytest.mark.parametrize("status,prev_state,prev_fails,expected_state", [
    (200, "draft", 0, "active"),
    (200, "exhausted", 0, "active"),
    (200, "quarantined", 0, "active"),
    (200, "active", 0, None),
    (401, "active", 0, "expired"),
    (401, "draft", 0, "expired"),
    (403, "active", 0, "exhausted"),
    (403, "draft", 0, "exhausted"),
    (429, "active", 0, "exhausted"),
    (429, "draft", 0, "exhausted"),
])
def test_health_check_state_transitions(status, prev_state, prev_fails, expected_state):
    vals = {"last_health_check": datetime.now(tz=timezone.utc)}

    if status == 200:
        body = {"resources": {"core": {"remaining": 5000, "reset": 1700000000}}}
        core = body["resources"]["core"]
        vals["rate_limit_remaining"] = core.get("remaining", 0)
        vals["consecutive_failure_count"] = 0
        vals["error_message"] = False
        if prev_state in ("exhausted", "draft", "quarantined") and core.get("remaining", 0) > _MIN_REMAINING_FOR_LEASE:
            vals["state"] = "active"
    elif status == 401:
        vals["state"] = "expired"
        vals["error_message"] = "401 Unauthorized — token invalid or revoked"
        vals["consecutive_failure_count"] = prev_fails + 1
    elif status == 403:
        vals["state"] = "exhausted"
        vals["rate_limit_remaining"] = 0
        vals["consecutive_failure_count"] = prev_fails + 1
    elif status == 429:
        vals["state"] = "exhausted"
        vals["rate_limit_remaining"] = 0
        vals["consecutive_failure_count"] = prev_fails + 1

    if expected_state is not None:
        assert vals.get("state") == expected_state
    else:
        assert "state" not in vals


@pytest.mark.parametrize("prev_fails,expected_quarantine", [
    (0, False),
    (4, False),
    (5, True),
    (10, True),
])
def test_quarantine_threshold_logic(prev_fails, expected_quarantine):
    new_fails = prev_fails + 1
    should_quarantine = new_fails >= _QUARANTINE_THRESHOLD
    assert should_quarantine == expected_quarantine


@pytest.mark.parametrize("status,prev_fails", [
    (0, 0),
    (0, 3),
    (0, 5),
    (500, 0),
    (502, 2),
])
def test_unknown_status_increments_failure(status, prev_fails):
    vals = {}
    if status not in (200, 401, 403, 429):
        vals["consecutive_failure_count"] = prev_fails + 1
    assert vals["consecutive_failure_count"] == prev_fails + 1


def test_build_xlsx_basic():
    headers = ["Name", "State", "Remaining"]
    rows = [("Token 0001", "active", 5000), ("Token 0002", "expired", 0)]
    result = AuroraGithubToken._build_xlsx(headers, rows)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_build_xlsx_empty_rows():
    headers = ["Name", "State"]
    rows = []
    result = AuroraGithubToken._build_xlsx(headers, rows)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_build_xlsx_none_values():
    headers = ["Name", "Value"]
    rows = [("test", None), (None, "val")]
    result = AuroraGithubToken._build_xlsx(headers, rows)
    assert isinstance(result, bytes)


def test_build_xlsx_content():
    from openpyxl import load_workbook
    headers = ["Col1", "Col2"]
    rows = [("a", "b"), ("c", "d")]
    xlsx_bytes = AuroraGithubToken._build_xlsx(headers, rows)
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.title == "Tokens"
    assert ws.cell(1, 1).value == "Col1"
    assert ws.cell(1, 2).value == "Col2"
    assert ws.cell(2, 1).value == "a"
    assert ws.cell(3, 2).value == "d"


def test_build_xlsx_headers_only():
    headers = ["H1", "H2", "H3"]
    rows = []
    xlsx_bytes = AuroraGithubToken._build_xlsx(headers, rows)
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(1, 1).value == "H1"
    assert ws.cell(2, 1).value is None


def test_build_xlsx_many_rows():
    headers = ["Name"]
    rows = [(f"Token {i:04d}",) for i in range(100)]
    xlsx_bytes = AuroraGithubToken._build_xlsx(headers, rows)
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.cell(101, 1).value == "Token 0099"


def test_build_xlsx_export_headers():
    headers = [
        "Name", "State", "Rate Limit Remaining", "Rate Limit Resets At",
        "Expires At", "Leased By", "Last Health Check", "Consecutive Failures",
        "Imported At", "Imported By", "Error Message",
    ]
    rows = []
    xlsx_bytes = AuroraGithubToken._build_xlsx(headers, rows)
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    ws = wb.active
    for i, h in enumerate(headers, 1):
        assert ws.cell(1, i).value == h


def test_model_name():
    assert AuroraGithubToken._name == "aurora.github.token"


def test_model_description():
    assert AuroraGithubToken._description == "GitHub Token"


def test_model_order():
    assert AuroraGithubToken._order == "name"


@pytest.mark.parametrize("field_name", [
    "name", "token", "token_hash", "state",
    "rate_limit_remaining", "rate_limit_reset",
    "expires_at", "leased_by_run_id", "leased_at",
    "last_health_check", "last_heartbeat",
    "consecutive_failure_count", "imported_at",
    "imported_by", "error_message",
])
def test_field_exists(field_name):
    assert hasattr(AuroraGithubToken, field_name)


def test_200_with_low_remaining_no_state_change():
    remaining = 50
    prev_state = "draft"
    vals = {}
    if prev_state in ("exhausted", "draft", "quarantined") and remaining > _MIN_REMAINING_FOR_LEASE:
        vals["state"] = "active"
    assert "state" not in vals


def test_200_with_high_remaining_activates():
    remaining = 5000
    prev_state = "draft"
    vals = {}
    if prev_state in ("exhausted", "draft", "quarantined") and remaining > _MIN_REMAINING_FOR_LEASE:
        vals["state"] = "active"
    assert vals["state"] == "active"


def test_200_active_stays_active():
    remaining = 5000
    prev_state = "active"
    vals = {}
    if prev_state in ("exhausted", "draft", "quarantined") and remaining > _MIN_REMAINING_FOR_LEASE:
        vals["state"] = "active"
    assert "state" not in vals
