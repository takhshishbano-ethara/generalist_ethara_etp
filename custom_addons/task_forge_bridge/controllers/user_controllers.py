from odoo import http
from odoo.http import request
from odoo.addons.api_auth_gateway.controllers.utility import (
    return_Response, validate_token, validate_request
)
from datetime import datetime, date
import json
import base64
import io


class TaskForgeUserController(http.Controller):

    @http.route('/api/v2/taskforge/users', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def list_users(self, **kwargs):
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)

            team_ids = employee._get_team_employee_ids()
            Employee = request.env['hr.employee'].sudo()
            employees = Employee.browse(team_ids)

            data = [self._format_employee(emp) for emp in employees]
            return return_Response(message="Users list", status=200, data={'data': data})
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/taskforge/users/team', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def team_users(self, **kwargs):
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)

            team_ids = employee._get_team_employee_ids()
            Employee = request.env['hr.employee'].sudo()
            employees = Employee.search([
                ('id', 'in', team_ids),
                ('task_forge_active', '=', True),
            ])

            today = date.today()
            Attendance = request.env['hr.attendance'].sudo()
            data = []
            for emp in employees:
                emp_data = self._format_employee(emp)
                att = Attendance.search([
                    ('employee_id', '=', emp.id),
                    ('check_in', '>=', datetime.combine(today, datetime.min.time())),
                    ('check_in', '<', datetime.combine(today, datetime.max.time())),
                ], limit=1)
                emp_data['tasking_status'] = 'Tasking' if att and not att.check_out else 'Idle'
                data.append(emp_data)

            return return_Response(message="Team list", status=200, data={'data': data})
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/taskforge/users/offboarded', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def offboarded_users(self, **kwargs):
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)

            role = employee._get_task_forge_role()
            Employee = request.env['hr.employee'].sudo()
            domain = [('task_forge_active', '=', False)]

            if role == 'pl':
                domain.append(('task_forge_pl_id', '=', employee.id))
            elif role in ('qr', 'ql'):
                domain.append(('task_forge_qr_id', '=', employee.id))
            elif role == 'tasker':
                domain.append(('id', '=', employee.id))

            employees = Employee.search(domain)
            data = [self._format_employee(emp) for emp in employees]
            return return_Response(message="Offboarded users", status=200, data={'data': data})
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/taskforge/users/pls', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def list_pls(self, **kwargs):
        try:
            Employee = request.env['hr.employee'].sudo()
            pl_group = request.env.ref('etp_user_roles.group_project_lead')
            pls = Employee.search([
                ('user_id.groups_id', 'in', [pl_group.id]),
                ('task_forge_active', '=', True),
            ])
            data = [{'id': p.id, 'name': p.name} for p in pls]
            return return_Response(message="PLs list", status=200, data={'data': data})
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/taskforge/users/qrs', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def list_qrs(self, **kwargs):
        try:
            Employee = request.env['hr.employee'].sudo()
            qr_group = request.env.ref('etp_user_roles.group_quality_reviewer')
            domain = [
                ('user_id.groups_id', 'in', [qr_group.id]),
                ('task_forge_active', '=', True),
            ]
            pl_id = kwargs.get('pl_id')
            if pl_id:
                domain.append(('task_forge_pl_id', '=', int(pl_id)))

            qrs = Employee.search(domain)
            data = [{'id': q.id, 'name': q.name, 'pl_id': q.task_forge_pl_id.id if q.task_forge_pl_id else None} for q in qrs]
            return return_Response(message="QRs list", status=200, data={'data': data})
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/taskforge/users/bulk_onboard', methods=['POST'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def bulk_onboard(self, **kwargs):
        try:
            user = request.env.user
            if not user.has_group('etp_user_roles.group_project_lead'):
                return return_Response(message="Insufficient permissions", status=403)

            file = request.httprequest.files.get('file')
            if not file:
                return return_Response(message="No file uploaded", status=400)

            filename = file.filename.lower()
            content = file.read()
            rows = []

            if filename.endswith('.csv'):
                import csv
                reader = csv.DictReader(io.StringIO(content.decode('utf-8')))
                rows = list(reader)
            elif filename.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content))
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
            else:
                return return_Response(message="Unsupported file format. Use CSV or XLSX.", status=400)

            Employee = request.env['hr.employee'].sudo()
            ResUsers = request.env['res.users'].sudo()
            created = []
            errors = []

            for idx, row in enumerate(rows, start=2):
                try:
                    name = row.get('name', '').strip()
                    email = row.get('email', '').strip()
                    if not name or not email:
                        errors.append(f"Row {idx}: name and email required")
                        continue
                    if not email.endswith('@ethara.ai'):
                        errors.append(f"Row {idx}: email must be @ethara.ai")
                        continue

                    existing = ResUsers.search([('login', '=', email)], limit=1)
                    if existing:
                        errors.append(f"Row {idx}: {email} already exists")
                        continue

                    new_user = ResUsers.create({
                        'name': name,
                        'login': email,
                        'email': email,
                        'password': row.get('password', 'Ethara@123'),
                    })
                    emp = new_user.employee_id
                    if not emp:
                        emp = Employee.create({
                            'name': name,
                            'work_email': email,
                            'user_id': new_user.id,
                        })

                    if row.get('pl_id'):
                        emp.write({'task_forge_pl_id': int(row['pl_id'])})
                    if row.get('qr_id'):
                        emp.write({'task_forge_qr_id': int(row['qr_id'])})

                    created.append({'id': emp.id, 'name': name, 'email': email})
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            return return_Response(
                message=f"Bulk onboard complete: {len(created)} created, {len(errors)} errors",
                status=200,
                data={'data': {'created': created, 'errors': errors}}
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @validate_token
    @http.route('/api/v2/get_tasker_list', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_request({})
    def get_tasker_list(self, **kwargs):
        temp = []
        try:
            user_id = request.env['res.users'].sudo().browse(request.env.uid)
            if not user_id.employee_id:
                return return_Response(message="Employee not found", status=404)
            domain = [('task_forge_qr_id', '=', user_id.employee_id.id)]
            if kwargs.get('qr_id'):
                domain = [('task_forge_qr_id', '=', int(kwargs.get('qr_id')))]
            total_tasker = request.env['hr.employee'].sudo().search(domain)
            temp = [self._format_employee(emp) for emp in total_tasker]
            return return_Response(message="Success", status=200, data={"records": temp, "count": len(temp)})
        except Exception as e:
            return return_Response(message="Something Went Wrong.", status=400, errors=[str(e)])


    def _format_employee(self, emp):
        Allocation = request.env['task.forge.allocation'].sudo()
        alloc = Allocation.search([('employee_id', '=', emp.id)], limit=1)
        return {
            'id': emp.id,
            'name': emp.name,
            'email': emp.work_email or '',
            'role': emp._get_task_forge_role(),
            'job_title': emp.job_title or '',
            'location': emp.work_location_name or '',
            'allocation_status': emp.tf_allocation_status or 'unallocated',
            'is_active': emp.task_forge_active,
            'pl_id': emp.task_forge_pl_id.id if emp.task_forge_pl_id else None,
            'pl_name': emp.task_forge_pl_id.name if emp.task_forge_pl_id else None,
            'qr_id': emp.task_forge_qr_id.id if emp.task_forge_qr_id else None,
            'qr_name': emp.task_forge_qr_id.name if emp.task_forge_qr_id else None,
            'current_project': alloc.project_id.name if alloc else None,
            'current_project_id': alloc.project_id.id if alloc else None,
        }
