from __future__ import annotations

import base64
import csv
import io
import logging

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)


HEADER_MAP = {
    "project type": "project_type",
    "instruction": "instruction",
    "original image url": "original_image_url",
    "edited image url": "edited_image_url",
    "does the edit make only the instructed change?": "edit_only_instructed",
    "are the two images aligned?": "images_aligned",
    "are both images free of ai slop?": "free_of_ai_slop",
    "tasker remarks": "tasker_remarks",
    "email address": "_email",
    "timestamp": "_skip",
    "qc status": "_skip",
    "ql remarks": "_skip",
    "final decision": "_skip",
    "qr remark": "_skip",
}

VALUE_MAP = {
    "edit_only_instructed": {
        "instruction aligned": "instruction_aligned",
        "no": "no",
    },
    "images_aligned": {
        "images aligned": "images_aligned",
        "no": "no",
    },
    "free_of_ai_slop": {
        "slop free": "slop_free",
        "no": "no",
    },
}

REQUIRED_FIELDS = ("project_type", "instruction", "original_image_url", "edited_image_url")


class I2IImportWizard(models.TransientModel):
    _name = "i2i.import.wizard"
    _description = "I2I Import Wizard (CSV / XLSX)"

    file = fields.Binary(string="Upload File (CSV or XLSX)", required=True)
    filename = fields.Char(string="Filename")
    skip_header = fields.Boolean(string="File has header row", default=True)
    created_count = fields.Integer(readonly=True)
    error_log = fields.Text(readonly=True)

    def action_import(self):
        self.ensure_one()
        if not self.file:
            raise UserError(_("Please upload a CSV or XLSX file."))
        if not self.skip_header:
            raise UserError(_("File must have a header row for column mapping."))

        filename = (self.filename or "").lower()
        raw = base64.b64decode(self.file)
        rows = self._read_xlsx(raw) if filename.endswith(".xlsx") else self._read_csv(raw)
        if not rows:
            raise UserError(_("File contains no data rows."))

        header_raw = [str(c or "").strip().lower() for c in rows[0]]
        header_map = [HEADER_MAP.get(h, None) for h in header_raw]

        mapped_fields = {f for f in header_map if f and not f.startswith("_")}
        missing = [f for f in REQUIRED_FIELDS if f not in mapped_fields]
        if missing:
            raise UserError(_(
                "Missing required columns: %s.\n"
                "Headers found: %s"
            ) % (", ".join(missing), ", ".join(header_raw)))

        Item = self.env["i2i.item"]
        Users = self.env["res.users"]
        created = 0
        errors = []
        created_items = self.env["i2i.item"]

        for idx, row in enumerate(rows[1:], start=2):
            try:
                vals = self._row_to_vals(header_map, row, Users, idx)
                item = Item.create(vals)
                created_items |= item
                created += 1
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Row {idx}: {exc}")
                _logger.warning("I2I import row %d failed: %s", idx, exc)

        for item in created_items:
            try:
                item.state = "human_qc"
                if item.llm_status in ("none", "failed"):
                    item._schedule_llm_qc()
            except Exception as exc:  # noqa: BLE001
                errors.append(f"Item {item.name}: post-import advance failed: {exc}")
                _logger.warning("I2I import: post-create advance for %s failed: %s", item.name, exc)

        self.write({
            "created_count": created,
            "error_log": "\n".join(errors[:50]) or False,
        })
        if errors:
            return {
                "type": "ir.actions.act_window",
                "name": _("Import I2I Items - Results"),
                "res_model": self._name,
                "res_id": self.id,
                "view_mode": "form",
                "target": "new",
            }
        message = _("%d items imported, advanced to Pending, LLM QC queued.") % created
        return {
            "type": "ir.actions.client",
            "tag": "display_notification",
            "params": {
                "title": _("Import Successful"),
                "message": message,
                "type": "success",
                "sticky": False,
                "next": {"type": "ir.actions.act_window_close"},
            },
        }

    def _row_to_vals(self, header_map, row, Users, row_idx):
        vals = {}
        email = None
        for col_idx, field_name in enumerate(header_map):
            if not field_name or field_name == "_skip":
                continue
            if col_idx >= len(row):
                continue
            raw_val = str(row[col_idx] or "").strip()
            if not raw_val:
                continue
            if field_name == "_email":
                email = raw_val
                continue
            if field_name in VALUE_MAP:
                key = VALUE_MAP[field_name].get(raw_val.lower())
                if not key:
                    raise UserError(_(
                        "Invalid value '%s' for column '%s'. Allowed: %s"
                    ) % (raw_val, field_name, ", ".join(VALUE_MAP[field_name].keys())))
                vals[field_name] = key
            else:
                vals[field_name] = raw_val

        for required in REQUIRED_FIELDS:
            if not vals.get(required):
                raise UserError(_("Missing required field: %s") % required)

        if email:
            user = Users.search([("login", "=ilike", email)], limit=1)
            if not user:
                raise UserError(_("User with email '%s' not found.") % email)
            vals["user_id"] = user.id

        return vals

    def _read_csv(self, raw):
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        return [list(r) for r in reader]

    def _read_xlsx(self, raw):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError(_(
                "openpyxl is required for XLSX import. Install: pip install openpyxl"
            )) from exc
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        return rows
