import base64
import binascii
import json
import logging
import re
from datetime import datetime

from odoo import http, SUPERUSER_ID
from odoo.http import request

from odoo.addons.api_auth_gateway.controllers.utility import (
    return_Response,
    is_valid_aadhaar,
    is_valid_email,
    is_valid_mobile,
    generate_s3_link,
)
from odoo.addons.employee_extension.services.aadhaar_extractor import (
    extract_aadhaar,
    mask_aadhaar,
    verify_against_input,
)

_logger = logging.getLogger(__name__)

EMP_CODE_PATTERN = re.compile(r'^GRP\d{3,}$')
PASSWORD_PATTERN = re.compile(r'^(?=.*[A-Z])(?=.*\d).{8,}$')
AADHAAR_ALLOWED_EXT = {'pdf', 'jpg', 'jpeg', 'png', 'webp'}
RESUME_ALLOWED_EXT = {'pdf', 'doc', 'docx'}
MAX_FILE_SIZE = 10 * 1024 * 1024
GENDER_VALUES = {'male', 'female', 'other'}
EXPERIENCE_VALUES = {'fresher', 'experienced'}
REGISTRATION_TYPES = {'employee', 'candidate'}
ETHARA_DOMAIN = '@ethara.ai'


def _ext_of(filename):
    if not filename or '.' not in filename:
        return ''
    return filename.rsplit('.', 1)[-1].lower()


class EmployeeSelfRegistrationController(http.Controller):

    def _safe_response(self, message, status=400):
        try:
            return return_Response(message=message, status=status)
        except Exception:
            _logger.exception("return_Response itself failed")
            return http.Response(
                json.dumps({'message': message, 'status_code': status, 'errors': []}),
                status=status,
                mimetype='application/json',
            )

    def _read_json_body(self, kwargs):
        data = {}
        try:
            raw = request.httprequest.get_data(cache=False, as_text=False) or b''
        except Exception as e:
            _logger.exception("Reading request body failed")
            return None, self._safe_response(f"Could not read request body: {e}", status=400)
        if raw:
            try:
                parsed = json.loads(raw)
            except (ValueError, TypeError):
                return None, self._safe_response("Request body must be valid JSON", status=400)
            if not isinstance(parsed, dict):
                return None, self._safe_response("Request body must be a JSON object", status=400)
            data.update(parsed)
        if kwargs:
            data.update({k: v for k, v in kwargs.items() if k not in data})
        return data, None

    def _decode_b64(self, b64_value, field_label):
        if not b64_value:
            return None, None
        try:
            return base64.b64decode(b64_value, validate=True), None
        except (binascii.Error, ValueError):
            return None, self._safe_response(f"{field_label} is not valid base64", status=400)

    def _process_employee(
        self, data,
        aadhaar_bytes, aadhaar_filename,
        resume_bytes, resume_filename,
    ):
        name = (data.get('name') or '').strip()
        gender = (data.get('gender') or '').strip().lower()
        phone = (data.get('phone') or '').strip()
        personal_email = (data.get('personal_email') or '').strip().lower()
        work_email = (data.get('work_email') or '').strip().lower()
        employee_code = (data.get('employee_code') or '').strip().upper()
        aadhaar_number = (data.get('aadhaar_number') or '').strip()
        birthday_raw = (data.get('birthday') or '').strip()
        password = data.get('password') or ''
        confirm_password = data.get('confirm_password') or ''

        try:
            department_id = int(data.get('department_id') or 0)
        except (TypeError, ValueError):
            department_id = 0
        try:
            designation_id = int(data.get('designation_id') or 0)
        except (TypeError, ValueError):
            designation_id = 0

        errors = []
        if not name:
            errors.append("Full name is required")
        if gender not in GENDER_VALUES:
            errors.append("Gender must be one of: male, female, other")
        if not is_valid_mobile(phone):
            errors.append("Phone number must be a valid 10-digit Indian mobile")
        if not is_valid_email(personal_email):
            errors.append("Personal email is invalid")
        if not is_valid_email(work_email) or not work_email.endswith(ETHARA_DOMAIN):
            errors.append("Ethara email must be a valid @ethara.ai address")
        # if not EMP_CODE_PATTERN.match(employee_code):
        #     errors.append("Employee code must match the format GRPXXXX (e.g. GRP1234)")
        if not department_id:
            errors.append("Department is required")
        if not designation_id:
            errors.append("Designation is required")
        if not is_valid_aadhaar(aadhaar_number):
            errors.append("Aadhaar number must be 12 digits and start with 2-9")
        if not PASSWORD_PATTERN.match(password):
            errors.append("Password must be at least 8 characters with 1 uppercase and 1 number")
        if password != confirm_password:
            errors.append("Passwords do not match")

        birthday_value = False
        if not birthday_raw:
            errors.append("Date of birth is required")
        else:
            try:
                birthday_value = datetime.strptime(birthday_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append("Date of birth must be in YYYY-MM-DD format")

        if not aadhaar_bytes:
            errors.append("Aadhaar card upload is required")
        else:
            ext = _ext_of(aadhaar_filename)
            if ext not in AADHAAR_ALLOWED_EXT:
                errors.append("Aadhaar card must be PDF, JPG, PNG or WEBP")
            elif len(aadhaar_bytes) > MAX_FILE_SIZE:
                errors.append("Aadhaar file exceeds 10 MB limit")

        if resume_bytes:
            resume_ext = _ext_of(resume_filename)
            if resume_ext not in RESUME_ALLOWED_EXT:
                errors.append("Resume must be PDF, DOC or DOCX")
            elif len(resume_bytes) > MAX_FILE_SIZE:
                errors.append("Resume exceeds 10 MB limit")

        if errors:
            return return_Response(message=errors[0], status=400, errors=errors)

        extracted = extract_aadhaar(aadhaar_bytes, aadhaar_filename)
        if extracted.get('error'):
            return return_Response(message=extracted['error'], status=400)
        ok, verify_msg, _ = verify_against_input(
            extracted,
            aadhaar_number,
            birthday_value.isoformat() if birthday_value else None,
            name or None,
        )
        if not ok:
            _logger.info(
                "Employee registration blocked: extracted=%s entered=%s reason=%s",
                mask_aadhaar(extracted.get('aadhaar_number')),
                mask_aadhaar(aadhaar_number),
                verify_msg,
            )
            return return_Response(message=verify_msg, status=400)

        admin_env = request.env(user=SUPERUSER_ID)
        ResUsers = admin_env['res.users']
        Employee = admin_env['hr.employee']
        Department = admin_env['hr.department']

        if department_id and not Department.browse(department_id).exists():
            return return_Response(message=f"Department id {department_id} does not exist", status=400)

        designation_model_name = None
        if 'designation_id' in Employee._fields:
            designation_model_name = Employee._fields['designation_id'].comodel_name
        if designation_id and designation_model_name:
            if not admin_env[designation_model_name].browse(designation_id).exists():
                return return_Response(
                    message=f"Designation id {designation_id} does not exist in {designation_model_name}",
                    status=400,
                )

        if ResUsers.search_count([('login', '=', work_email)]):
            return return_Response(message=f"An account already exists for {work_email}", status=400)
        if 'employee_code' in Employee._fields and Employee.search_count(
            [('employee_code', '=ilike', employee_code)]
        ):
            return return_Response(message=f"Employee code {employee_code} already exists", status=400)
        if Employee.search_count([('identification_id', '=', aadhaar_number)]):
            return return_Response(message="Aadhaar number already registered", status=400)
        if Employee.search_count([('work_email', '=ilike', work_email)]):
            return return_Response(message=f"Employee with email {work_email} already exists", status=400)

        company = (
            admin_env.ref('base.main_company', raise_if_not_found=False)
            or admin_env['res.company'].search([], limit=1, order='id asc')
        )
        if not company:
            return return_Response(message="No company configured", status=500)

        new_user = ResUsers.with_company(company).with_context(
            no_reset_password=True,
            mail_create_nosubscribe=True,
            mail_create_nolog=True,
            tracking_disable=True,
        ).create({
            'name': name,
            'login': work_email,
            'email': work_email,
            'password': password,
            'company_id': company.id,
            'company_ids': [(6, 0, [company.id])],
        })

        employee_vals = {
            'name': name,
            'user_id': new_user.id,
            'company_id': company.id,
        }
        # Write to the canonical fields the employee-onboarding detail form reads
        # back (employee_onboarding/controllers/onboarding_controller.py serializer)
        # so Aadhaar / phone / gender captured here are pre-filled on first load.
        # `sex` is the hr.version gender field; `private_phone` backs the
        # onboarding "contact number"; `aadhaar_number` is the dedicated field
        # alongside the standard `identification_id`.
        for fname, fval in (
            ('gender', gender),
            ('sex', gender),
            ('work_phone', phone),
            ('private_phone', phone),
            ('work_email', work_email),
            ('identification_id', aadhaar_number),
            ('aadhaar_number', aadhaar_number),
            ('private_email', personal_email),
            ('employee_code', employee_code),
        ):
            if fval and fname in Employee._fields:
                employee_vals[fname] = fval
        if department_id and 'department_id' in Employee._fields:
            employee_vals['department_id'] = department_id
        if designation_id and 'designation_id' in Employee._fields:
            employee_vals['designation_id'] = designation_id
        if birthday_value and 'birthday' in Employee._fields:
            employee_vals['birthday'] = birthday_value

        employee = new_user.employee_id
        if employee:
            employee.write(employee_vals)
        else:
            employee = Employee.create(employee_vals)

        aadhaar_b64 = base64.b64encode(aadhaar_bytes).decode('utf-8')
        aadhaar_url = generate_s3_link(
            aadhaar_b64,
            prefix='employee/aadhaar',
            uid=employee.id,
            filename=aadhaar_filename or 'aadhaar',
        )
        if not aadhaar_url:
            return return_Response(message="Failed to upload Aadhaar card to storage", status=500)

        write_vals = {}
        if 'aadhaar_card_url' in employee._fields:
            write_vals['aadhaar_card_url'] = aadhaar_url

        resume_url = ''
        if resume_bytes:
            resume_b64 = base64.b64encode(resume_bytes).decode('utf-8')
            resume_url = generate_s3_link(
                resume_b64,
                prefix='employee/resume',
                uid=employee.id,
                filename=resume_filename or 'resume',
            )
            if not resume_url:
                return return_Response(message="Failed to upload resume to storage", status=500)
            if 'resume_url' in employee._fields:
                write_vals['resume_url'] = resume_url

        if write_vals:
            employee.write(write_vals)

        return return_Response(
            message="Registration successful. You can log in with your Ethara email and password.",
            status=200,
            data={'data': {
                'type': 'employee',
                'employee_id': employee.id,
                'user_id': new_user.id,
                'work_email': work_email,
                'employee_code': employee_code,
                'aadhaar_card_url': aadhaar_url,
                'resume_url': resume_url,
            }},
        )

    def _process_candidate(
        self, data,
        aadhaar_bytes, aadhaar_filename,
        resume_bytes, resume_filename,
    ):
        name = (data.get('name') or '').strip()
        gender = (data.get('gender') or '').strip().lower()
        phone = (data.get('phone') or '').strip()
        personal_email = (data.get('personal_email') or '').strip().lower()
        aadhaar_number = (data.get('aadhaar_number') or '').strip()
        birthday_raw = (data.get('birthday') or '').strip()
        password = data.get('password') or ''
        confirm_password = data.get('confirm_password') or ''
        experience = (data.get('experience') or 'fresher').strip().lower()

        try:
            experience_years = float(data.get('experience_years') or 0)
        except (TypeError, ValueError):
            experience_years = -1.0
        try:
            college_id = int(data.get('college_id') or 0)
        except (TypeError, ValueError):
            college_id = 0
        college_name = (data.get('college_name') or '').strip()

        errors = []
        if not name:
            errors.append("Full name is required")
        if gender and gender not in GENDER_VALUES:
            errors.append("Gender must be one of: male, female, other")
        if phone and not is_valid_mobile(phone):
            errors.append("Phone number must be a valid 10-digit Indian mobile")
        if not is_valid_email(personal_email):
            errors.append("Personal email is invalid")
        if not is_valid_aadhaar(aadhaar_number):
            errors.append("Aadhaar number must be 12 digits and start with 2-9")
        if not PASSWORD_PATTERN.match(password):
            errors.append("Password must be at least 8 characters with 1 uppercase and 1 number")
        if password != confirm_password:
            errors.append("Passwords do not match")
        if experience not in EXPERIENCE_VALUES:
            errors.append("Experience must be 'fresher' or 'experienced'")
        if experience == 'experienced' and experience_years <= 0:
            errors.append("Experience years is required for experienced candidates")
        if experience_years < 0:
            errors.append("Experience years must be a non-negative number")

        birthday_value = False
        if birthday_raw:
            try:
                birthday_value = datetime.strptime(birthday_raw, '%Y-%m-%d').date()
            except ValueError:
                errors.append("Date of birth must be in YYYY-MM-DD format")

        if not aadhaar_bytes:
            errors.append("Aadhaar card upload is required")
        else:
            ext = _ext_of(aadhaar_filename)
            if ext not in AADHAAR_ALLOWED_EXT:
                errors.append("Aadhaar card must be PDF, JPG, PNG or WEBP")
            elif len(aadhaar_bytes) > MAX_FILE_SIZE:
                errors.append("Aadhaar file exceeds 10 MB limit")

        if not resume_bytes:
            errors.append("Resume upload is required for candidates")
        else:
            resume_ext = _ext_of(resume_filename)
            if resume_ext not in RESUME_ALLOWED_EXT:
                errors.append("Resume must be PDF, DOC or DOCX")
            elif len(resume_bytes) > MAX_FILE_SIZE:
                errors.append("Resume exceeds 10 MB limit")

        if errors:
            return return_Response(message=errors[0], status=400, errors=errors)

        extracted = extract_aadhaar(aadhaar_bytes, aadhaar_filename)
        if extracted.get('error'):
            return return_Response(message=extracted['error'], status=400)
        ok, verify_msg, _ = verify_against_input(
            extracted,
            aadhaar_number,
            birthday_value.isoformat() if birthday_value else None,
            name or None,
        )
        if not ok:
            _logger.info(
                "Candidate registration blocked: extracted=%s entered=%s reason=%s",
                mask_aadhaar(extracted.get('aadhaar_number')),
                mask_aadhaar(aadhaar_number),
                verify_msg,
            )
            return return_Response(message=verify_msg, status=400)

        admin_env = request.env(user=SUPERUSER_ID)
        ResUsers = admin_env['res.users']
        Applicant = admin_env['hr.applicant']
        College = admin_env['employee.college']

        if not college_id and college_name:
            existing_college = College.with_context(active_test=False).search(
                [('name', '=ilike', college_name)], limit=1,
            )
            if existing_college:
                college_id = existing_college.id
            else:
                college_id = College.create({
                    'name': college_name,
                    'active': True,
                }).id

        if college_id and not College.browse(college_id).exists():
            return return_Response(message=f"College id {college_id} does not exist", status=400)

        if ResUsers.search_count([('login', '=', personal_email)]):
            return return_Response(message=f"An account already exists for {personal_email}", status=400)
        if Applicant.search_count([('email_from', '=ilike', personal_email)]):
            return return_Response(message=f"A candidate profile already exists for {personal_email}", status=400)
        if aadhaar_number and Applicant.search_count([('aadhaar_number', '=', aadhaar_number)]):
            return return_Response(message="Aadhaar number already registered as a candidate", status=400)

        company = (
            admin_env.ref('base.main_company', raise_if_not_found=False)
            or admin_env['res.company'].search([], limit=1, order='id asc')
        )
        if not company:
            return return_Response(message="No company configured", status=500)

        portal_group = admin_env.ref('base.group_portal', raise_if_not_found=False)
        user_vals = {
            'name': name,
            'login': personal_email,
            'email': personal_email,
            'password': password,
            'company_id': company.id,
            'company_ids': [(6, 0, [company.id])],
        }
        if portal_group:
            user_vals['group_ids'] = [(6, 0, [portal_group.id])]
        new_user = ResUsers.with_company(company).with_context(
            no_reset_password=True,
            mail_create_nosubscribe=True,
            mail_create_nolog=True,
            tracking_disable=True,
        ).create(user_vals)

        applicant_vals = {
            'partner_name': name,
            'partner_id': new_user.partner_id.id,
            'email_from': personal_email,
            'partner_phone': phone or False,
            'company_id': company.id,
            'gender': gender or False,
            'birthday': birthday_value or False,
            'aadhaar_number': aadhaar_number,
            'experience': experience,
            'experience_years': experience_years if experience == 'experienced' else 0.0,
            'college_id': college_id or False,
            'candidate_user_id': new_user.id,
        }
        applicant = Applicant.with_company(company).with_context(
            mail_create_nosubscribe=True,
            mail_create_nolog=True,
            tracking_disable=True,
        ).create(applicant_vals)

        aadhaar_b64 = base64.b64encode(aadhaar_bytes).decode('utf-8')
        aadhaar_url = generate_s3_link(
            aadhaar_b64,
            prefix='candidate/aadhaar',
            uid=applicant.id,
            filename=aadhaar_filename or 'aadhaar',
        )
        if not aadhaar_url:
            return return_Response(message="Failed to upload Aadhaar card to storage", status=500)

        resume_b64 = base64.b64encode(resume_bytes).decode('utf-8')
        resume_url = generate_s3_link(
            resume_b64,
            prefix='candidate/resume',
            uid=applicant.id,
            filename=resume_filename or 'resume',
        )
        if not resume_url:
            return return_Response(message="Failed to upload resume to storage", status=500)

        applicant.write({
            'aadhaar_card_url': aadhaar_url,
            'resume_url': resume_url,
        })

        return return_Response(
            message="Candidate profile created. You can log in with your personal email and password.",
            status=200,
            data={'data': {
                'type': 'candidate',
                'applicant_id': applicant.id,
                'user_id': new_user.id,
                'personal_email': personal_email,
                'experience': experience,
                'experience_years': applicant.experience_years,
                'aadhaar_card_url': aadhaar_url,
                'resume_url': resume_url,
            }},
        )

    @http.route(
        '/api/v2/employees/registration_options',
        methods=['GET'], type='http', auth='public', csrf=False, cors='*',
    )
    def registration_options(self, **kwargs):
        try:
            admin_env = request.env(user=SUPERUSER_ID)

            departments = admin_env['hr.department'].search_read(
                domain=[('active', '=', True)],
                fields=['id', 'name'],
                order='name asc',
            )

            designations = []
            if 'hr.employee.designation' in admin_env:
                designations = admin_env['hr.employee.designation'].search_read(
                    domain=[('active', '=', True)],
                    fields=['id', 'name', 'code'],
                    order='name asc',
                )

            return return_Response(
                message="ok",
                status=200,
                data={'data': {
                    'departments': departments,
                    'designations': designations,
                }},
            )
        except Exception as e:
            _logger.exception("registration_options failed")
            return self._safe_response(str(e) or "Internal error", status=500)

    @http.route(
        '/api/v2/employees/colleges',
        methods=['GET'], type='http', auth='public', csrf=False, cors='*',
    )
    def colleges(self, **kwargs):
        try:
            admin_env = request.env(user=SUPERUSER_ID)
            if 'employee.college' not in admin_env:
                return self._safe_response("College model not available", status=500)

            search_term = (kwargs.get('q') or '').strip()
            try:
                limit = int(kwargs.get('limit') or 0) or None
            except (TypeError, ValueError):
                limit = None

            domain = [('active', '=', True)]
            if search_term:
                domain.append(('name', 'ilike', search_term))

            colleges = admin_env['employee.college'].search_read(
                domain=domain,
                fields=['id', 'name', 'code', 'city'],
                order='name asc',
                limit=limit,
            )

            return return_Response(
                message="ok",
                status=200,
                data={'data': {
                    'colleges': colleges,
                    'count': len(colleges),
                }},
            )
        except Exception as e:
            _logger.exception("colleges endpoint failed")
            return self._safe_response(str(e) or "Internal error", status=500)

    @http.route(
        '/api/v2/employees/colleges/import',
        methods=['POST'], type='http', auth='public', csrf=False, cors='*',
    )
    def import_colleges(self, **kwargs):
        try:
            data, err_resp = self._read_json_body(kwargs)
            if err_resp is not None:
                return err_resp

            excel_url = (data.get('excel_url') or '').strip()
            excel_b64 = data.get('excel_file_base64')

            if not excel_url and not excel_b64:
                return self._safe_response(
                    "Provide either 'excel_url' or 'excel_file_base64'", status=400,
                )

            try:
                from openpyxl import load_workbook
            except ImportError:
                return self._safe_response("Server is missing the 'openpyxl' library", status=500)

            import io

            if excel_b64:
                excel_bytes, err_resp = self._decode_b64(excel_b64, 'excel_file_base64')
                if err_resp is not None:
                    return err_resp
            else:
                if not (excel_url.startswith('http://') or excel_url.startswith('https://')):
                    return self._safe_response("excel_url must be a valid http(s) URL", status=400)
                import urllib.request
                try:
                    with urllib.request.urlopen(excel_url, timeout=60) as resp:
                        excel_bytes = resp.read()
                except Exception as e:
                    return self._safe_response(f"Failed to download Excel: {e}", status=400)

            if len(excel_bytes) > 50 * 1024 * 1024:
                return self._safe_response("Excel file exceeds 50 MB limit", status=413)

            try:
                wb = load_workbook(filename=io.BytesIO(excel_bytes), read_only=True, data_only=True)
            except Exception as e:
                return self._safe_response(f"Could not parse Excel: {e}", status=400)

            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                return self._safe_response("Excel is empty", status=400)
            header = [str(h or '').strip().lower() for h in header_row]
            if 'name' not in header:
                return self._safe_response("Excel must have a 'name' column", status=400)

            idx_name = header.index('name')
            idx_code = header.index('code') if 'code' in header else None
            idx_city = header.index('city') if 'city' in header else None
            idx_active = header.index('active') if 'active' in header else None

            admin_env = request.env(user=SUPERUSER_ID)
            if 'employee.college' not in admin_env:
                return self._safe_response("College model not available", status=500)

            College = admin_env['employee.college']
            existing = College.with_context(active_test=False).search_read(
                [], ['name', 'code']
            )
            existing_codes = {(c['code'] or '').strip().lower() for c in existing if c.get('code')}
            existing_names = {(c['name'] or '').strip().lower() for c in existing if c.get('name')}

            to_create = []
            skipped_duplicates = 0
            skipped_blank = 0

            for row in rows_iter:
                if row is None:
                    continue

                def cell(i):
                    if i is None or i >= len(row):
                        return ''
                    return str(row[i]).strip() if row[i] is not None else ''

                name = cell(idx_name)
                code = cell(idx_code)
                city = cell(idx_city)

                active = True
                if idx_active is not None and idx_active < len(row):
                    v = row[idx_active]
                    if isinstance(v, bool):
                        active = v
                    elif isinstance(v, (int, float)):
                        active = bool(v)
                    elif isinstance(v, str):
                        active = v.strip().lower() in ('1', 'true', 'yes', 'y')

                if not name:
                    skipped_blank += 1
                    continue

                code_key = code.lower()
                if code and code_key in existing_codes:
                    skipped_duplicates += 1
                    continue
                if not code and name.lower() in existing_names:
                    skipped_duplicates += 1
                    continue

                to_create.append({
                    'name': name,
                    'code': code or False,
                    'city': city or False,
                    'active': active,
                })
                if code:
                    existing_codes.add(code_key)
                existing_names.add(name.lower())

            created_count = 0
            if to_create:
                batch_size = 1000
                for i in range(0, len(to_create), batch_size):
                    College.create(to_create[i:i + batch_size])
                    created_count += len(to_create[i:i + batch_size])

            return return_Response(
                message=f"Imported {created_count} colleges. Skipped {skipped_duplicates} duplicates, {skipped_blank} blank rows.",
                status=200,
                data={'data': {
                    'created': created_count,
                    'skipped_duplicates': skipped_duplicates,
                    'skipped_blank': skipped_blank,
                }},
            )
        except Exception as e:
            _logger.exception("import_colleges failed")
            return self._safe_response(str(e) or "Internal error", status=500)

    @http.route(
        '/api/v2/employees/aadhaar/verify',
        methods=['POST'], type='http', auth='public', csrf=False, cors='*',
    )
    def verify_aadhaar(self, **kwargs):
        try:
            content_length = request.httprequest.content_length or 0
            if content_length > (MAX_FILE_SIZE + 64 * 1024):
                return self._safe_response(
                    "Aadhaar file must be <= 10 MB", status=413,
                )

            data, err_resp = self._read_json_body(kwargs)
            if err_resp is not None:
                return err_resp

            entered_aadhaar = (data.get('aadhaar_number') or '').strip()
            entered_dob = (data.get('birthday') or '').strip() or None
            entered_name = (data.get('name') or '').strip() or None
            aadhaar_b64 = data.get('aadhaar_file_base64')
            aadhaar_filename = (data.get('aadhaar_file_name') or 'aadhaar.pdf').strip()

            errors = []
            if not is_valid_aadhaar(entered_aadhaar):
                errors.append("Aadhaar number must be 12 digits and start with 2-9")
            if not aadhaar_b64:
                errors.append("Aadhaar card upload is required")
            if entered_dob:
                try:
                    datetime.strptime(entered_dob, '%Y-%m-%d')
                except ValueError:
                    errors.append("Date of birth must be in YYYY-MM-DD format")
            if errors:
                return return_Response(message=errors[0], status=400, errors=errors)

            aadhaar_bytes, err_resp = self._decode_b64(aadhaar_b64, 'aadhaar_file_base64')
            if err_resp is not None:
                return err_resp

            ext = _ext_of(aadhaar_filename)
            if ext not in AADHAAR_ALLOWED_EXT:
                return self._safe_response(
                    "Aadhaar card must be PDF, JPG, PNG or WEBP", status=400,
                )
            if len(aadhaar_bytes) > MAX_FILE_SIZE:
                return self._safe_response(
                    "Aadhaar file exceeds 10 MB limit", status=413,
                )

            extracted = extract_aadhaar(aadhaar_bytes, aadhaar_filename)
            if extracted.get('error'):
                return return_Response(
                    message=extracted['error'],
                    status=400,
                    data={'data': {
                        'verified': False,
                        'extracted': None,
                        'checks': None,
                        'source': None,
                    }},
                )

            verified, message, checks = verify_against_input(
                extracted, entered_aadhaar, entered_dob, entered_name,
            )

            full = extracted.get('aadhaar_number')
            last4 = extracted.get('aadhaar_last4')
            _logger.info(
                "Aadhaar verify: entered=%s extracted=%s verified=%s source=%s",
                mask_aadhaar(entered_aadhaar),
                mask_aadhaar(full) if full else (f"XXXX XXXX {last4}" if last4 else "?"),
                verified, extracted.get('source'),
            )

            return return_Response(
                message=message,
                status=200,
                data={'data': {
                    'verified': verified,
                    'extracted': {
                        'aadhaar_number': full,
                        'aadhaar_number_masked': mask_aadhaar(full) if full else (
                            f"XXXX XXXX {last4}" if last4 else None
                        ),
                        'aadhaar_last4': last4,
                        'dob': extracted.get('dob'),
                        'name': extracted.get('name'),
                        'gender': extracted.get('gender'),
                    },
                    'checks': checks,
                    'source': extracted.get('source'),
                }},
            )
        except Exception as e:
            _logger.exception("verify_aadhaar failed")
            return self._safe_response(str(e) or "Internal error", status=500)

    @http.route(
        '/api/v2/employees/self_register',
        methods=['POST'], type='http', auth='public', csrf=False, cors='*',
    )
    def self_register(self, **kwargs):
        try:
            content_length = request.httprequest.content_length or 0
            if content_length > (MAX_FILE_SIZE * 2 + 64 * 1024):
                return self._safe_response(
                    "Request body too large. Aadhaar/resume must be <= 10 MB each.",
                    status=413,
                )

            data, err_resp = self._read_json_body(kwargs)
            if err_resp is not None:
                return err_resp

            reg_type = (data.pop('type', '') or '').strip().lower()
            if reg_type not in REGISTRATION_TYPES:
                return self._safe_response(
                    "'type' is required and must be 'employee' or 'candidate'",
                    status=400,
                )

            aadhaar_b64 = data.pop('aadhaar_file_base64', None)
            aadhaar_filename = data.pop('aadhaar_file_name', None) or 'aadhaar.pdf'
            data.pop('aadhaar_file_mimetype', None)

            resume_b64 = data.pop('resume_file_base64', None)
            resume_filename = data.pop('resume_file_name', None) or 'resume.pdf'
            data.pop('resume_file_mimetype', None)

            aadhaar_bytes, err_resp = self._decode_b64(aadhaar_b64, 'aadhaar_file_base64')
            if err_resp is not None:
                return err_resp
            resume_bytes, err_resp = self._decode_b64(resume_b64, 'resume_file_base64')
            if err_resp is not None:
                return err_resp

            if reg_type == 'employee':
                return self._process_employee(
                    data, aadhaar_bytes, aadhaar_filename, resume_bytes, resume_filename,
                )
            return self._process_candidate(
                data, aadhaar_bytes, aadhaar_filename, resume_bytes, resume_filename,
            )
        except Exception as e:
            _logger.exception("Self-registration dispatch failed")
            return self._safe_response(str(e) or "Internal error", status=500)
