from odoo import http, fields
from odoo.http import request
from odoo.addons.api_auth_gateway.controllers.utility import (
    return_Response, validate_token, validate_request
)
import json
import pandas as pd
import io
from datetime import datetime, date, timedelta, time
import calendar
import logging

_logger = logging.getLogger(__name__)


class EmployeeController(http.Controller):

    def _send_onboarding_email(self, user, plain_password):
        """Send welcome email with login credentials to newly onboarded employee."""
        try:
            template = request.env.ref(
                'employee_extension.email_template_employee_onboarding',
                raise_if_not_found=False,
            )
            login_url = request.env['ir.config_parameter'].sudo().get_param(
                'employee_extension.onboarding_login_url',
                default='https://etp.stage.ethara.ai',
            )
            role_name = ''
            if user.user_role:
                role_name = user.user_role.name or ''

            ctx = {
                'plain_password': plain_password,
                'login_url': login_url,
                'role_name': role_name,
            }

            if template:
                template.sudo().with_context(**ctx).send_mail(user.id, force_send=True)
            else:
                mail_values = {
                    'subject': 'Welcome to Ethara - Your Login Credentials',
                    'email_from': request.env['ir.config_parameter'].sudo().get_param(
                        'mail.catchall.email', 'noreply@ethara.ai'),
                    'email_to': user.email,
                    'body_html': '''
                        <div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                            <h2 style="color: #007bff;">Welcome to Ethara!</h2>
                            <p>Hi %s,</p>
                            <p>Your account has been created. Here are your login credentials:</p>
                            <div style="background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 20px; margin: 20px 0;">
                                <p><strong>Login Email:</strong> %s</p>
                                <p><strong>Password:</strong> %s</p>
                                <p><strong>Role:</strong> %s</p>
                            </div>
                            <p style="color: #856404; background: #fff3cd; border: 1px solid #ffc107; padding: 12px; border-radius: 5px;">
                                <strong>Important:</strong> Please change your password after your first login.
                            </p>
                            <p style="text-align: center; margin: 25px 0;">
                                <a href="%s" style="background: #007bff; color: white; padding: 12px 30px;
                                   text-decoration: none; border-radius: 5px; font-size: 16px;">Login to Ethara ETP</a>
                            </p>
                            <hr style="border: none; border-top: 1px solid #eee;"/>
                            <p style="color: #999; font-size: 11px;">Ethara ETP Team</p>
                        </div>
                    ''' % (user.name, user.login, plain_password, role_name, login_url),
                }
                request.env['mail.mail'].sudo().create(mail_values).send()
        except Exception as e:
            _logger.error('Failed to send onboarding email to %s: %s', user.login, str(e))

    @http.route('/api/v2/employees', methods=['POST'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    @validate_request({
        'name': {'type': 'string', 'required': True},
        'email': {'type': 'email', 'required': True},
        'job_title': {'type': 'int', 'required': True},
        'role_id': {'type': 'int', 'required': True},
    })
    def create_employee(self, **kwargs):
        try:
            user = request.env.user
            if user.user_role.id not in [request.env.ref('api_auth_gateway.role_qc_technical').id, request.env.ref('api_auth_gateway.role_qc_stem').id, request.env.ref('api_auth_gateway.role_qc_non_stem').id, request.env.ref('api_auth_gateway.role_pl_technical').id, request.env.ref('api_auth_gateway.role_cto_technical').id, request.env.ref('api_auth_gateway.role_pl_stem').id, request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                return return_Response(message="Permission denied: You are not authorized to use this feature.", status=400)

            jdata = kwargs.get('jdata')
            Employee = request.env['hr.employee'].sudo()
            ResUsers = request.env['res.users'].sudo()

            email = jdata.get('email', '').strip().lower()
            if not email.endswith('@ethara.ai'):
                return return_Response(message="Email must be @ethara.ai domain", status=400)

            existing_user = ResUsers.search([('login', '=', email)], limit=1)
            if existing_user:
                return return_Response(message=f"User with email {email} already exists", status=400)

            user_vals = {
                'name': jdata['name'],
                'login': email,
                'email': email,
                'user_role': jdata.get('role_id'),
                'password': 'Ethara@123',
            }
            new_user = ResUsers.create(user_vals)
            employee_vals = {}
            if user.user_role.id in [request.env.ref('api_auth_gateway.role_pl_technical').id, request.env.ref('api_auth_gateway.role_pl_stem').id, request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                employee_vals['task_forge_pl_id'] = user.employee_id.id
            if user.user_role.id in [request.env.ref('api_auth_gateway.role_qc_technical').id, request.env.ref('api_auth_gateway.role_qc_stem').id, request.env.ref('api_auth_gateway.role_qc_non_stem').id]:
                employee_vals['task_forge_qr_id'] = user.employee_id.id
                employee_vals['task_forge_pl_id'] = user.employee_id.task_forge_pl_id.id if user.employee_id else False
            if jdata.get('work_location_name'):
                employee_vals['work_location_name'] = jdata.get('work_location_name')
            if jdata.get('job_title'):
                employee_vals['designation_id'] = jdata.get('job_title')
            if jdata.get('department_id'):
                employee_vals['department_id'] = jdata['department_id']
            if jdata.get('pl_id'):
                employee_vals['task_forge_pl_id'] = jdata['pl_id']
            if jdata.get('qr_id'):
                employee_vals['task_forge_qr_id'] = jdata['qr_id']
            if not new_user.employee_id:
                employee_vals['name'] = jdata['name']
                employee_vals['work_email'] = email
                employee_vals['user_id'] = new_user.id
                employee = Employee.create(employee_vals)
            else:
                employee = new_user.employee_id
                new_user.employee_id.sudo().write(employee_vals)
            if kwargs.get('project_id'):
                ProjectRequest = request.env['project.project'].sudo().browse(int(kwargs.get('project_id')))
                emp_list = []
                if ProjectRequest.exists():
                    if employee.user_id.user_role.id in [request.env.ref('api_auth_gateway.role_pl_technical').id,
                                                         request.env.ref('api_auth_gateway.role_pl_stem').id,
                                                         request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                        emp_list = ProjectRequest.project_lead.ids
                        emp_list.append(employee.id)
                        ProjectRequest.sudo().project_lead = [(6, 0, emp_list)]

                    elif employee.user_id.user_role.id in [request.env.ref('api_auth_gateway.role_qc_technical').id,
                                                           request.env.ref('api_auth_gateway.role_qc_stem').id,
                                                           request.env.ref('api_auth_gateway.role_qc_non_stem').id]:
                        emp_list = ProjectRequest.project_qc_reviewer.ids
                        emp_list.append(employee.id)
                        ProjectRequest.sudo().project_qc_reviewer = [(6, 0, emp_list)]

                    elif employee.user_id.user_role.id in [
                        request.env.ref('api_auth_gateway.role_tasker_technical').id,
                        request.env.ref('api_auth_gateway.role_tasker_stem').id,
                        request.env.ref('api_auth_gateway.role_tasker_non_stem').id]:
                        emp_list = ProjectRequest.project_tasker.ids
                        emp_list.append(employee.id)
                        ProjectRequest.sudo().project_tasker = [(6, 0, emp_list)]

            self._send_onboarding_email(new_user, 'Ethara@123')

            try:
                request.env['kubera.notification'].sudo().create({
                    'title': 'New Employee Created',
                    'message': f'Employee "{jdata["name"]}" has been created.',
                    'user_id': request.env.user.id,
                    'priority': '1',
                    'res_model': 'hr.employee',
                    'res_id': new_user.employee_id.id if new_user.employee_id else 0,
                })
            except Exception:
                pass

            return return_Response(message="Employee created successfully", status=200)
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/employees/check_bulk_create_file', methods=['POST'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def check_bulk_create_file(self, **kwargs):
        try:
            user = request.env.user
            if user.user_role.id not in [request.env.ref('api_auth_gateway.role_qc_technical').id,
                                         request.env.ref('api_auth_gateway.role_qc_stem').id,
                                         request.env.ref('api_auth_gateway.role_qc_non_stem').id,
                                         request.env.ref('api_auth_gateway.role_pl_technical').id,
                                         request.env.ref('api_auth_gateway.role_cto_technical').id,
                                         request.env.ref('api_auth_gateway.role_pl_stem').id,
                                         request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                return return_Response(message="Permission denied: You are not authorized to use this feature.", status=400)

            file_obj = request.httprequest.files.get('file')
            if not file_obj:
                return return_Response(message="No file uploaded. Use key 'file'", status=400)
            filename = file_obj.filename.lower()
            file_content = file_obj.read()
            try:
                if filename.endswith('.csv'):
                    df = pd.read_csv(io.BytesIO(file_content))
                elif filename.endswith(('.xlsx', '.xls')):
                    df = None
                    last_error = None
                    
                    # Try openpyxl first for .xlsx files
                    if filename.endswith('.xlsx'):
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
                            sheet = wb.active
                            data = []
                            for row in sheet.iter_rows(values_only=True):
                                data.append(row)
                            wb.close()
                            if data:
                                df = pd.DataFrame(data[1:], columns=data[0])
                        except Exception as oxl_error:
                            last_error = str(oxl_error)
                    
                    # Try xlrd for .xls files
                    if df is None and filename.endswith('.xls'):
                        try:
                            df = pd.read_excel(io.BytesIO(file_content), engine='xlrd')
                        except Exception as xlrd_error:
                            last_error = str(xlrd_error)
                    
                    # Try with different engines as fallback
                    if df is None:
                        for engine in ['openpyxl', 'xlrd', 'odf']:
                            try:
                                df = pd.read_excel(io.BytesIO(file_content), engine=engine)
                                break
                            except Exception as engine_error:
                                last_error = str(engine_error)
                                continue
                    
                    if df is None:
                        return return_Response(
                            message="Unable to parse the Excel file. Please save it as a newer .xlsx format or convert to CSV.",
                            status=400
                        )
                else:
                    return return_Response(message="Unsupported file format. Use .csv, .xlsx, or .xls", status=400)
            except Exception as e:
                error_msg = str(e)
                # Check for common corruption patterns
                if 'META-INF' in error_msg or 'manifest.xml' in error_msg:
                    return return_Response(
                        message="The file appears to be corrupted or in an older format. Please save it as a new .xlsx file or use CSV format.",
                        status=400
                    )
                return return_Response(message=f"Error parsing file: {error_msg}", status=400)
            df = df.fillna('').astype(str)
            ResUsers = request.env['res.users'].sudo()
            record_temp = []
            for index, row in df.iterrows():
                idx = index + 1
                try:
                    name = row.get('name', '').strip()
                    email = row.get('email', '').strip().lower()
                    if name and email:
                        error_msg = ""
                        if not name or name == '':
                            error_msg = f"Row {idx}: name is missing"

                        if not email or not email.endswith('@ethara.ai'):
                            error_msg = f"Row {idx}: invalid ethara.ai email"

                        if ResUsers.search_count([('login', '=', email)]):
                            error_msg = f"Row {idx}: {email} already exists"

                        record_temp.append({
                            'index': index,
                            'name': name,
                            'email': email,
                            'user_role': row.get('user_role'),
                            'job_title': row.get('job_title'),
                            'error_msg': error_msg,
                        })
                except Exception as e:
                    print(f"Row {idx}: {str(e)}")

            return return_Response(
                message=f"Success",
                status=200,
                data={'data': record_temp}
            )

        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/employees/bulk_create', methods=['POST'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def bulk_create_employees_file(self, **kwargs):
        try:
            user = request.env.user
            if user.user_role.id not in [request.env.ref('api_auth_gateway.role_qc_technical').id,
                                         request.env.ref('api_auth_gateway.role_qc_stem').id,
                                         request.env.ref('api_auth_gateway.role_qc_non_stem').id,
                                         request.env.ref('api_auth_gateway.role_pl_technical').id,
                                         request.env.ref('api_auth_gateway.role_cto_technical').id,
                                         request.env.ref('api_auth_gateway.role_pl_stem').id,
                                         request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                return return_Response(message="Permission denied: You are not authorized to use this feature.", status=400)

            file_obj = request.httprequest.files.get('file')
            if not file_obj:
                return return_Response(message="No file uploaded. Use key 'file'", status=400)
            filename = file_obj.filename.lower()
            file_content = file_obj.read()
            try:
                if filename.endswith('.csv'):
                    df = pd.read_csv(io.BytesIO(file_content))
                elif filename.endswith(('.xlsx', '.xls')):
                    df = None
                    last_error = None

                    # Try openpyxl first for .xlsx files
                    if filename.endswith('.xlsx'):
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
                            sheet = wb.active
                            data = []
                            for row in sheet.iter_rows(values_only=True):
                                data.append(row)
                            wb.close()
                            if data:
                                df = pd.DataFrame(data[1:], columns=data[0])
                        except Exception as oxl_error:
                            last_error = str(oxl_error)

                    # Try xlrd for .xls files
                    if df is None and filename.endswith('.xls'):
                        try:
                            df = pd.read_excel(io.BytesIO(file_content), engine='xlrd')
                        except Exception as xlrd_error:
                            last_error = str(xlrd_error)

                    # Try with different engines as fallback
                    if df is None:
                        for engine in ['openpyxl', 'xlrd', 'odf']:
                            try:
                                df = pd.read_excel(io.BytesIO(file_content), engine=engine)
                                break
                            except Exception as engine_error:
                                last_error = str(engine_error)
                                continue

                    if df is None:
                        return return_Response(
                            message="Unable to parse the Excel file. Please save it as a newer .xlsx format or convert to CSV.",
                            status=400
                        )
                else:
                    return return_Response(message="Unsupported file format. Use .csv, .xlsx, or .xls", status=400)
            except Exception as e:
                error_msg = str(e)
                # Check for common corruption patterns
                if 'META-INF' in error_msg or 'manifest.xml' in error_msg:
                    return return_Response(
                        message="The file appears to be corrupted or in an older format. Please save it as a new .xlsx file or use CSV format.",
                        status=400
                    )
                return return_Response(message=f"Error parsing file: {error_msg}", status=400)
            df = df.fillna('').astype(str)
            Employee = request.env['hr.employee'].sudo()
            ResUsers = request.env['res.users'].sudo()
            created = []
            errors = []
            for index, row in df.iterrows():
                idx = index + 1
                try:
                    name = row.get('name', '').strip()
                    email = row.get('email', '').strip().lower()
                    if not name or name == '':
                        errors.append(f"Row {idx}: name is missing")
                        continue
                    if not email or not email.endswith('@ethara.ai'):
                        errors.append(f"Row {idx}: invalid ethara.ai email")
                        continue

                    if ResUsers.search_count([('login', '=', email)]):
                        errors.append(f"Row {idx}: {email} already exists")
                        continue

                    user_vals = {
                        'name': name,
                        'login': email,
                        'email': email,
                        'password': row.get('password') if row.get('password') else 'Ethara@123',
                    }
                    if row.get('user_role'):
                        user_role_domain = [('project_type', '=', 'non-stem'), ('name', '=', row.get('user_role'))]
                        if row.get('user_role') in ['CTO']:
                            user_role_domain = [('name', '=', row.get('user_role'))]
                        user_role = request.env['api.role'].sudo().search(user_role_domain, limit=1)
                        if user_role:
                            user_vals['user_role'] = user_role.id

                    new_user = ResUsers.create(user_vals)
                    employee_vals = {}
                    if row.get('job_title'):
                        designation_id = request.env['hr.employee.designation'].sudo().search([('name', '=', row.get('job_title'))], limit=1)
                        if designation_id:
                            employee_vals['designation_id'] = designation_id.id

                    if user.user_role.id in [request.env.ref('api_auth_gateway.role_pl_technical').id,
                                             request.env.ref('api_auth_gateway.role_pl_stem').id,
                                             request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                        employee_vals['task_forge_pl_id'] = user.employee_id.id
                    if user.user_role.id in [request.env.ref('api_auth_gateway.role_qc_technical').id,
                                             request.env.ref('api_auth_gateway.role_qc_stem').id,
                                             request.env.ref('api_auth_gateway.role_qc_non_stem').id]:
                        employee_vals['task_forge_qr_id'] = user.employee_id.id
                        employee_vals['task_forge_pl_id'] = user.employee_id.task_forge_pl_id.id

                    if row.get('qr_email'):
                        qr_email = request.env['res.users'].sudo().search([('login', '=', row.get('qr_email'))], limit=1)
                        if qr_email:
                            if qr_email.employee_id:
                                employee_vals['task_forge_qr_id'] = qr_email.employee_id.id

                    if row.get('pl_email'):
                        pl_email = request.env['res.users'].sudo().search([('login', '=', row.get('pl_email'))], limit=1)
                        if pl_email:
                            if pl_email.employee_id:
                                employee_vals['task_forge_pl_id'] = pl_email.employee_id.id

                    if row.get('department_id'):
                        employee_vals['department_id'] = int(row.get('department_id'))

                    if row.get('pl_id'):
                        employee_vals['task_forge_pl_id'] = int(row.get('pl_id'))

                    if row.get('qr_id'):
                        employee_vals['task_forge_qr_id'] = int(row.get('qr_id'))

                    if not new_user.employee_id:
                        employee_vals['name'] = name
                        employee_vals['work_email'] = email
                        employee_vals['user_id'] = new_user.id
                        employee_vals['work_location_name'] = row.get('work_location_name', '')

                        employee = Employee.create(employee_vals)
                    else:
                        employee = new_user.employee_id
                        new_user.employee_id.sudo().write(employee_vals)
                    if kwargs.get('project_id'):
                        ProjectRequest = request.env['project.project'].sudo().browse(int(kwargs.get('project_id')))
                        emp_list = []
                        if ProjectRequest.exists():
                            if employee.user_id.user_role.id in [request.env.ref('api_auth_gateway.role_pl_technical').id,
                                                            request.env.ref('api_auth_gateway.role_pl_stem').id,
                                                            request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                                emp_list = ProjectRequest.project_lead.ids
                                emp_list.append(employee.id)
                                ProjectRequest.sudo().project_lead = [(6, 0, emp_list)]

                            elif employee.user_id.user_role.id in [request.env.ref('api_auth_gateway.role_qc_technical').id,
                                                              request.env.ref('api_auth_gateway.role_qc_stem').id,
                                                              request.env.ref('api_auth_gateway.role_qc_non_stem').id]:
                                emp_list = ProjectRequest.project_qc_reviewer.ids
                                emp_list.append(employee.id)
                                ProjectRequest.sudo().project_qc_reviewer = [(6, 0, emp_list)]

                            elif employee.user_id.user_role.id in [
                                request.env.ref('api_auth_gateway.role_tasker_technical').id,
                                request.env.ref('api_auth_gateway.role_tasker_stem').id,
                                request.env.ref('api_auth_gateway.role_tasker_non_stem').id]:
                                emp_list = ProjectRequest.project_tasker.ids
                                emp_list.append(employee.id)
                                ProjectRequest.sudo().project_tasker = [(6, 0, emp_list)]

                        #
                    # if not new_user.employee_id:
                    #     employee_vals = {
                    #         'name': name,
                    #         'work_email': email,
                    #         'user_id': new_user.id,
                    #         'work_location_name': row.get('work_location_name', ''),
                    #         'department_id': int(row['department_id']) if row.get('department_id') else False,
                    #         'task_forge_pl_id': int(row['pl_id']) if row.get('pl_id') else False,
                    #         'task_forge_qr_id': int(row['qr_id']) if row.get('qr_id') else False,
                    #     }
                    #     if row.get('job_title'):
                    #         designation_id = request.env['hr.employee.designation'].sudo().search([('name', '=', row.get('job_title'))], limit=1)
                    #         if designation_id:
                    #             employee_vals['designation_id'] = designation_id.id
                    #
                    #     employee = Employee.create(employee_vals)
                        created.append({'id': employee.id, 'name': employee.name, 'email': employee.work_email})
                    plain_password = row.get('password') if row.get('password') else 'Ethara@123'
                    self._send_onboarding_email(new_user, plain_password)
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")

            return return_Response(
                message=f"File processed: {len(created)} created, {len(errors)} errors",
                status=200,
                data={'data': {'created': created, 'errors': errors}}
            )

        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v1/employees/update', methods=['PUT', 'PATCH'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    @validate_request({
        'employee_id': {'type': 'str', 'required': True},
    })
    def update_employee(self, **kwargs):
        try:
            jdata = kwargs.get('jdata')
            Employee = request.env['hr.employee'].sudo()
            employee = Employee.browse(int(jdata.get('employee_id')))

            if not employee.exists():
                return return_Response(message="Employee not found", status=404)
            employee_vals = {}

            if jdata.get('work_location_name'):
                employee_vals['work_location_name'] = jdata['work_location_name']

            if jdata.get('job_title'):
                employee_vals['designation_id'] = jdata['job_title']

            if jdata.get('department_id'):
                employee_vals['department_id'] = jdata['department_id']

            if jdata.get('pl_id'):
                employee_vals['task_forge_pl_id'] = jdata['pl_id']

            if jdata.get('qr_id'):
                employee_vals['task_forge_qr_id'] = jdata['qr_id']

            if employee_vals:
                employee.sudo().write(employee_vals)

            if jdata.get('role_id'):
                if employee.user_id:
                    employee.user_id.user_role = int(jdata.get('role_id'))

            try:
                request.env['kubera.notification'].sudo().create({
                    'title': 'Employee Updated',
                    'message': f'Employee "{employee.name}" has been updated.',
                    'user_id': request.env.user.id,
                    'priority': '1',
                    'res_model': 'hr.employee',
                    'res_id': employee.id,
                })
            except Exception:
                pass

            return return_Response(message="Employee updated successfully", status=200)
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v1/employees/offboard', methods=['POST'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    @validate_request({
        'employee_id': {'type': 'int', 'required': True},
        'action': {'type': 'string', 'required': True},
        'reason_id': {'type': 'int', 'required': True},
        'offboard_notes': {'type': 'string', 'required': True}
    })
    def offboard_employee(self, **kwargs):
        try:
            jdata = kwargs.get('jdata')
            Employee = request.env['hr.employee'].sudo()
            employee = Employee.browse(int(jdata.get('employee_id')))

            if not employee.exists():
                return return_Response(message="Employee not found", status=404)

            action = jdata.get('action', '').lower()

            if action == 'offboard':
                employee.sudo().write({
                    'offboarding_state': 'offboarded',
                    # 'active': False,
                    'task_forge_active': False,
                    'offboard_date': fields.Date.today(),
                    'reason_id': int(jdata.get('reason_id')),
                    'offboard_notes': jdata.get('offboard_notes')
                })

                try:
                    request.env['kubera.notification'].sudo().create({
                        'title': 'Employee Offboarded',
                        'message': f'Employee "{employee.name}" has been offboarded.',
                        'user_id': request.env.user.id,
                        'priority': '2',
                        'res_model': 'hr.employee',
                        'res_id': employee.id,
                    })
                except Exception:
                    pass

                return return_Response(message="Employee offboarded successfully", status=200)

            elif action == 'reactivate':
                if employee.offboarding_state == 'active':
                    return return_Response(message="Employee is already active", status=400)

                employee.write({
                    'offboarding_state': 'active',
                    'active': True,
                    'task_forge_active': True,
                    'offboard_date': False,
                })

                try:
                    request.env['kubera.notification'].sudo().create({
                        'title': 'Employee Reactivated',
                        'message': f'Employee "{employee.name}" has been reactivated.',
                        'user_id': request.env.user.id,
                        'priority': '1',
                        'res_model': 'hr.employee',
                        'res_id': employee.id,
                    })
                except Exception:
                    pass

                return return_Response(message="Employee reactivated successfully", status=200)
            else:
                return return_Response(message="Invalid action. Use: offboard or reactivate", status=400)
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/employees/detail_view', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    @validate_request({
        'employee_id': {'type': 'str', 'required': True},
    })
    def get_employee_detail_view(self, **kwargs):
        try:
            Employee = request.env['hr.employee'].sudo()
            employee = Employee.browse(int(kwargs['employee_id']))

            if not employee.exists():
                return return_Response(message="Employee not found", status=404)

            today = date.today()
            start_this_week = today - timedelta(days=today.weekday())
            start_last_week = start_this_week - timedelta(days=7)
            end_last_week = start_this_week - timedelta(days=1)
            Log = request.env['task.forge.log'].sudo()

            def get_prod_stats(target_emp_id, start_date, end_date):
                domain = [
                    ('employee_id', '=', target_emp_id),
                    ('date', '>=', start_date),
                    ('date', '<=', end_date)
                ]
                total = Log.search_count(domain)
                done = Log.search_count(domain + [('state', '=', 'completed')])

                percentage = (done / total * 100) if total > 0 else 0.0
                return round(percentage, 2), total

            last_7_days = [str(today - timedelta(days=i)) for i in range(6, -1, -1)]

            productivity_report = {
                "labels": last_7_days,
                'values': []
            }

            for day in last_7_days:
                domain = [
                    ('employee_id', '=', employee.id),
                    ('date', '=', day)
                ]

                total = Log.search_count(domain)
                done = Log.search_count(domain + [('state', '=', 'completed')])

                # Calculate percentage
                percentage = (done / total * 100) if total > 0 else 0.0

                productivity_report['values'].append(round(percentage, 2))
            # Calculations
            this_week_prod, _ = get_prod_stats(employee.id, start_this_week, today)
            last_week_prod, _ = get_prod_stats(employee.id, start_last_week, end_last_week)
            task_record = request.env['task.forge.log'].sudo().search([('employee_id', '=', employee.id), ('date', '=', today)])
            active_task = request.env['task.forge.log'].sudo().search([('employee_id', '=', employee.id), ('date', '=', today), ('state', '=', 'in_progress')], order='write_date desc', limit=1)
            punch_in_status, emp_session = self.get_employee_current_status(employee)
            abs_diff = round(this_week_prod - last_week_prod, 2)

            first_day_of_month = today.replace(day=1)
            days_passed = today.day

            attendances = request.env['hr.attendance'].sudo().search([
                ('employee_id', '=', employee.id),
                ('check_in', '>=', first_day_of_month),
                ('check_in', '<=', datetime.combine(today, time.max)), ('attendance_status', '=', 'present')])

            # total_present = len(attendances.mapped(lambda a: a.check_in.date() if a.check_in else ""))

            present_dates = []
            for att in attendances:
                if att.check_in and isinstance(att.check_in, datetime):
                    present_dates.append(att.check_in.date())
            total_present = len(set(present_dates))
            
            punch_in_times = []
            for att in attendances:
                if att.check_in and isinstance(att.check_in, datetime):
                    punch_in_times.append(att.check_in.hour * 60 + att.check_in.minute)
            # punch_in_times = []
            # for att in attendances:
            #     check_in_time = att.check_in.time()
            #     punch_in_times.append(check_in_time.hour * 60 + check_in_time.minute)

            if punch_in_times:
                avg_minutes = sum(punch_in_times) / len(punch_in_times)
                avg_punch_in = f"{int(avg_minutes // 60):02d}:{int(avg_minutes % 60):02d}"
            else:
                avg_punch_in = "00:00"

            # 3. Total Working Days (Excluding Weekends)
            # Calculate how many Mon-Fri have passed this month
            working_days_count = 0
            for day in range(1, days_passed + 1):
                if calendar.weekday(today.year, today.month, day) < 5:  # 0-4 are Mon-Fri
                    working_days_count += 1

            # 4. Leaves Taken
            leaves = request.env['hr.leave'].sudo().search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'validate'),
                ('date_from', '>=', first_day_of_month),
                ('date_to', '<=', datetime.combine(today, time.max))
            ])
            leave_taken = sum(leaves.mapped('number_of_days'))
            absent = working_days_count - total_present - leave_taken
            return return_Response(
                message="Employee details",
                status=200,
                data={'data': {
                    'id': employee.id if employee else 0,
                    'name': employee.name if employee and employee.name else "",
                    'email': employee.work_email if employee and employee.work_email else "",
                    'mobile': employee.work_phone if employee and employee.work_phone else "",
                    'job_title_id': employee.designation_id.id if employee.designation_id else 0,
                    'job_title': employee.designation_id.name if employee.designation_id and employee.designation_id.name else "",
                    'department_id': employee.department_id.id if employee.department_id else 0,
                    'department': employee.department_id.name if employee.department_id and employee.department_id.name else '',
                    'offboarding_state': employee.offboarding_state or "",
                    # 'is_offboarded': employee.is_offboarded,
                    'role_id': employee.user_id.user_role.id if employee.user_id.user_role else 0,
                    'role': employee.user_id.user_role.name if employee.user_id.user_role and employee.user_id.user_role.name else "",
                    'current_status': punch_in_status,
                    'emp_session': emp_session,
                    'abs_diff': abs_diff,
                    'this_week_productivity': round(this_week_prod, 2),
                    'last_week_productivity': round(last_week_prod, 2),
                    'current_task': active_task.name if active_task else "",
                    'task_today': len(task_record),
                    'today_task_record': [{'id': t.id, 'name': t.name, 'status': t.state or ""} for t in task_record],
                    'work_location': employee.work_location_name or "",
                    'offboard_date': employee.offboard_date.isoformat() if employee.offboard_date else "",
                    'pl_id': employee.task_forge_pl_id.id if employee.task_forge_pl_id else 0,
                    'pl_name': employee.task_forge_pl_id.name if employee.task_forge_pl_id and employee.task_forge_pl_id.name else "",
                    'qr_id': employee.task_forge_qr_id.id if employee.task_forge_qr_id else 0,
                    'qr_name': employee.task_forge_qr_id.name if employee.task_forge_qr_id and employee.task_forge_qr_id.name else "",
                    'active': employee.active or False,
                    'productivity_report': productivity_report,
                    'last_active': str(active_task.write_date) if active_task else "",
                    'this_month_total_present': total_present,
                    'this_month_total_working_days': working_days_count,
                    'leave_taken': leave_taken,
                    'absent': absent,
                    'avg_punch_in_time': avg_punch_in
                }}
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/employees_list', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def list_employees(self, **kwargs):
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)
            team_ids = employee._get_team_employee_ids()

            Employee = request.env['hr.employee'].sudo()

            domain = [('employee_id', 'in', team_ids)]

            if kwargs.get('active') == 'true':
                domain.append(('active', '=', True))

            elif kwargs.get('active') == 'false':
                in_team_ids = employee._get_inactive_team_employee_ids()
                domain = [('employee_id', 'in', in_team_ids)]
                domain.append(('task_forge_active', '=', False))

            search = kwargs.get('search')
            if search:
                domain += [('name', 'ilike', search)]

            if kwargs.get('offboarding_state'):
                domain.append(('offboarding_state', '=', kwargs['offboarding_state']))

            if kwargs.get('role'):
                domain.append(('user_id.user_role', '=', int(kwargs['role'])))

            if kwargs.get('pl_record') in [1, '1']:
                domain.append(('user_id.user_role', '=', request.env.ref('api_auth_gateway.role_pl_non_stem').id))
                
            if kwargs.get('qr_record') in [1, '1']:
                domain.append(('user_id.user_role', '=', request.env.ref('api_auth_gateway.role_qc_non_stem').id))

            if kwargs.get('department_id'):
                domain.append(('department_id', '=', int(kwargs['department_id'])))

            page = int(kwargs.get('page')) if kwargs.get('page') else 1
            limit = int(kwargs.get('limit')) if kwargs.get('limit') else 10
            offset = (page - 1) * limit
            total_count = request.env['hr.employee'].sudo().search_count(domain)
            if not kwargs.get('page'):
                limit = total_count
                offset = 0

            employees = Employee.search(domain, limit=limit, offset=offset)
            data = []
            today = date.today()
            start_this_week = today - timedelta(days=today.weekday())

            for emp in employees:
                Log = request.env['task.forge.log'].sudo()
                def get_prod_stats(target_emp_id, start_date, end_date):
                    domain = [
                        ('employee_id', '=', target_emp_id),
                        ('date', '>=', start_date),
                        ('date', '<=', end_date)
                    ]
                    total = Log.search_count(domain)
                    done = Log.search_count(domain + [('state', '=', 'completed')])

                    percentage = (done / total * 100) if total > 0 else 0.0
                    return round(percentage, 2), total

                this_week_prod, _ = get_prod_stats(emp.id, start_this_week, today)
                task_record = request.env['task.forge.log'].sudo().search_count([('employee_id', '=', emp.id), ('date', '=', today)])
                active_task = request.env['task.forge.log'].sudo().search([('employee_id', '=', emp.id), ('date', '=', today), ('state', '=', 'in_progress')], order='write_date desc', limit=1)
                current_status, _ =  self.get_employee_current_status(emp)
                vals = {
                    'id': emp.id if emp else 0,
                    'name': emp.name if emp and emp.name else "",
                    'email': emp.work_email if emp and emp.work_email else "",
                    'job_title_id': emp.designation_id.id if emp.designation_id else 0,
                    'job_title': emp.designation_id.name if emp.designation_id and emp.designation_id.name else "",
                    'department_id': emp.department_id.id if emp.department_id else 0,
                    'department': emp.department_id.name if emp.department_id and emp.department_id.name else '',
                    'offboarding_state': emp.offboarding_state or "",
                    # 'is_offboarded': emp.is_offboarded,
                    'role_id': emp.user_id.user_role.id if emp.user_id.user_role else 0,
                    'role': emp.user_id.user_role.name if emp.user_id.user_role and emp.user_id.user_role.name else "",
                    'current_status': current_status,
                    'current_task': active_task.name if active_task else "",
                    'productivity': this_week_prod,
                    'task_today': task_record,
                    'work_location': emp.work_location_name or "",
                    'offboard_date': emp.offboard_date.isoformat() if emp.offboard_date else "",
                    'pl_id': emp.task_forge_pl_id.id if emp.task_forge_pl_id else 0,
                    'pl_name': emp.task_forge_pl_id.name if emp.task_forge_pl_id and emp.task_forge_pl_id.name else "",
                    'qr_id': emp.task_forge_qr_id.id if emp.task_forge_qr_id else 0,
                    'qr_name': emp.task_forge_qr_id.name if emp.task_forge_qr_id and emp.task_forge_qr_id.name else "",
                    'active': emp.active or False,
                    'last_active': str(active_task.write_date) if active_task else ""
                }
                if kwargs.get('status'):
                    if kwargs.get('status') == 'offline' and current_status == 'Offline':
                        data.append(vals)
                    elif kwargs.get('status') == 'active' and current_status == 'Active':
                        data.append(vals)
                    elif kwargs.get('status') == 'idle' and current_status == 'Idle':
                        data.append(vals)
                else:
                    data.append(vals)

            active_projects = request.env['project.project'].sudo().search([('non_stemp_project_status', 'in', ['not_started', 'production'])])
            assigned_ids = set()
            for ap in active_projects:
                assigned_ids.update(ap.project_lead.ids)
                assigned_ids.update(ap.project_qc_reviewer.ids)
                assigned_ids.update(ap.project_tasker.ids)

            role_pl = [request.env.ref('api_auth_gateway.role_pl_non_stem').id,
                       request.env.ref('api_auth_gateway.role_pl_technical').id,
                       request.env.ref('api_auth_gateway.role_pl_stem').id]

            role_qr = [request.env.ref('api_auth_gateway.role_qc_technical').id,
                       request.env.ref('api_auth_gateway.role_qc_stem').id,
                       request.env.ref('api_auth_gateway.role_qc_non_stem').id]

            role_tk = [request.env.ref('api_auth_gateway.role_tasker_technical').id,
                       request.env.ref('api_auth_gateway.role_tasker_stem').id,
                       request.env.ref('api_auth_gateway.role_tasker_non_stem').id]

            pl_count = Employee.search_count([('id', 'in', team_ids), ('user_id.user_role', 'in', role_pl)])
            qr_count = Employee.search_count([('id', 'in', team_ids), ('user_id.user_role', 'in', role_qr)])
            tk_count = Employee.search_count([('id', 'in', team_ids), ('user_id.user_role', 'in', role_tk)])

            on_bench_count = len(set(team_ids) - assigned_ids)
            return return_Response(
                message=f"{len(data)} employees found",
                status=200,
                data={
                    'total_record_count': total_count,
                    'data': data,
                    "total": len(data),
                    "on_bench": on_bench_count,
                    'pl_count': pl_count,
                    'qr_count': qr_count,
                    'tasker_count': tk_count,
                    'offboarded_count': request.env['hr.employee'].sudo().search_count([('task_forge_active', '=', False)]),
                    'request_count': request.env['employee.allocation.request'].sudo().search_count([]),
                })
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/employees/<int:employee_id>', methods=['DELETE'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def delete_employee(self, employee_id, **kwargs):
        try:
            Employee = request.env['hr.employee'].sudo()
            employee = Employee.browse(employee_id)

            if not employee.exists():
                return return_Response(message="Employee not found", status=404)

            employee.write({'active': False})

            return return_Response(
                message="Employee archived successfully",
                status=200
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/get_user_role', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    def get_user_role(self, **kwargs):
        temp = []
        try:
            domain = [('project_type', '=', 'non-stem')]
            if kwargs.get('project_type'):
                domain = [('project_type', '=', kwargs.get('project_type'))]
            roles = request.env['api.role'].sudo().search(domain)
            for role in roles:
                temp.append({'id': role.id, 'name': role.name})
            return return_Response(
                message="success",
                status=200,
                data={'data': temp}
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/get_pl_employees_list', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def get_pl_employees_list(self, **kwargs):
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)

            Employee = request.env['hr.employee'].sudo()
            domain = [('user_id.user_role', 'in', [request.env.ref('api_auth_gateway.role_pl_technical').id,
                                       request.env.ref('api_auth_gateway.role_pl_stem').id,
                                       request.env.ref('api_auth_gateway.role_pl_non_stem').id]), ('task_forge_active', '=', True)]
            employees = Employee.search(domain)
            data = []
            for emp in employees:
                data.append({
                    'id': emp.id if emp else 0,
                    'name': emp.name if emp and emp.name else "",
                    'email': emp.work_email if emp and emp.work_email else "",
                    'role_id': emp.user_id.user_role.id if emp.user_id.user_role else 0,
                    'role': emp.user_id.user_role.name if emp.user_id.user_role and emp.user_id.user_role.name else ""
                })
            return return_Response(
                message=f"{len(data)} employees found",
                status=200,
                data={
                    'data': data,
                    "total": len(data)
                })
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/get_qr_employees_list', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    def get_qr_employees_list(self, **kwargs):
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)

            Employee = request.env['hr.employee'].sudo()
            domain = [('user_id.user_role', 'in', [request.env.ref('api_auth_gateway.role_qc_technical').id, request.env.ref('api_auth_gateway.role_qc_stem').id, request.env.ref('api_auth_gateway.role_qc_non_stem').id]), ('task_forge_active', '=', True)]
            employees = Employee.search(domain)
            data = []
            for emp in employees:
                data.append({
                    'id': emp.id if emp else 0,
                    'name': emp.name if emp and emp.name else "",
                    'email': emp.work_email if emp and emp.work_email else "",
                    'role_id': emp.user_id.user_role.id if emp.user_id.user_role else 0,
                    'role': emp.user_id.user_role.name if emp.user_id.user_role and emp.user_id.user_role.name else ""
                })
            return return_Response(
                message=f"{len(data)} employees found",
                status=200,
                data={
                    'data': data,
                    "total": len(data)
                })
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @validate_token
    @http.route('/api/v2/get_user_role_employee_onboarding', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    def get_user_role_employee_onboarding(self, **kwargs):
        temp = []
        try:
            user = request.env.user
            employee = user.employee_id
            if not employee:
                return return_Response(message="Employee profile not found", status=404)
            pl_emp_id = 0
            qr_emp_id = 0
            # if user.user_role.id == request.env.ref('api_auth_gateway.role_cto_technical').id:
            #     for role in [request.env.ref('api_auth_gateway.role_qc_stem'),
            #                  request.env.ref('api_auth_gateway.role_qc_technical'),
            #                  request.env.ref('api_auth_gateway.role_pl_stem'),
            #                  request.env.ref('api_auth_gateway.role_pl_technical'),
            #                  request.env.ref('api_auth_gateway.role_qc_non_stem'),
            #                  request.env.ref('api_auth_gateway.role_pl_non_stem'),
            #                  request.env.ref('api_auth_gateway.role_tasker_technical'),
            #                  request.env.ref('api_auth_gateway.role_tasker_stem'),
            #                  request.env.ref('api_auth_gateway.role_tasker_non_stem')]:
            #         temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})

            if user.user_role.id in [request.env.ref('api_auth_gateway.role_pl_non_stem').id]:
                pl_emp_id = employee.id
                for role in [request.env.ref('api_auth_gateway.role_qc_non_stem'), request.env.ref('api_auth_gateway.role_tasker_non_stem')]:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})

            elif user.user_role.id in [request.env.ref('api_auth_gateway.role_pl_technical').id]:
                pl_emp_id = employee.id
                for role in [request.env.ref('api_auth_gateway.role_qc_technical'), request.env.ref('api_auth_gateway.role_tasker_technical')]:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})

            elif user.user_role.id in [request.env.ref('api_auth_gateway.role_pl_stem').id]:
                pl_emp_id = employee.id
                for role in [request.env.ref('api_auth_gateway.role_qc_stem'), request.env.ref('api_auth_gateway.role_tasker_stem')]:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})

            elif user.user_role.id in [request.env.ref('api_auth_gateway.role_qc_stem').id]:
                pl_emp_id = employee.task_forge_pl_id.id
                qr_emp_id = employee.id
                for role in [request.env.ref('api_auth_gateway.role_tasker_stem')]:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})

            elif user.user_role.id in [request.env.ref('api_auth_gateway.role_qc_non_stem').id]:
                pl_emp_id = employee.task_forge_pl_id.id
                qr_emp_id = employee.id
                for role in [request.env.ref('api_auth_gateway.role_tasker_non_stem')]:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})

            elif user.user_role.id in [request.env.ref('api_auth_gateway.role_qc_technical').id]:
                pl_emp_id = employee.task_forge_pl_id.id
                qr_emp_id = employee.id
                for role in [request.env.ref('api_auth_gateway.role_tasker_technical')]:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})
            else:
                domain = []
                if kwargs.get('project_type'):
                    domain = [('project_type', '=', kwargs.get('project_type'))]
                roles = request.env['api.role'].sudo().search(domain)
                for role in roles:
                    temp.append({'id': role.id, 'name': role.name,'project_type': role.project_type})
            return return_Response(
                message="success",
                status=200,
                data={'data': temp, 'pl_emp_id': pl_emp_id, 'qr_emp_id': qr_emp_id}
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)

    @http.route('/api/v2/get_offboarding_reasons', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    def get_offboarding_reasons(self, **kwargs):
        temp = []
        try:
            offboarding_reasons = request.env['hr.employee.offboarding.reasons'].sudo().search([])
            for role in offboarding_reasons:
                temp.append({'id': role.id, 'reason': role.reason})
            return return_Response(
                message="success",
                status=200,
                data={'data': temp}
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)

    def get_employee_current_status(self, emp):
        TaskLog = request.env['task.forge.log'].sudo()
        tasks = TaskLog.search_count([
            ('employee_id', '=', emp.id),
            ('state', '=', 'in_progress'),
        ])
        today = date.today()
        Attendance = request.env['hr.attendance'].sudo()
        attendance = Attendance.search([
            ('employee_id', '=', emp.id),
            ('check_in', '>=', datetime.combine(today, datetime.min.time())),
            ('check_in', '<', datetime.combine(today, datetime.max.time())),
            ('attendance_status', '=', 'present')
        ], limit=1)

        duration_display = "00:00"
        if attendance and attendance.check_in and attendance.check_out:
            diff = attendance.check_out - attendance.check_in
            total_seconds = int(diff.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            duration_display = f"{hours:02d}:{minutes:02d}"

        elif attendance and attendance.check_in and not attendance.check_out:
            diff = datetime.now() - attendance.check_in
            total_seconds = int(diff.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            duration_display = f"{hours:02d}:{minutes:02d}"

        current_status = ""
        if not attendance:
            current_status = "Offline"
        elif tasks:
            current_status = "Active"
        else:
            current_status = "Idle"
        return current_status, duration_display


    @http.route('/api/v2/get_on_bench_employees', methods=['GET'], type='http', auth='none', csrf=False, cors='*')
    @validate_token
    @validate_request({})
    def get_on_bench_employees(self, **kwargs):
        temp = []
        try:
            active_projects = request.env['project.project'].sudo().search([('non_stemp_project_status', 'in', ['not_started', 'production'])])
            assign_employee = []
            for ap in active_projects:
                assign_employee.extend(ap.project_lead.ids)
                assign_employee.extend(ap.project_qc_reviewer.ids)
                assign_employee.extend(ap.project_tasker.ids)
            domain = [('id', 'not in', assign_employee)]

            if kwargs.get('role_id'):
                domain.append(('user_id.user_role', '=', int(kwargs.get('role_id'))))

            employees = self.env['hr.employee'].sudo().search(domain)
            temp = [{
                'id': emp.id or 0,
                'name': emp.name or "",
                'email': emp.work_email or '',
                'role': emp._get_task_forge_role(),
                'pl_id': emp.task_forge_pl_id.id if emp.task_forge_pl_id else "",
                'pl_name': emp.task_forge_pl_id.name if emp.task_forge_pl_id and emp.task_forge_pl_id.name else "",
                'qr_id': emp.task_forge_qr_id.id if emp.task_forge_qr_id else "",
                'qr_name': emp.task_forge_qr_id.name if emp.task_forge_qr_id and emp.task_forge_qr_id.name else ""
            } for emp in employees]
            return return_Response(
                message="success",
                status=200,
                data={'data': temp}
            )
        except Exception as e:
            return return_Response(message=str(e), status=400)


