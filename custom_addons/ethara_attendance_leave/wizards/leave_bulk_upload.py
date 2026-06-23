import base64
import io
import logging

from odoo import models, fields, _
from odoo.exceptions import UserError

from odoo.addons.ethara_attendance_leave.models.hr_employee import (
    ETHARA_BALANCE_FIELDS,
)

_logger = logging.getLogger(__name__)

# (ethara_leave_code, column header) — mirrors the Leave Management → Employees
# grid columns. One column per leave type; one row per employee.
GRID_COLUMNS = [
    ('sl', 'Sick'),
    ('cl', 'Casual'),
    ('el', 'Earned'),
    ('lop', 'LOP'),
    ('marriage', 'Marriage'),
    ('maternity', 'Maternity'),
    ('bereavement', 'Bereavement'),
    ('comp_off', 'Comp-Off'),
    ('wfh', 'WFH'),
    ('menstrual', 'Menstrual'),
    ('restricted_holiday', 'Restricted'),
]
EMPLOYEE_HEADER = 'Employee (email or ID)'
TEMPLATE_HEADERS = [EMPLOYEE_HEADER] + [label for _code, label in GRID_COLUMNS]


class LeaveAllocationBulkUpload(models.TransientModel):
    _name = 'hr.leave.allocation.bulk.upload'
    _description = 'Bulk Upload Leave Allocations (XLSX)'

    upload_file = fields.Binary(string='Excel File (.xlsx)')
    upload_filename = fields.Char(string='Filename')
    result_summary = fields.Text(string='Result', readonly=True)

    # ------------------------------------------------------------------ #
    #  Template download (empty grid — headers only)                      #
    # ------------------------------------------------------------------ #
    def action_download_template(self):
        """Generate an empty .xlsx in the same grid layout as the employee list
        (Employee + one column per leave type) and return a download."""
        self.ensure_one()
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leave Allocations'
        ws.append(TEMPLATE_HEADERS)

        for idx, _h in enumerate(TEMPLATE_HEADERS, start=1):
            ws.cell(row=1, column=idx).font = Font(bold=True)
            width = 26 if idx == 1 else 12
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'leave_allocation_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(buffer.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    # ------------------------------------------------------------------ #
    #  Import                                                             #
    # ------------------------------------------------------------------ #
    def action_import(self):
        """Parse the grid .xlsx: one row per employee, one column per leave type.
        Each non-empty cell sets that employee's balance for that type (reusing
        the same grid logic + validations as the editable list)."""
        self.ensure_one()
        if not self.upload_file:
            raise UserError(_('Please choose an .xlsx file to upload first.'))

        import openpyxl
        try:
            data = base64.b64decode(self.upload_file)
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as exc:  # noqa: BLE001 - surface a clean message to HR
            raise UserError(_('Could not read the file as .xlsx: %s') % exc)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise UserError(_('The file has no data rows (only a header was found).'))

        col_code = self._map_columns(rows[0])
        if not col_code:
            raise UserError(_(
                'No leave-type columns recognised. Download the template to get '
                'the correct column headers.'))

        updated, errors = 0, []
        for line_no, row in enumerate(rows[1:], start=2):
            if row is None or all(cell in (None, '') for cell in row):
                continue  # blank line
            try:
                with self.env.cr.savepoint():
                    employee = self._resolve_employee(row[0] if row else None)
                    vals = {}
                    for idx, code in col_code.items():
                        if idx >= len(row):
                            continue
                        cell = row[idx]
                        if cell in (None, ''):
                            continue  # leave this type unchanged
                        vals[ETHARA_BALANCE_FIELDS[code]] = self._coerce_days(cell)
                    if vals:
                        employee.write(vals)
                        updated += 1
            except Exception as exc:  # noqa: BLE001 - per-row, keep importing
                errors.append('Row %d: %s' % (line_no, exc))

        summary = _('%d employee row(s) applied.') % updated
        if errors:
            summary += '\n\n' + _('%d row(s) skipped:') % len(errors) + '\n' + '\n'.join(errors)
        self.result_summary = summary

        # All good -> toast "Done" and reload the list so new balances show.
        if not errors:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Done'),
                    'message': _('%d employee row(s) applied successfully.') % updated,
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                },
            }

        # Some rows failed -> keep the wizard open so HR can read which ones.
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
            'name': _('Bulk Upload — Result'),
        }

    # ------------------------------------------------------------------ #
    #  Helpers                                                            #
    # ------------------------------------------------------------------ #
    def _map_columns(self, header):
        """Map each spreadsheet column index -> ethara_leave_code.

        Column 0 is the employee column and is skipped. Other columns are matched
        first against the template labels, then against leave type name/code."""
        label_to_code = {label.lower(): code for code, label in GRID_COLUMNS}
        col_code = {}
        for idx, cell in enumerate(header or []):
            if idx == 0 or cell in (None, ''):
                continue
            key = str(cell).strip().lower()
            code = label_to_code.get(key)
            if not code:
                lt = self.env['hr.leave.type'].search([
                    '|', ('name', '=ilike', str(cell).strip()),
                    ('ethara_leave_code', '=', key),
                ], limit=1)
                code = lt.ethara_leave_code or None
            if code in ETHARA_BALANCE_FIELDS:
                col_code[idx] = code
        return col_code

    def _resolve_employee(self, ref):
        if ref in (None, ''):
            raise UserError(_('Employee is empty.'))
        Employee = self.env['hr.employee']
        if isinstance(ref, (int, float)) or (isinstance(ref, str) and ref.strip().isdigit()):
            emp = Employee.browse(int(ref))
            if emp.exists():
                return emp
        ref = str(ref).strip()
        emp = Employee.search(
            ['|', ('work_email', '=ilike', ref), ('user_id.login', '=ilike', ref)], limit=1)
        if not emp:
            emp = Employee.search([('name', '=ilike', ref)], limit=1)
        if not emp:
            raise UserError(_('No employee found for "%s".') % ref)
        return emp

    @staticmethod
    def _coerce_days(value):
        try:
            days = float(value)
        except (TypeError, ValueError):
            raise UserError(_('"%s" is not a number.') % value)
        if days < 0:
            raise UserError(_('Days cannot be negative (%s).') % value)
        return days
