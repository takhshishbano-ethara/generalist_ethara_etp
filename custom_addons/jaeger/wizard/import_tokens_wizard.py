import base64
import csv
import hashlib
import io
import logging

from odoo import fields, models
from odoo.exceptions import UserError

from ..models.credential_manager import encrypt_value

_logger = logging.getLogger(__name__)

_VALID_PREFIXES = ("ghp_", "gho_", "github_pat_")
_MIN_TOKEN_LENGTH = 30
_HEADER_NAMES = {"token", "pat", "github_token"}
_BATCH_SIZE = 500


class JaegerImportTokensWizard(models.TransientModel):
    _name = "jaeger.import.tokens.wizard"
    _description = "Import GitHub Tokens"

    token_file = fields.Binary(string="Token File", required=True)
    token_filename = fields.Char()
    result_message = fields.Text(readonly=True)

    def action_import(self):
        self.ensure_one()
        if not self.token_file:
            raise UserError("Please upload a file.")

        raw = base64.b64decode(self.token_file)
        filename = (self.token_filename or "").lower()

        if filename.endswith(".xlsx"):
            raw_tokens = self._parse_xlsx(raw)
        elif filename.endswith(".csv"):
            raw_tokens = self._parse_csv(raw)
        else:
            raise UserError("Unsupported file format. Upload .xlsx or .csv")

        if not raw_tokens:
            raise UserError("No tokens found in file.")

        valid = []
        invalid = []
        for tok in raw_tokens:
            tok = tok.strip()
            if not tok:
                continue
            if any(tok.startswith(p) for p in _VALID_PREFIXES) and len(tok) >= _MIN_TOKEN_LENGTH:
                valid.append(tok)
            else:
                invalid.append(tok[:20] + "...")

        if not valid:
            raise UserError(f"No valid tokens found. {len(invalid)} rejected (missing ghp_/gho_/github_pat_ prefix).")

        hashes = {hashlib.sha256(t.encode()).hexdigest(): t for t in valid}

        hash_list = list(hashes.keys())
        self.env.cr.execute("""
            SELECT token_hash FROM jaeger_github_token
            WHERE token_hash IN (SELECT unnest(%s::text[]))
        """, (hash_list,))
        existing = {r[0] for r in self.env.cr.fetchall()}

        new_hashes = [h for h in hash_list if h not in existing]
        if not new_hashes:
            self.result_message = f"All {len(valid)} tokens already exist in the pool."
            return self._show_result()

        ICP = self.env["ir.config_parameter"].sudo()
        existing_count = self.env["jaeger.github.token"].search_count([])

        vals_list = []
        for i, h in enumerate(new_hashes):
            raw_tok = hashes[h]
            seq_num = existing_count + i + 1
            vals_list.append({
                "name": f"Token {seq_num:04d}",
                "token": encrypt_value(ICP, raw_tok),
                "token_hash": h,
                "state": "draft",
            })

        created = 0
        for start in range(0, len(vals_list), _BATCH_SIZE):
            batch = vals_list[start:start + _BATCH_SIZE]
            self.env["jaeger.github.token"].create(batch)
            created += len(batch)

        skipped = len(valid) - len(new_hashes)
        msg_parts = [f"Import complete: {created} tokens imported"]
        if skipped:
            msg_parts.append(f"{skipped} duplicates skipped")
        if invalid:
            msg_parts.append(f"{len(invalid)} invalid tokens rejected")
        msg_parts.append("Tokens are in 'draft' state pending health check verification")

        self.result_message = ". ".join(msg_parts) + "."
        return self._show_result()

    def _show_result(self):
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    @staticmethod
    def _parse_xlsx(raw_bytes):
        tokens = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=io.BytesIO(raw_bytes), read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                val = row[0]
                if val is not None:
                    val = str(val).strip()
                    if val:
                        tokens.append(val)
            wb.close()
        except Exception as exc:
            raise UserError(f"Failed to parse Excel file: {exc}") from exc

        if tokens and tokens[0].lower() in _HEADER_NAMES:
            tokens = tokens[1:]
        return tokens

    @staticmethod
    def _parse_csv(raw_bytes):
        tokens = []
        try:
            text = raw_bytes.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            for row in reader:
                if row:
                    val = row[0].strip()
                    if val:
                        tokens.append(val)
        except Exception as exc:
            raise UserError(f"Failed to parse CSV file: {exc}") from exc

        if tokens and tokens[0].lower() in _HEADER_NAMES:
            tokens = tokens[1:]
        return tokens
